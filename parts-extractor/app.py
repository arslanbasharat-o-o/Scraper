from flask import Flask, request, jsonify, render_template, send_file, url_for, Response, session as flask_session, redirect
import atexit
import csv
import datetime
import html
import io
import ipaddress
import json
import logging
import os
import re
import requests
import shutil
import signal
import socket
import subprocess
import sys
import time
from logging.handlers import RotatingFileHandler
from pathlib import Path
from urllib.parse import urljoin, urlparse
from dataclasses import asdict
from typing import List, Dict, Tuple
from concurrent.futures import ThreadPoolExecutor, as_completed
from copy import copy
from functools import wraps
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import mimetypes
import pytz
import threading
import unicodedata
import uuid
from xml.etree import ElementTree as ET
from database import db_manager
from automation_service import discover_category_targets
from auth import init_auth, require_login, require_role, validate_credentials, is_auth_configured
from flask_login import current_user, login_user, logout_user

AUTOMATION_CHECKPOINT_ITEM_LIMIT = 100
AUTOMATION_LIVE_DETAIL_ITEM_LIMIT = 500


def load_local_env_file(path: str = ".env") -> None:
    """Load simple KEY=VALUE pairs from a local .env without overriding the shell."""
    if os.getenv("PYTEST_CURRENT_TEST") or "pytest" in sys.modules:
        return
    if not os.path.exists(path):
        return
    try:
        with open(path, "r", encoding="utf-8") as handle:
            for raw_line in handle:
                line = raw_line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                key, value = line.split("=", 1)
                key = key.strip().lstrip("\ufeff")
                if not key or key in os.environ:
                    continue
                value = value.strip().strip('"').strip("'")
                os.environ[key] = value
    except OSError:
        pass


load_local_env_file()

# Import scraper engines (separated for maintainability)
from scrapers.scraper_engine import (
    Item, build_session, scrape_url,
    parse_price_number,
    enrich_item_details as enrich_standard_item_details
)

# Import XCellParts specialized scraper
from scrapers import SCRAPER_CONFIG, detect_scraper_key, xcell_scraper_engine

# Import TXParts specialized scraper
from scrapers import txparts_scraper_engine

# Import Parts4Cells specialized scraper
from scrapers import parts4cells_scraper_engine

# Import PhoneLCDParts and GadgetFix specialized scrapers
from scrapers import phonelcdparts_scraper_engine, gadgetfix_scraper_engine
from scrapers.browser_fetcher import browser_fetch_mode

SCRAPER_MODULES = {
    'standard': None,
    'mobilesentrix_canada': None,
    'xcell': xcell_scraper_engine,
    'txparts': txparts_scraper_engine,
    'parts4cells': parts4cells_scraper_engine,
    'phonelcdparts': phonelcdparts_scraper_engine,
    'gadgetfix': gadgetfix_scraper_engine,
}

app = Flask(__name__)
app.logger.setLevel(logging.INFO)
APP_ROOT = Path(__file__).resolve().parent
app.config['MAX_CONTENT_LENGTH'] = max(
    1,
    int(os.getenv('MAX_REQUEST_SIZE_MB', '25') or 25),
) * 1024 * 1024

# Secret key: required for Flask sessions. In production, always set SECRET_KEY env var.
_secret_key = os.getenv('SECRET_KEY', '').strip()
if not _secret_key:
    import secrets as _secrets
    _secret_key = _secrets.token_hex(32)
    app.logger.warning(
        "[security] SECRET_KEY not set in environment â€” using a random ephemeral key. "
        "Sessions will not survive server restarts. Set SECRET_KEY in .env for production."
    )
app.secret_key = _secret_key
del _secret_key

# Initialise Flask-Login auth system. Auth is backward-compatible:
# when AUTH_PASSWORD / AUTH_PASSWORD_HASH are not configured, all endpoints
# remain open (existing behavior preserved).
init_auth(app)


def configure_app_file_logging() -> None:
    log_path = APP_ROOT / 'server.log'
    log_path_text = str(log_path)
    root_logger = logging.getLogger()
    existing_paths = {
        getattr(handler, 'baseFilename', '')
        for handler in root_logger.handlers
        if isinstance(handler, RotatingFileHandler)
    }
    if log_path_text not in existing_paths:
        handler = RotatingFileHandler(log_path, maxBytes=2_000_000, backupCount=5, encoding='utf-8')
        handler.setLevel(logging.INFO)
        handler.setFormatter(logging.Formatter(
            '%(asctime)s %(levelname)s [%(name)s] %(message)s'
        ))
        root_logger.addHandler(handler)
    if root_logger.level > logging.INFO:
        root_logger.setLevel(logging.INFO)
    app.logger.propagate = True
    logging.getLogger('werkzeug').setLevel(logging.INFO)
    logging.getLogger('werkzeug').propagate = True


configure_app_file_logging()

PROXIED_IMAGE_TTL_SECONDS = 15 * 60
MAX_PROXIED_IMAGE_CACHE_ITEMS = 512
MAX_PROXIED_IMAGE_BYTES = max(
    1,
    int(os.getenv('MAX_PROXIED_IMAGE_MB', '15') or 15),
) * 1024 * 1024
SUPPLIER_REMOTE_HOSTS = tuple(sorted({
    str(domain or '').strip().lower().lstrip('.')
    for config in SCRAPER_CONFIG.values()
    for domain in config.get('domains', ())
    if str(domain or '').strip()
}))
DESTRUCTIVE_CONFIRMATION_VALUE = 'permanently-delete'
PROXIED_IMAGE_CACHE: Dict[str, Dict[str, object]] = {}
PROXIED_IMAGE_CACHE_LOCK = threading.Lock()
AUTO_DETAIL_SCAN_MAX_ITEMS = 20
AUTOMATION_POLL_INTERVAL_SECONDS = 45
AUTOMATION_SCHEDULER_LOCK = threading.Lock()
AUTOMATION_SCHEDULER_STARTED = False
AUTOMATION_SCHEDULER_THREAD: threading.Thread | None = None
AUTOMATION_STOP_EVENT = threading.Event()
AUTOMATION_ACTIVE_JOBS = set()
AUTOMATION_ACTIVE_JOBS_LOCK = threading.Lock()
MENU_MAP_JOBS: Dict[str, Dict[str, object]] = {}
MENU_MAP_JOBS_LOCK = threading.Lock()
SHUTDOWN_HOOKS_REGISTERED = False


class AutomationRunPaused(RuntimeError):
    """Raised by progress callbacks when a running automation run is paused."""

MENU_MAP_SITES = {
    'xcellparts': {
        'name': 'XCell Parts',
        'url': 'https://xcellparts.com/',
        'module': 'scrapers.menu_map.xcellparts',
        'behavior': 'Astra/WooCommerce desktop menu. Parent items open by hover; hierarchy lives in category-column blocks.',
    },
    'parts4cells': {
        'name': 'Parts4Cells',
        'url': 'https://parts4cells.com/',
        'module': 'scrapers.menu_map.parts4cells',
        'behavior': 'Magento desktop menu. Parent items open by click; sub-child groups are megamenu blocks and child links use sub-category links.',
    },
    'phonelcdparts': {
        'name': 'Phone LCD Parts',
        'url': 'https://www.phonelcdparts.com/',
        'module': 'scrapers.menu_map.phonelcdparts',
        'behavior': 'Ninja/Magezon menu. Parent items use hover/click; left tab titles select sub-child categories and right tab panels contain child links.',
    },
    'mobilesentrix': {
        'name': 'MobileSentrix',
        'url': 'https://www.mobilesentrix.com/',
        'module': 'scrapers.menu_map.mobilesentrix',
        'behavior': 'Large custom desktop menu. Parent items open by click; groups are sview blocks.',
    },
    'mobilesentrix_canada': {
        'name': 'MobileSentrix Canada',
        'url': 'https://www.mobilesentrix.ca/',
        'module': 'scrapers.menu_map.mobilesentrix_canada',
        'behavior': 'Large custom desktop menu. Parent items open by click; groups are sview blocks.',
    },
    'txparts': {
        'name': 'TXParts',
        'url': 'https://txparts.com/',
        'module': 'scrapers.menu_map.txparts',
        'behavior': 'Shop/category navigation. The scraper extracts category links from desktop navigation and shop/category menus.',
    },
    'txparts_canada': {
        'name': 'TXParts Canada',
        'url': 'https://txpartscanada.ca/',
        'module': 'scrapers.menu_map.txparts_canada',
        'behavior': 'Shop/category navigation. The scraper extracts category links from desktop navigation and shop/category menus.',
        'timeout': 180000,
    },
    'gadgetfix': {
        'name': 'GadgetFix',
        'url': 'https://gadgetfix.com/',
        'module': 'scrapers.menu_map.gadgetfix',
        'behavior': 'Category-link navigation. The scraper extracts GadgetFix category URLs and groups model links under their visible menu sections.',
    },
}


def asset_url(filename: str) -> str:
    """Generate cache-busted static asset URLs so frontend fixes load immediately."""
    version = None
    try:
        static_path = os.path.join(app.static_folder or '', filename)
        if static_path and os.path.exists(static_path):
            version = int(os.path.getmtime(static_path))
    except Exception:
        version = None
    return url_for('static', filename=filename, v=version)


@app.context_processor
def inject_asset_url():
    return {
        'asset_url': asset_url,
        'browser_rendering_enabled': True,
    }


@app.before_request
def ensure_background_services():
    ensure_automation_scheduler_started()


@app.before_request
def reject_cross_origin_state_changes():
    if request.method not in {'POST', 'PUT', 'PATCH', 'DELETE'}:
        return None
    origin = str(request.headers.get('Origin') or '').strip()
    if not origin:
        return None
    parsed_origin = urlparse(origin)
    allowed_origins = {
        value.strip().rstrip('/')
        for value in str(os.getenv('CORS_ALLOWED_ORIGINS') or '').split(',')
        if value.strip()
    }
    if parsed_origin.netloc == request.host or origin.rstrip('/') in allowed_origins:
        return None
    return jsonify({'error': 'Cross-origin state-changing requests are not allowed.'}), 403


@app.after_request
def apply_security_headers(response):
    response.headers.setdefault('X-Content-Type-Options', 'nosniff')
    response.headers.setdefault('X-Frame-Options', 'SAMEORIGIN')
    response.headers.setdefault('Referrer-Policy', 'same-origin')
    response.headers.setdefault('Permissions-Policy', 'camera=(), microphone=(), geolocation=()')
    if request.path.startswith('/api/'):
        response.headers.setdefault('Cache-Control', 'no-store')
    return response


def require_destructive_confirmation(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if request.headers.get('X-Confirm-Destructive') != DESTRUCTIVE_CONFIRMATION_VALUE:
            return jsonify({
                'error': 'Explicit destructive-action confirmation is required.',
            }), 428
        return view(*args, **kwargs)
    return wrapped


def get_public_base_url() -> str:
    """Resolve the canonical external base URL when available, otherwise use the active request."""
    configured_url = os.getenv('PUBLIC_BASE_URL') or os.getenv('SITE_URL')
    if configured_url:
        return configured_url.rstrip('/')

    return request.url_root.rstrip('/')


def get_template_lastmod(template_name: str) -> str:
    template_path = os.path.join(app.template_folder or '', template_name)
    try:
        modified_at = datetime.datetime.fromtimestamp(os.path.getmtime(template_path), tz=datetime.timezone.utc)
    except OSError:
        modified_at = datetime.datetime.now(datetime.timezone.utc)
    return modified_at.replace(microsecond=0).isoformat().replace('+00:00', 'Z')


def build_public_site_pages() -> List[Dict[str, str]]:
    return [
        {
            'path': '/',
            'endpoint': 'index',
            'changefreq': 'daily',
            'priority': '1.0',
            'lastmod': get_template_lastmod('index.html'),
        },
        {
            'path': '/history',
            'endpoint': 'history',
            'changefreq': 'daily',
            'priority': '0.8',
            'lastmod': get_template_lastmod('history.html'),
        },
        {
            'path': '/automation',
            'endpoint': 'automation',
            'changefreq': 'daily',
            'priority': '0.8',
            'lastmod': get_template_lastmod('automation.html'),
        },

        {
            'path': '/menu-map',
            'endpoint': 'menu_map',
            'changefreq': 'daily',
            'priority': '0.8',
            'lastmod': get_template_lastmod('menu_map.html'),
        },
    ]

@app.teardown_appcontext
def close_db_connection(_exception=None):
    """Release SQLite connections after each request to avoid stale thread-local snapshots."""
    db_manager.close_connection()

def normalize_compare_text(value) -> str:
    return re.sub(r'\s+', ' ', str(value or '')).strip()


_SEMANTIC_PUNCTUATION_TRANSLATION = str.maketrans({
    '\u00a0': ' ',
    '\u2010': '-',
    '\u2011': '-',
    '\u2012': '-',
    '\u2013': '-',
    '\u2014': '-',
    '\u2015': '-',
    '\u2212': '-',
    '\u2018': "'",
    '\u2019': "'",
    '\u201a': "'",
    '\u201b': "'",
    '\u2032': "'",
    '\u201c': '"',
    '\u201d': '"',
    '\u201e': '"',
    '\u201f': '"',
    '\u2033': '"',
})


def normalize_semantic_compare_text(value) -> str:
    """Normalize presentation-only text differences without changing display values."""
    text = html.unescape(str(value or '')).translate(_SEMANTIC_PUNCTUATION_TRANSLATION)
    text = unicodedata.normalize('NFKC', text).casefold()
    text = ''.join(
        ' ' if unicodedata.category(character) in {'Cc', 'Cf', 'Zl', 'Zp'} else character
        for character in text
    )
    text = text.replace("''", '"')
    return re.sub(r'\s+', ' ', text).strip()


def normalize_identifier_compare_text(value) -> str:
    """Normalize product identifiers without merging different SKUs/models."""
    text = normalize_semantic_compare_text(value)
    if not text:
        return ''
    return re.sub(r'[^a-z0-9]+', '', text)


def normalize_compare_url(value) -> str:
    """Return a stable URL identity without tracking queries or trailing-slash noise."""
    text = normalize_compare_text(html.unescape(str(value or '')))
    if not text:
        return ''
    parsed = urlparse(text)
    if not parsed.scheme or not parsed.netloc:
        return text.rstrip('/')

    host = (parsed.hostname or '').lower()
    if not host:
        return text.rstrip('/')
    port = parsed.port
    default_port = (parsed.scheme.lower() == 'http' and port == 80) or (
        parsed.scheme.lower() == 'https' and port == 443
    )
    netloc = host if not port or default_port else f'{host}:{port}'
    path = re.sub(r'/+', '/', parsed.path or '/')
    path = path.rstrip('/') or '/'
    return parsed._replace(
        scheme=parsed.scheme.lower(),
        netloc=netloc,
        path=path,
        params='',
        query='',
        fragment='',
    ).geturl()


def get_item_extra_value(item_dict: Dict[str, object], extra: Dict[str, object], *keys) -> str:
    for key in keys:
        value = item_dict.get(key)
        if value not in (None, ''):
            return normalize_compare_text(value)
        value = extra.get(key)
        if value not in (None, ''):
            return normalize_compare_text(value)
    return ''


def normalize_stock_status_for_compare(value) -> str:
    """Collapse stock quantity variants into stable compare values."""
    text = normalize_compare_text(value).lower()
    if not text:
        return ''
    if 'outofstock' in text or 'out of stock' in text:
        return 'out of stock'
    if re.search(r'\b\d+\s+in stock\b', text) or 'in stock' in text:
        return 'in stock'
    if 'backorder' in text or 'back order' in text:
        return 'backorder'
    if 'preorder' in text or 'pre-order' in text:
        return 'preorder'
    return text


def has_specific_stock_detail(value) -> bool:
    """Detect whether a stock label carries quantity/low-stock detail instead of a generic state."""
    text = normalize_compare_text(value).lower()
    if not text:
        return False
    return bool(
        re.search(r'\b\d+\s+in stock\b', text)
        or 'left in stock' in text
        or re.search(r'\bonly\b.*\bin stock\b', text)
    )


def needs_specific_stock_refresh(items) -> bool:
    """Detect when a fast listing scrape only has generic stock values."""
    for item in items:
        item_dict = asdict(item) if hasattr(item, '__dict__') else dict(item or {})
        extra = item_dict.get('extra') if isinstance(item_dict.get('extra'), dict) else {}
        stock_status = normalize_compare_text(item_dict.get('stock_status') or extra.get('stock_status'))
        if not stock_status:
            return True
        if normalize_stock_status_for_compare(stock_status) == 'in stock' and not has_specific_stock_detail(stock_status):
            return True
    return False


def get_scraper_for_url(url: str):
    """Determine which scraper engine to use based on domain."""
    scraper_key = detect_scraper_key(url)
    return scraper_key, SCRAPER_MODULES.get(scraper_key)


def format_category_label_from_url(url: str) -> str:
    """Turn a category URL tail into a readable model/category label."""
    parsed = urlparse(str(url or '').strip())
    path = parsed.path.rstrip('/')
    if path.endswith('.html'):
        path = path[:-5]
    segments = [segment for segment in path.split('/') if segment]
    if not segments:
        return ''
    label = segments[-1]
    label = re.sub(r'[-_]+', ' ', label)
    label = re.sub(r'\s+', ' ', label).strip()
    return label.title()


def annotate_items_with_target(items, target_url: str, target_label: str = '', automation_job: Dict | None = None):
    """Attach the originating category URL/label so later comparisons can group by model."""
    normalized_url = str(target_url or '').strip()
    normalized_label = str(target_label or '').strip() or format_category_label_from_url(normalized_url)
    for item in items or []:
        extra = getattr(item, 'extra', None)
        if not isinstance(extra, dict):
            extra = {}
            try:
                setattr(item, 'extra', extra)
            except Exception:
                continue
        extra['target_url'] = normalized_url
        extra['target_label'] = normalized_label
        if normalized_label:
            extra['model_label'] = normalized_label
        if automation_job:
            extra['automation_job_id'] = automation_job.get('id')
            extra['automation_job_name'] = automation_job.get('name')
    return items


def apply_enriched_item_data(item, enriched_data: Dict[str, object]):
    """Copy dataclass-exported enriched fields into another item object."""
    for key, value in (enriched_data or {}).items():
        if key == 'extra':
            current_extra = getattr(item, 'extra', None)
            if isinstance(current_extra, dict) and isinstance(value, dict):
                current_extra.update(value)
            elif hasattr(item, 'extra') and isinstance(value, dict):
                setattr(item, 'extra', dict(value))
            continue
        if not hasattr(item, key):
            continue
        if value is None:
            continue
        if isinstance(value, str) and not value.strip():
            continue
        setattr(item, key, value)
    return item


def hydrate_items_from_previous_history(items, previous_history: Dict | None) -> int:
    """Reuse saved metadata to avoid false diffs when deep detail scan is disabled."""
    if not previous_history:
        return 0

    previous_by_url = {}
    for previous_item in previous_history.get('items', []):
        url = normalize_compare_text(previous_item.get('url'))
        if url:
            previous_by_url[url] = previous_item

    hydrated = 0
    for item in items:
        item_url = normalize_compare_text(getattr(item, 'url', ''))
        previous_item = previous_by_url.get(item_url)
        if not previous_item:
            continue

        item_dict = asdict(item) if hasattr(item, '__dict__') else dict(item or {})
        extra = item_dict.get('extra') if isinstance(item_dict.get('extra'), dict) else {}
        patch = {}

        if not normalize_compare_text(item_dict.get('sku') or extra.get('sku')) and previous_item.get('sku'):
            patch['sku'] = previous_item.get('sku')
        current_stock = normalize_compare_text(item_dict.get('stock_status') or extra.get('stock_status'))
        previous_stock = normalize_compare_text(previous_item.get('stock_status'))
        if previous_stock and (
            not current_stock or
            (not has_specific_stock_detail(current_stock) and has_specific_stock_detail(previous_stock))
        ):
            patch['stock_status'] = previous_item.get('stock_status')
        if not normalize_compare_text(item_dict.get('description') or extra.get('description')) and previous_item.get('description'):
            patch['description'] = previous_item.get('description')

        if patch:
            apply_enriched_item_data(item, patch)
            if hasattr(item, 'extra') and isinstance(item.extra, dict):
                item.extra.update(patch)
            hydrated += 1

    return hydrated


def enrich_scraped_items(items, rules: Dict, retries: int, verify_ssl: bool, use_curl: bool, enrich_details: bool = True, logger=None, use_browser: bool = False, progress_callback=None):
    """Open each unique product detail page and merge richer metadata into scrape results."""
    if not enrich_details or not items:
        return items, 0

    allow_txparts_detail = str(os.getenv('TXPARTS_ENRICH_DETAILS', '')).strip().lower() in {'1', 'true', 'yes', 'on'}
    skipped_detail_urls = 0
    url_to_indexes: Dict[str, List[int]] = {}
    for idx, item in enumerate(items):
        item_sku = str(getattr(item, 'sku', '') or '').strip()
        if item_sku:
            continue
        item_url = normalize_compare_text(getattr(item, 'url', ''))
        if item_url:
            engine_type, _ = get_scraper_for_url(item_url)
            if engine_type == 'txparts' and not allow_txparts_detail:
                skipped_detail_urls += 1
                continue
            url_to_indexes.setdefault(item_url, []).append(idx)

    if not url_to_indexes:
        if skipped_detail_urls and logger:
            logger.info(
                "[detail] Skipped %s TXParts detail URL(s); set TXPARTS_ENRICH_DETAILS=1 to enable TXParts phase-2 detail fetches.",
                skipped_detail_urls,
            )
        return items, 0

    thread_state = threading.local()
    created_sessions = []
    created_sessions_lock = threading.Lock()

    def get_thread_session(engine_type: str):
        if engine_type == 'parts4cells':
            return None

        sessions = getattr(thread_state, 'sessions', None)
        if sessions is None:
            sessions = {}
            thread_state.sessions = sessions

        with created_sessions_lock:
            if engine_type in sessions:
                return sessions[engine_type]

            if engine_type == 'xcell':
                session, _ = xcell_scraper_engine.build_session(retries=retries, verify_ssl=verify_ssl)
            elif engine_type == 'txparts':
                session, _ = txparts_scraper_engine.build_session(retries=retries, verify_ssl=verify_ssl)
            elif engine_type == 'phonelcdparts':
                session, _ = phonelcdparts_scraper_engine.build_session(retries=retries, verify_ssl=verify_ssl)
            elif engine_type == 'gadgetfix':
                session, _ = gadgetfix_scraper_engine.build_session(retries=retries, verify_ssl=verify_ssl)
            else:
                session, _ = build_session(retries=retries, verify_ssl=verify_ssl, use_curl=use_curl)

            sessions[engine_type] = session
            created_sessions.append(session)
            return session

    def enrich_one(item_url: str, item):
        def _do_enrich(use_browser_mode: bool):
            with browser_fetch_mode(use_browser_mode):
                engine_type, _ = get_scraper_for_url(item_url)
                if engine_type == 'xcell':
                    return xcell_scraper_engine.enrich_item_details(get_thread_session(engine_type), item, rules, logger)
                elif engine_type == 'txparts':
                    return txparts_scraper_engine.enrich_item_details(get_thread_session(engine_type), item, rules, logger)
                elif engine_type == 'parts4cells':
                    return parts4cells_scraper_engine.enrich_item_details(None, item, rules, logger)
                elif engine_type == 'phonelcdparts':
                    return phonelcdparts_scraper_engine.enrich_item_details(get_thread_session(engine_type), item, rules, logger)
                elif engine_type == 'gadgetfix':
                    return gadgetfix_scraper_engine.enrich_item_details(get_thread_session(engine_type), item, rules, logger)
                else:
                    return enrich_standard_item_details(get_thread_session(engine_type), item, rules, logger)

        # Fast Safari 15.5 TLS detail enrichment
        enriched = item
        engine_type, _ = get_scraper_for_url(item_url)
        try:
            candidate = _do_enrich(False)
            if candidate:
                enriched = candidate
        except Exception as http_exc:
            if logger:
                logger.debug(f"[detail] HTTP enrichment skipped for {item_url}: {http_exc}")
            enriched = item

        # Browser fallback for xcell only when HTTP fails or is blocked
        if engine_type == 'xcell' and use_browser:
            result_dict = asdict(enriched) if hasattr(enriched, '__dataclass_fields__') else {}
            if not result_dict.get('sku') and not result_dict.get('description'):
                try:
                    browser_enriched = _do_enrich(True)
                    if browser_enriched:
                        enriched = browser_enriched
                except Exception as browser_exc:
                    if logger:
                        logger.debug(f"[detail] Browser fallback also failed for {item_url}: {browser_exc}")

        return item_url, asdict(enriched)

    enriched_count = 0
    total_to_enrich = len(url_to_indexes)
    max_workers = min(16, max(1, len(url_to_indexes)))

    if progress_callback:
        try:
            progress_callback({
                'phase': 2,
                'phase_name': 'Phase 2: Product SKU & Detail Scan',
                'phase2_completed': 0,
                'phase2_total': total_to_enrich,
                'current_items': len(items),
            })
        except Exception:
            pass

    try:
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = {
                executor.submit(enrich_one, item_url, items[indexes[0]]): item_url
                for item_url, indexes in url_to_indexes.items()
            }
            for future in as_completed(futures):
                item_url = futures[future]
                try:
                    _, enriched_data = future.result()
                except Exception as exc:
                    if logger:
                        logger.warning(f"[detail] Failed to enrich {item_url}: {exc}")
                    continue

                for idx in url_to_indexes[item_url]:
                    apply_enriched_item_data(items[idx], enriched_data)
                enriched_count += 1

                if progress_callback:
                    try:
                        progress_callback({
                            'phase': 2,
                            'phase_name': 'Phase 2: Product SKU & Detail Scan',
                            'phase2_completed': enriched_count,
                            'phase2_total': total_to_enrich,
                            'current_items': len(items),
                            'last_item_url': item_url,
                        })
                    except Exception:
                        pass
    finally:
        for session in created_sessions:
            if session and hasattr(session, 'close'):
                try:
                    session.close()
                except Exception:
                    pass

    if logger:
        logger.info(f"[detail] Enriched {enriched_count} item(s) across {len(url_to_indexes)} unique product URLs")
    return items, enriched_count

def extract_item_price(item, *, prefer_adjusted: bool = True) -> float | None:
    """Extract either the adjusted or source price from any scraper item shape."""
    item_dict = asdict(item) if hasattr(item, '__dict__') else dict(item or {})

    numeric_keys = (
        ('discounted_value', 'price_value', 'discounted', 'original')
        if prefer_adjusted
        else ('price_value', 'original', 'discounted_value', 'discounted')
    )
    text_keys = (
        ('discounted_formatted', 'original_formatted', 'price_text')
        if prefer_adjusted
        else ('original_formatted', 'price_text', 'discounted_formatted')
    )

    for key in numeric_keys:
        value = item_dict.get(key)
        if isinstance(value, (int, float)) and float(value) > 0:
            return round(float(value), 2)

    for key in text_keys:
        value = parse_price_number(str(item_dict.get(key) or ''))
        if value is not None and float(value) > 0:
            return round(float(value), 2)

    return None


def get_effective_item_price(item) -> float | None:
    """Extract the adjusted/display price from any scraper item shape."""
    return extract_item_price(item, prefer_adjusted=True)


def get_comparable_item_price(item) -> float | None:
    """Extract the stable source price used for run-to-run comparisons."""
    return extract_item_price(item, prefer_adjusted=False)

def build_price_drop_alerts(items, previous_prices: Dict[str, Dict], drop_pct: float) -> List[Dict]:
    """Compare current items against the latest saved DB prices for the same URL."""
    alerts = []

    for item in items:
        item_dict = asdict(item) if hasattr(item, '__dict__') else dict(item or {})
        url = str(item_dict.get('url') or '').strip()
        if not url:
            continue

        previous = previous_prices.get(url)
        if not previous or previous.get('price') is None:
            continue

        current_price = get_comparable_item_price(item)
        previous_price = float(previous['price'])
        if current_price is None or previous_price <= 0 or current_price >= previous_price:
            continue

        percent_drop = round(((previous_price - current_price) / previous_price) * 100, 2)
        if percent_drop < drop_pct:
            continue

        alerts.append({
            'url': url,
            'title': item_dict.get('title', ''),
            'site': item_dict.get('site', ''),
            'previous_price': previous_price,
            'current_price': round(current_price, 2),
            'drop_percent': percent_drop,
            'previous_history_id': previous.get('history_id'),
            'previous_timestamp': previous.get('timestamp'),
        })

    alerts.sort(key=lambda alert: alert['drop_percent'], reverse=True)
    return alerts

def normalize_item_snapshot(item) -> Dict[str, object]:
    """Create a consistent comparison snapshot from current or historical item data."""
    item_dict = asdict(item) if hasattr(item, '__dict__') else dict(item or {})
    extra = item_dict.get('extra') if isinstance(item_dict.get('extra'), dict) else {}
    comparison_price = get_comparable_item_price(item_dict)
    target_url = normalize_compare_text(item_dict.get('target_url') or extra.get('target_url'))
    target_label = normalize_compare_text(item_dict.get('target_label') or extra.get('target_label'))
    model_label = normalize_compare_text(item_dict.get('model_label') or extra.get('model_label') or target_label or format_category_label_from_url(target_url))
    url = normalize_compare_text(item_dict.get('url'))
    title = normalize_compare_text(item_dict.get('title'))
    sku = normalize_compare_text(item_dict.get('sku') or extra.get('sku'))
    stock_status = normalize_compare_text(item_dict.get('stock_status') or extra.get('stock_status'))
    description = normalize_compare_text(item_dict.get('description') or extra.get('description'))
    canonical_url = get_item_extra_value(
        item_dict,
        extra,
        'canonical_url',
        'canonical',
        'product_url',
        'permalink',
    ) or url
    product_id = get_item_extra_value(
        item_dict,
        extra,
        'product_id',
        'productId',
        'data_product_id',
        'woocommerce_product_id',
        'id',
    )
    variant_id = get_item_extra_value(
        item_dict,
        extra,
        'variant_id',
        'variation_id',
        'variantId',
        'data_variation_id',
    )
    product_type = get_item_extra_value(item_dict, extra, 'product_type', 'type')
    brand = get_item_extra_value(item_dict, extra, 'brand')
    model = get_item_extra_value(item_dict, extra, 'model', 'device_model')
    display_position = get_item_extra_value(item_dict, extra, 'display_position', 'position')
    frame_type = get_item_extra_value(item_dict, extra, 'frame_type')
    display_quality = get_item_extra_value(item_dict, extra, 'display_quality', 'quality', 'grade')
    return {
        'url': url,
        'url_compare': normalize_compare_url(url),
        'canonical_url': canonical_url,
        'canonical_url_compare': normalize_compare_url(canonical_url),
        'site': normalize_compare_text(item_dict.get('site')),
        'source': normalize_compare_text(item_dict.get('source')),
        'title': title,
        'title_compare': normalize_semantic_compare_text(title),
        'sku': sku,
        'sku_compare': normalize_identifier_compare_text(sku),
        'product_id': product_id,
        'product_id_compare': normalize_identifier_compare_text(product_id),
        'variant_id': variant_id,
        'variant_id_compare': normalize_identifier_compare_text(variant_id),
        'product_type': product_type,
        'product_type_compare': normalize_semantic_compare_text(product_type),
        'brand': brand,
        'brand_compare': normalize_semantic_compare_text(brand),
        'model': model,
        'model_compare': normalize_semantic_compare_text(model),
        'display_position': display_position,
        'display_position_compare': normalize_semantic_compare_text(display_position),
        'frame_type': frame_type,
        'frame_type_compare': normalize_semantic_compare_text(frame_type),
        'display_quality': display_quality,
        'display_quality_compare': normalize_semantic_compare_text(display_quality),
        'stock_status': stock_status,
        'stock_status_compare': normalize_stock_status_for_compare(stock_status),
        'description': description,
        'description_compare': normalize_semantic_compare_text(description),
        'comparison_price': comparison_price,
        'target_url': target_url,
        'target_label': target_label,
        'model_label': model_label,
        'price_formatted': normalize_compare_text(
            item_dict.get('original_formatted')
            or item_dict.get('price_text')
            or item_dict.get('discounted_formatted')
        ) if comparison_price is not None else '',
        'adjusted_price_formatted': normalize_compare_text(
            item_dict.get('discounted_formatted')
            or item_dict.get('original_formatted')
            or item_dict.get('price_text')
        ) if comparison_price is not None else '',
    }


def is_comparable_product_snapshot(snapshot: Dict[str, object]) -> bool:
    """Reject category/navigation rows and scraper placeholders from comparisons."""
    url = str(snapshot.get('url_compare') or '')
    title = str(snapshot.get('title') or '')
    if not url or not title or str(snapshot.get('source') or '').lower() == 'error':
        return False

    parsed = urlparse(url)
    host = (parsed.hostname or '').lower()
    path = (parsed.path or '').lower()
    if '/product-category/' in path:
        return False
    if host == 'xcellparts.com' or host.endswith('.xcellparts.com'):
        if '/product/' not in path:
            return False
    if (host == 'gadgetfix.com' or host.endswith('.gadgetfix.com')) and '/category/' in path:
        return False

    looks_like_category_count = bool(re.search(r'\b\d+\s+products?$', title, re.IGNORECASE))
    has_product_metadata = bool(
        snapshot.get('sku')
        or snapshot.get('comparison_price') is not None
        or snapshot.get('stock_status')
    )
    if looks_like_category_count and not has_product_metadata:
        return False
    return True


def snapshot_quality(snapshot: Dict[str, object]) -> tuple:
    """Prefer the most complete row when overlapping targets return one product."""
    return (
        sum(bool(snapshot.get(key)) for key in (
            'sku',
            'stock_status',
            'description',
            'title',
            'site',
            'model_label',
            'target_label',
        )),
        int(snapshot.get('comparison_price') is not None),
        len(str(snapshot.get('description') or '')),
        len(str(snapshot.get('title') or '')),
    )


def prepare_comparison_snapshots(items) -> Tuple[List[Dict[str, object]], Dict[str, int]]:
    """Filter and deduplicate rows for product-level history comparisons."""
    raw_snapshots = [normalize_item_snapshot(item) for item in items or []]
    comparable = [snapshot for snapshot in raw_snapshots if is_comparable_product_snapshot(snapshot)]
    by_url: Dict[str, Dict[str, object]] = {}
    for snapshot in comparable:
        key = str(snapshot.get('url_compare') or '')
        existing = by_url.get(key)
        if existing is None or snapshot_quality(snapshot) > snapshot_quality(existing):
            by_url[key] = snapshot
    return list(by_url.values()), {
        'rows': len(raw_snapshots),
        'excluded_non_products': len(raw_snapshots) - len(comparable),
        'duplicate_rows': len(comparable) - len(by_url),
        'unique_products': len(by_url),
    }


def deduplicate_comparable_items(items) -> List[Dict[str, object]]:
    """Return original product records deduplicated by canonical URL for API display, preserving category occurrences."""
    by_url: Dict[str, Dict[str, object]] = {}
    category_map: Dict[str, List[str]] = {}
    url_counts: Dict[str, int] = {}
    out: List[Dict[str, object]] = []

    for item in items or []:
        item_dict = asdict(item) if hasattr(item, '__dataclass_fields__') else dict(item or {})
        url = str(item_dict.get('url') or '').strip().rstrip('/')
        if not url:
            continue

        url_counts[url] = url_counts.get(url, 0) + 1

        extra_val = item_dict.get('extra') or item_dict.get('extra_json')
        if isinstance(extra_val, str):
            try:
                extra = json.loads(extra_val)
            except Exception:
                extra = {}
        elif isinstance(extra_val, dict):
            extra = extra_val
        else:
            extra = {}

        cat = (
            extra.get('model_label')
            or extra.get('target_label')
            or item_dict.get('category')
            or extra.get('category')
            or ''
        )
        if not cat and extra.get('target_url'):
            raw_cat = str(extra['target_url']).rstrip('/').split('/')[-1].replace('.html', '').replace('-', ' ').replace('_', ' ')
            cat = ' '.join(word.capitalize() for word in raw_cat.split())

        if url not in category_map:
            category_map[url] = []
        if cat and cat not in category_map[url]:
            category_map[url].append(cat)

        if url not in by_url:
            if not isinstance(item_dict.get('extra'), dict):
                item_dict['extra'] = extra
            by_url[url] = item_dict
            out.append(item_dict)

    for item_dict in out:
        url = str(item_dict.get('url') or '').strip().rstrip('/')
        cats = category_map.get(url, [])
        occurrences = url_counts.get(url, 1)
        is_dup = occurrences > 1 or len(cats) > 1
        item_dict['duplicate_categories'] = cats
        item_dict['duplicate_count'] = max(occurrences, len(cats))
        item_dict['is_duplicate'] = is_dup
        if not item_dict.get('category') and cats:
            item_dict['category'] = cats[0]

    return out


def public_comparison_snapshot(snapshot: Dict[str, object]) -> Dict[str, object]:
    return {
        key: value
        for key, value in snapshot.items()
        if not key.endswith('_compare') and key != 'source'
    }


def is_usable_scraped_item(item) -> bool:
    """Return whether a scraped row is a real product rather than an error placeholder."""
    item_dict = asdict(item) if hasattr(item, '__dict__') else dict(item or {})
    title = normalize_compare_text(item_dict.get('title'))
    url = normalize_compare_text(item_dict.get('url'))
    source = normalize_compare_text(item_dict.get('source')).lower()
    price_text = normalize_compare_text(
        item_dict.get('price_text')
        or item_dict.get('original_formatted')
        or item_dict.get('discounted_formatted')
    ).lower()
    if source == 'error' or price_text.startswith(('fetch_failed:', 'parallel_scrape_failed:')):
        return False
    return bool(title and url)


def serialize_scraped_item(item) -> Dict[str, object]:
    """Return a JSON-safe item snapshot for API responses and live run previews."""
    item_dict = asdict(item) if hasattr(item, '__dict__') else dict(item or {})
    if not isinstance(item_dict.get('extra'), dict):
        item_dict['extra'] = {}
    return item_dict

def build_session_comparison(
    previous_history: Dict | None,
    current_items,
    current_target_urls=None,
    run_validation: Dict | None = None,
    target_errors: List[Dict[str, object]] | None = None,
) -> Dict:
    """Compare current scrape results against the latest previous run for the same target URLs."""
    previous_items = (previous_history or {}).get('items', []) if previous_history else []
    current_snapshots, current_metrics = prepare_comparison_snapshots(current_items)
    previous_snapshots, previous_metrics = prepare_comparison_snapshots(previous_items)
    current_target_scope = {
        normalize_compare_url(url)
        for url in (current_target_urls or [])
        if normalize_compare_url(url)
    }

    comparison = {
        'has_previous_run': previous_history is not None,
        'previous_history_id': previous_history.get('id') if previous_history else None,
        'previous_timestamp': previous_history.get('timestamp') if previous_history else None,
        'summary': {
            'previous_items': len(previous_snapshots),
            'current_items': len(current_snapshots),
            'previous_rows': previous_metrics['rows'],
            'current_rows': current_metrics['rows'],
            'excluded_previous_non_products': previous_metrics['excluded_non_products'],
            'excluded_current_non_products': current_metrics['excluded_non_products'],
            'duplicate_previous_rows': previous_metrics['duplicate_rows'],
            'duplicate_current_rows': current_metrics['duplicate_rows'],
            'out_of_scope_previous_products': 0,
            'added': 0,
            'removed': 0,
            'changed': 0,
            'price_changes': 0,
            'stock_changes': 0,
            'title_changes': 0,
            'sku_changes': 0,
            'description_changes': 0,
            'url_changes': 0,
            'category_changes': 0,
            'temporarily_missing': 0,
            'verification_required': 0,
            'scrape_failures': 0,
            'removed_confirmed': 0,
            'review_required': 0,
        },
        'added': [],
        'removed': [],
        'removed_confirmed': [],
        'temporarily_missing': [],
        'verification_required': [],
        'review_required': [],
        'scrape_failures': [],
        'changed': [],
    }

    if not previous_history:
        return comparison

    matched_previous = set()
    matched_current = set()
    matched_pairs = []

    def match_ref(snapshot: Dict[str, object]) -> str:
        return str(snapshot.get('url_compare') or id(snapshot))

    matched_prev_ids = set()
    matched_curr_ids = set()

    identity_matchers = (
        ('product_id_compare', 'product_id'),
        ('variant_id_compare', 'variant_id'),
        ('sku_compare', 'sku'),
        ('canonical_url_compare', 'canonical_url'),
        ('url_compare', 'url'),
        ('title_compare', 'title'),
    )

    for key_name, reason in identity_matchers:
        prev_map = {}
        curr_map = {}
        for item in previous_snapshots:
            item_id = match_ref(item)
            if item_id in matched_prev_ids:
                continue
            k = item.get(key_name)
            if k:
                prev_map.setdefault(k, []).append(item)
        for item in current_snapshots:
            item_id = match_ref(item)
            if item_id in matched_curr_ids:
                continue
            k = item.get(key_name)
            if k:
                curr_map.setdefault(k, []).append(item)

        for k in set(prev_map) & set(curr_map):
            p_list = prev_map[k]
            c_list = curr_map[k]
            if len(p_list) == 1 and len(c_list) == 1:
                p_item = p_list[0]
                c_item = c_list[0]
                matched_pairs.append((p_item, c_item, reason))
                matched_prev_ids.add(match_ref(p_item))
                matched_curr_ids.add(match_ref(c_item))

    remaining_previous = [item for item in previous_snapshots if match_ref(item) not in matched_prev_ids]
    remaining_current = [item for item in current_snapshots if match_ref(item) not in matched_curr_ids]

    error_targets = {
        normalize_guard_url((error or {}).get('url') or '')
        for error in (target_errors or [])
        if normalize_guard_url((error or {}).get('url') or '')
    }
    validation_approved = bool((run_validation or {}).get('approved', True))

    for item in remaining_current:
        public_item = public_comparison_snapshot(item)
        if validation_approved:
            public_item['status'] = 'Newly Added'
            comparison['added'].append(public_item)
        else:
            public_item['status'] = 'Review Required'
            public_item['reason'] = 'The current run failed scrape-completeness validation.'
            comparison['review_required'].append(public_item)

    def missing_status_for(item: Dict[str, object]) -> Tuple[str, str]:
        target_url = normalize_guard_url(item.get('target_url') or '')
        if target_url and target_url in error_targets:
            return 'Scrape Failure', 'The product target had a fetch or parse error in the current run.'
        if not validation_approved:
            return 'Verification Required', 'The current run failed scrape-completeness validation.'
        return 'Temporarily Missing', 'Missing from this run only; direct verification is required before confirming removal.'

    for item in remaining_previous:
        target_url = normalize_compare_url(item.get('target_url'))
        if current_target_scope and target_url and target_url not in current_target_scope:
            comparison['summary']['out_of_scope_previous_products'] += 1
            continue
        status, reason = missing_status_for(item)
        public_item = public_comparison_snapshot(item)
        public_item['status'] = status
        public_item['reason'] = reason
        comparison['verification_required'].append(public_item)
        if status == 'Temporarily Missing':
            # A product missing from one scrape only is NOT confirmed removed.
            # It goes into temporarily_missing for monitoring â€” never into removed.
            comparison['temporarily_missing'].append(public_item)
        elif status == 'Scrape Failure':
            # Target failed to fetch/parse â€” product fate unknown, not removed.
            comparison['scrape_failures'].append(public_item)
        else:
            # 'Review Required' (run failed completeness validation) â€” treated as
            # a potential removal candidate requiring manual review.
            comparison['removed'].append(public_item)

    for prev_item, current_item, match_reason in matched_pairs:
        field_changes = {}

        if prev_item['url_compare'] != current_item['url_compare']:
            field_changes['url'] = {'before': prev_item['url'], 'after': current_item['url']}
        if prev_item['title_compare'] != current_item['title_compare']:
            field_changes['title'] = {'before': prev_item['title'], 'after': current_item['title']}
        if normalize_compare_url(prev_item.get('target_url')) != normalize_compare_url(current_item.get('target_url')):
            field_changes['target_url'] = {'before': prev_item.get('target_url', ''), 'after': current_item.get('target_url', '')}
        if prev_item['sku_compare'] and current_item['sku_compare'] and prev_item['sku_compare'] != current_item['sku_compare']:
            field_changes['sku'] = {'before': prev_item['sku'], 'after': current_item['sku']}
        if (
            prev_item['description_compare']
            and current_item['description_compare']
            and prev_item['description_compare'] != current_item['description_compare']
        ):
            field_changes['description'] = {'before': prev_item['description'], 'after': current_item['description']}
        if (
            prev_item['stock_status_compare']
            and current_item['stock_status_compare']
            and prev_item['stock_status_compare'] != current_item['stock_status_compare']
        ):
            field_changes['stock_status'] = {'before': prev_item['stock_status'], 'after': current_item['stock_status']}

        prev_price = prev_item.get('comparison_price')
        current_price = current_item.get('comparison_price')
        if prev_price is not None and current_price is not None and abs(prev_price - current_price) > 0.009:
            field_changes['price'] = {
                'before': prev_price,
                'after': current_price,
                'before_formatted': prev_item.get('price_formatted', ''),
                'after_formatted': current_item.get('price_formatted', ''),
            }

        if field_changes:
            status = 'Updated'
            if set(field_changes) == {'url'}:
                status = 'URL Changed'
            elif set(field_changes) == {'target_url'}:
                status = 'Category Changed'
            elif 'title' in field_changes and not ({'price', 'stock_status'} & set(field_changes)):
                status = 'Product Renamed'
            elif set(field_changes) == {'price'}:
                status = 'Price Changed'
            elif set(field_changes) == {'stock_status'}:
                status = 'Stock Changed'
            comparison['changed'].append({
                'before': public_comparison_snapshot(prev_item),
                'after': public_comparison_snapshot(current_item),
                'changes': field_changes,
                'status': status,
                'match_reason': match_reason,
            })

    comparison['summary']['added'] = len(comparison['added'])
    comparison['summary']['removed'] = len(comparison['removed'])
    comparison['summary']['removed_confirmed'] = len(comparison['removed_confirmed'])
    comparison['summary']['temporarily_missing'] = len(comparison['temporarily_missing'])
    comparison['summary']['verification_required'] = len(comparison['verification_required'])
    comparison['summary']['scrape_failures'] = len(comparison['scrape_failures'])
    comparison['summary']['review_required'] = len(comparison['review_required'])
    comparison['summary']['changed'] = len(comparison['changed'])
    comparison['summary']['price_changes'] = sum(1 for change in comparison['changed'] if 'price' in change['changes'])
    comparison['summary']['stock_changes'] = sum(1 for change in comparison['changed'] if 'stock_status' in change['changes'])
    comparison['summary']['title_changes'] = sum(1 for change in comparison['changed'] if 'title' in change['changes'])
    comparison['summary']['sku_changes'] = sum(1 for change in comparison['changed'] if 'sku' in change['changes'])
    comparison['summary']['description_changes'] = sum(1 for change in comparison['changed'] if 'description' in change['changes'])
    comparison['summary']['url_changes'] = sum(1 for change in comparison['changed'] if 'url' in change['changes'])
    comparison['summary']['category_changes'] = sum(1 for change in comparison['changed'] if 'target_url' in change['changes'])
    return comparison


def build_public_history_id(scraper_key: str, raw_history_id: str) -> str:
    normalized_scraper = str(scraper_key or '').strip()
    normalized_history_id = str(raw_history_id or '').strip()
    if not normalized_scraper or not normalized_history_id:
        return ''
    return f"{normalized_scraper}:{normalized_history_id}"


def save_automation_partial_history(
    run_id: int,
    job: Dict | None,
    target_urls: List[str],
    *,
    previous_history: Dict | None = None,
    error_text: str = '',
) -> Tuple[str, int, Dict]:
    latest_run = db_manager.get_automation_run(run_id)
    run_summary = latest_run.get('summary') if isinstance((latest_run or {}).get('summary'), dict) else {}
    checkpoint_items = run_summary.get('preview_items') if isinstance(run_summary.get('preview_items'), list) else []
    persisted_items = db_manager.get_automation_run_items(run_id)
    checkpoint_count = int(
        run_summary.get('current_items')
        or run_summary.get('items_count')
        or (latest_run or {}).get('items_count')
        or len(persisted_items)
        or len(checkpoint_items)
        or 0
    )
    usable_items = [dict(item) for item in persisted_items if isinstance(item, dict)]
    if not usable_items:
        usable_items = [dict(item) for item in checkpoint_items if isinstance(item, dict)]
    if not usable_items:
        return '', 0, run_summary

    raw_history_id = str(int(time.time() * 1000))
    history_rules = dict((job or {}).get('rules', {}) or {})
    if job:
        history_rules['_automation_job_id'] = job.get('id')
        history_rules['_automation_job_name'] = job.get('name')
        history_rules['_automation_category_query'] = job.get('category_query')
    history_rules['_automation_partial'] = True
    history_rules['_automation_partial_run_id'] = run_id
    history_rules['_automation_partial_reason'] = str(error_text or 'Automation run failed before completion.')

    if not db_manager.save_fetch_history(raw_history_id, target_urls, usable_items, history_rules):
        app.logger.error(f"[automation] Failed to save partial history for run {run_id}")
        return '', len(usable_items), run_summary

    scraper_key = str((job or {}).get('scraper_key') or (latest_run or {}).get('scraper_key') or '').strip()
    public_history_id = build_public_history_id(scraper_key, raw_history_id)
    partial_summary = dict(run_summary)
    partial_summary['partial_history_saved'] = True
    partial_summary['partial_history_id'] = public_history_id
    partial_summary['partial_history_items'] = len(usable_items)
    if error_text:
        partial_summary['partial_history_reason'] = str(error_text)
    if previous_history is not None:
        partial_summary['partial_comparison'] = build_session_comparison(
            previous_history,
            usable_items,
            current_target_urls=target_urls,
        )
    return public_history_id, len(usable_items), partial_summary


def get_comparison_model_label(item: Dict[str, object]) -> str:
    snapshot = dict(item or {})
    return (
        normalize_compare_text(snapshot.get('model_label'))
        or normalize_compare_text(snapshot.get('target_label'))
        or format_category_label_from_url(snapshot.get('target_url'))
        or format_category_label_from_url(snapshot.get('url'))
        or normalize_compare_text(snapshot.get('title'))
    )


def build_automation_model_summary(comparison: Dict) -> List[Dict[str, object]]:
    buckets: Dict[str, Dict[str, object]] = {}

    def touch(label: str):
        clean_label = normalize_compare_text(label) or 'Uncategorized'
        return buckets.setdefault(clean_label, {
            'model': clean_label,
            'added': 0,
            'removed': 0,
            'changed': 0,
            'price_changes': 0,
            'stock_changes': 0,
        })

    for item in comparison.get('added', []):
        touch(get_comparison_model_label(item))['added'] += 1

    for item in comparison.get('removed', []):
        touch(get_comparison_model_label(item))['removed'] += 1

    for entry in comparison.get('changed', []):
        bucket = touch(get_comparison_model_label(entry.get('after') or entry.get('before') or {}))
        bucket['changed'] += 1
        if 'price' in (entry.get('changes') or {}):
            bucket['price_changes'] += 1
        if 'stock_status' in (entry.get('changes') or {}):
            bucket['stock_changes'] += 1

    return sorted(
        buckets.values(),
        key=lambda item: (-int(item['changed']) - int(item['added']) - int(item['removed']), str(item['model']).lower())
    )


def build_automation_run_summary(target_urls: List[str], comparison: Dict, price_drops: List[Dict]) -> Dict[str, object]:
    summary = dict((comparison or {}).get('summary') or {})
    summary['target_count'] = len(target_urls or [])
    summary['price_drop_alerts'] = len(price_drops or [])
    summary['models'] = build_automation_model_summary(comparison or {})
    return summary


def scraper_guard_enabled() -> bool:
    value = os.getenv('SCRAPER_ANOMALY_GUARD', '1')
    return str(value or '').strip().lower() not in {'0', 'false', 'no', 'off'}


def baseline_protection_enabled() -> bool:
    value = os.getenv('SCRAPER_BASELINE_PROTECTION', '1')
    return str(value or '').strip().lower() not in {'0', 'false', 'no', 'off'}


def history_retention_keep_count() -> int:
    try:
        return max(0, int(os.getenv('SCRAPER_HISTORY_KEEP_PER_URL_SET') or '2'))
    except (TypeError, ValueError):
        return 2


def chatgpt_incident_reporting_enabled() -> bool:
    value = os.getenv('SCRAPER_CHATGPT_AUTO_REPORT', '0')
    return str(value or '').strip().lower() in {'1', 'true', 'yes', 'on'}


def normalize_guard_url(value: str) -> str:
    text = str(value or '').strip()
    if not text:
        return ''
    parsed = urlparse(text)
    path = parsed.path.rstrip('/') or '/'
    return parsed._replace(path=path, params='', query='', fragment='').geturl().lower()


def target_counts_from_history(history: Dict | None) -> Dict[str, int]:
    counts: Dict[str, int] = {}
    if not history:
        return counts
    for item in history.get('items', []) or []:
        snapshot = normalize_item_snapshot(item)
        target_url = normalize_guard_url(snapshot.get('target_url') or '')
        if not target_url:
            continue
        counts[target_url] = counts.get(target_url, 0) + 1
    return counts


def count_items_by_target(items) -> Dict[str, int]:
    counts: Dict[str, int] = {}
    for item in items or []:
        if not is_usable_scraped_item(item):
            continue
        item_dict = asdict(item) if hasattr(item, '__dict__') else dict(item or {})
        extra = item_dict.get('extra') if isinstance(item_dict.get('extra'), dict) else {}
        target_url = normalize_guard_url(item_dict.get('target_url') or extra.get('target_url') or '')
        if not target_url:
            continue
        counts[target_url] = counts.get(target_url, 0) + 1
    return counts


def detect_sparse_target_anomalies(
    urls: List[str],
    current_items,
    previous_history: Dict | None,
    engine_used: Dict[str, str],
) -> List[Dict[str, object]]:
    if not scraper_guard_enabled() or not previous_history:
        return []
    try:
        min_previous = max(1, int(os.getenv('SCRAPER_ANOMALY_MIN_PREVIOUS') or '10'))
    except (TypeError, ValueError):
        min_previous = 10
    try:
        sparse_ratio = max(0.01, min(0.9, float(os.getenv('SCRAPER_ANOMALY_SPARSE_RATIO') or '0.25')))
    except (TypeError, ValueError):
        sparse_ratio = 0.25
    try:
        max_sparse_items = max(0, int(os.getenv('SCRAPER_ANOMALY_MAX_SPARSE_ITEMS') or '2'))
    except (TypeError, ValueError):
        max_sparse_items = 2

    normalized_urls = [normalize_guard_url(url) for url in (urls or []) if str(url or '').strip()]
    parent_target_urls = {
        url
        for url in normalized_urls
        if any(other != url and other.startswith(f"{url.rstrip('/')}/") for other in normalized_urls)
    }

    previous_counts = target_counts_from_history(previous_history)
    current_counts = count_items_by_target(current_items)
    anomalies = []
    for url in urls or []:
        normalized_url = normalize_guard_url(url)
        if normalized_url in parent_target_urls:
            continue
        previous_count = int(previous_counts.get(normalized_url) or 0)
        if previous_count < min_previous:
            continue
        current_count = int(current_counts.get(normalized_url) or 0)
        ratio_limit = max_sparse_items
        ratio_limit = max(ratio_limit, int(previous_count * sparse_ratio))
        if current_count <= ratio_limit and current_count <= max_sparse_items:
            anomalies.append({
                'type': 'sparse_target_result',
                'url': url,
                'engine': engine_used.get(url, detect_scraper_key(url)),
                'previous_count': previous_count,
                'current_count': current_count,
                'threshold_ratio': sparse_ratio,
                'max_sparse_items': max_sparse_items,
            })
    return anomalies


def count_comparison_snapshots_by_target(snapshots: List[Dict[str, object]]) -> Dict[str, int]:
    counts: Dict[str, int] = {}
    for snapshot in snapshots or []:
        target_url = normalize_guard_url(snapshot.get('target_url') or '')
        if not target_url:
            continue
        counts[target_url] = counts.get(target_url, 0) + 1
    return counts


def parse_guard_float(name: str, default: float, *, lower: float = 0.0, upper: float | None = None) -> float:
    try:
        value = float(os.getenv(name) or default)
    except (TypeError, ValueError):
        value = default
    value = max(lower, value)
    if upper is not None:
        value = min(upper, value)
    return value


def parse_guard_int(name: str, default: int, *, lower: int = 0) -> int:
    try:
        value = int(os.getenv(name) or default)
    except (TypeError, ValueError):
        value = default
    return max(lower, value)


def validate_scrape_completeness(
    urls: List[str],
    current_items,
    previous_history: Dict | None,
    target_errors: List[Dict[str, object]] | None = None,
) -> Dict[str, object]:
    """Protect saved history/comparisons from partial or failed scrape runs."""
    current_snapshots, current_metrics = prepare_comparison_snapshots(current_items)
    previous_items = (previous_history or {}).get('items', []) if previous_history else []
    previous_snapshots, previous_metrics = prepare_comparison_snapshots(previous_items)
    validation = {
        'approved': True,
        'status': 'Approved for Comparison',
        'reasons': [],
        'metrics': {
            'previous_items': len(previous_snapshots),
            'current_items': len(current_snapshots),
            'previous_rows': previous_metrics.get('rows', 0),
            'current_rows': current_metrics.get('rows', 0),
            'target_count': len(urls or []),
            'target_error_count': len(target_errors or []),
            'total_drop_items': max(0, len(previous_snapshots) - len(current_snapshots)),
            'total_drop_ratio': 0.0,
            'target_anomaly_count': 0,
        },
        'target_anomalies': [],
        'target_errors': list(target_errors or [])[:50],
    }
    if not baseline_protection_enabled() or not previous_history:
        return validation

    previous_count = len(previous_snapshots)
    current_count = len(current_snapshots)
    if previous_count > 0:
        validation['metrics']['total_drop_ratio'] = round(max(0, previous_count - current_count) / previous_count, 6)

    max_drop_ratio = parse_guard_float('SCRAPER_MAX_TOTAL_DROP_RATIO', 0.03, lower=0.0, upper=0.95)
    max_drop_items = parse_guard_int('SCRAPER_MAX_TOTAL_DROP_ITEMS', 100, lower=1)
    max_target_error_ratio = parse_guard_float('SCRAPER_MAX_TARGET_ERROR_RATIO', 0.0, lower=0.0, upper=1.0)
    min_target_previous = parse_guard_int('SCRAPER_TARGET_DROP_MIN_PREVIOUS', 10, lower=1)
    target_drop_ratio = parse_guard_float('SCRAPER_TARGET_DROP_RATIO', 0.50, lower=0.01, upper=0.99)

    if previous_count > 0 and current_count == 0:
        validation['reasons'].append('Current scrape returned zero comparable products while a previous baseline exists.')

    total_drop = previous_count - current_count
    if previous_count > 0 and total_drop > 0:
        allowed_drop = max(max_drop_items, int(previous_count * max_drop_ratio))
        if total_drop > allowed_drop:
            validation['reasons'].append(
                f'Comparable product count dropped by {total_drop} ({validation["metrics"]["total_drop_ratio"]:.2%}), '
                f'exceeding allowed drop of {allowed_drop}.'
            )

    normalized_urls = [normalize_guard_url(url) for url in (urls or []) if str(url or '').strip()]
    parent_target_urls = {
        url
        for url in normalized_urls
        if any(other != url and other.startswith(f"{url.rstrip('/')}/") for other in normalized_urls)
    }
    previous_counts = count_comparison_snapshots_by_target(previous_snapshots)
    current_counts = count_comparison_snapshots_by_target(current_snapshots)
    for url in normalized_urls:
        if url in parent_target_urls:
            continue
        previous_target_count = int(previous_counts.get(url) or 0)
        if previous_target_count < min_target_previous:
            continue
        current_target_count = int(current_counts.get(url) or 0)
        if current_target_count < int(previous_target_count * target_drop_ratio):
            validation['target_anomalies'].append({
                'type': 'target_count_drop',
                'url': url,
                'previous_count': previous_target_count,
                'current_count': current_target_count,
                'threshold_ratio': target_drop_ratio,
            })

    validation['metrics']['target_anomaly_count'] = len(validation['target_anomalies'])
    if validation['target_anomalies']:
        validation['reasons'].append(
            f'{len(validation["target_anomalies"])} target(s) dropped below the configured per-target threshold.'
        )

    previous_targets_with_items = {url for url, count in previous_counts.items() if count > 0}
    failed_targets = {
        normalize_guard_url((error or {}).get('url') or '')
        for error in (target_errors or [])
        if normalize_guard_url((error or {}).get('url') or '')
    }
    relevant_failed_targets = sorted(previous_targets_with_items & failed_targets)
    if relevant_failed_targets:
        allowed_errors = int(max_target_error_ratio * max(1, len(previous_targets_with_items)))
        if len(relevant_failed_targets) > allowed_errors:
            validation['reasons'].append(
                f'{len(relevant_failed_targets)} target(s) with previous products had scrape errors in this run.'
            )

    if validation['reasons']:
        validation['approved'] = False
        validation['status'] = 'Rejected by Validation'
    return validation


def validation_from_history_rejection(history: Dict | None) -> Dict | None:
    rules = (history or {}).get('rules')
    if not isinstance(rules, dict) or not rules.get('_baseline_rejected'):
        return None
    reason = normalize_compare_text(
        rules.get('_baseline_rejection_reason')
        or 'This saved history was rejected by scrape-completeness validation.'
    )
    return {
        'approved': False,
        'status': 'Rejected by Validation',
        'reasons': [reason],
        'metrics': {
            'current_rows': (history or {}).get('items_count', 0),
            'target_count': len((history or {}).get('urls') or []),
        },
        'target_anomalies': [],
        'target_errors': [],
    }


def backup_scraper_databases(incident_id: str) -> List[str]:
    backup_dir = APP_ROOT / 'data' / 'backups' / 'scraper_incidents' / incident_id
    backup_dir.mkdir(parents=True, exist_ok=True)
    candidates = set()
    db_path = getattr(db_manager, 'db_path', '')
    if db_path:
        candidates.add(Path(db_path))
    site_db_dir = APP_ROOT / 'data' / 'site_dbs'
    if site_db_dir.exists():
        candidates.update(site_db_dir.glob('*.db'))

    backups = []
    for source in sorted(candidates):
        try:
            if not source.exists() or not source.is_file():
                continue
            destination = backup_dir / source.name
            shutil.copy2(source, destination)
            backups.append(str(destination))
        except Exception as exc:
            app.logger.warning(f"[scraper-guard] Could not back up {source}: {exc}")
    return backups


def build_scraper_incident_prompt(incident: Dict[str, object]) -> str:
    anomalies = incident.get('anomalies') or []
    anomaly_lines = '\n'.join(
        f"- {item.get('engine')} | previous={item.get('previous_count')} current={item.get('current_count')} | {item.get('url')}"
        for item in anomalies[:20]
    )
    return f"""A scraper data-quality guard stopped a run before saving a bad comparison.

Incident ID: {incident.get('id')}
Job: {incident.get('job_name')} ({incident.get('scraper_key')})
Reason: categories that previously had many products returned only a tiny number of products.

Evidence:
{anomaly_lines or '- No anomaly rows attached.'}

Repository context:
- Main workflow: app.py execute_scrape_workflow
- Supplier engines: scrapers/*_scraper_engine.py
- Browser helper: scrapers/browser_fetcher.py
- Tests: tests/test_supplier_scrapers.py and tests/test_api_scrape.py

Please inspect the affected supplier UI/parser, identify why the product count dropped, patch the scraper conservatively, add a regression test using representative rendered HTML, run targeted tests, then restart the local server only after backing up databases. Do not save a new scrape history until the target counts are plausible again."""


def maybe_send_incident_to_chatgpt(prompt: str) -> Dict[str, object]:
    if not chatgpt_incident_reporting_enabled():
        return {'enabled': False, 'sent': False, 'reason': 'SCRAPER_CHATGPT_AUTO_REPORT is disabled'}
    api_key = os.getenv('OPENAI_API_KEY', '').strip()
    if not api_key:
        return {'enabled': True, 'sent': False, 'reason': 'OPENAI_API_KEY is not set'}
    model = os.getenv('SCRAPER_CHATGPT_MODEL', 'gpt-4.1-mini').strip() or 'gpt-4.1-mini'
    try:
        response = requests.post(
            'https://api.openai.com/v1/responses',
            headers={
                'Authorization': f'Bearer {api_key}',
                'Content-Type': 'application/json',
            },
            json={
                'model': model,
                'input': prompt,
            },
            timeout=60,
        )
        payload = response.json() if response.content else {}
        return {
            'enabled': True,
            'sent': response.ok,
            'status_code': response.status_code,
            'model': model,
            'response_id': payload.get('id', ''),
            'error': '' if response.ok else payload.get('error', payload),
        }
    except Exception as exc:
        return {'enabled': True, 'sent': False, 'reason': str(exc)}


def create_scraper_guard_incident(
    anomalies: List[Dict[str, object]],
    *,
    urls: List[str],
    automation_job: Dict | None,
    previous_history: Dict | None,
    items_count: int,
) -> Dict[str, object]:
    incident_id = f"scraper-guard-{int(time.time() * 1000)}"
    incident_dir = APP_ROOT / 'output' / 'scraper_incidents'
    incident_dir.mkdir(parents=True, exist_ok=True)
    incident = {
        'id': incident_id,
        'created_at': datetime.datetime.now(pytz.timezone('Asia/Karachi')).isoformat(),
        'scraper_key': detect_scraper_key(urls[0]) if urls else '',
        'job_id': (automation_job or {}).get('id'),
        'job_name': (automation_job or {}).get('name', ''),
        'previous_history_id': (previous_history or {}).get('id', ''),
        'items_count': items_count,
        'target_count': len(urls or []),
        'anomaly_count': len(anomalies or []),
        'anomalies': anomalies[:50],
        'database_backups': backup_scraper_databases(incident_id),
    }
    prompt = build_scraper_incident_prompt(incident)
    incident['prompt_path'] = str(incident_dir / f'{incident_id}.prompt.md')
    incident['report_path'] = str(incident_dir / f'{incident_id}.json')
    incident['chatgpt_report'] = maybe_send_incident_to_chatgpt(prompt)
    Path(incident['prompt_path']).write_text(prompt, encoding='utf-8')
    Path(incident['report_path']).write_text(json.dumps(incident, indent=2, ensure_ascii=True), encoding='utf-8')
    return incident


def execute_scrape_workflow(
    urls_input,
    *,
    crawl_pagination: bool = True,
    max_pages: int = 10,
    delay_ms: int = 50,
    retries: int = 1,
    verify_ssl: bool = True,
    use_curl: bool = False,
    use_browser: bool | None = None,
    use_parallel: bool = True,
    enrich_details: bool = True,
    rules: Dict | None = None,
    drop_pct: float = 10.0,
    target_labels: Dict[str, str] | None = None,
    automation_job: Dict | None = None,
    previous_history_override: Dict | None = None,
    progress_callback = None,
):
    rules = dict(rules or {})
    target_labels = {str(key).strip(): str(value or '').strip() for key, value in (target_labels or {}).items() if str(key).strip()}
    urls = [u.strip() for u in (urls_input.splitlines() if isinstance(urls_input, str) else urls_input or []) if str(u).strip()]
    seen_urls = set()
    urls = [u for u in urls if not (u in seen_urls or seen_urls.add(u))]

    if not urls:
        return {
            "error": "At least one URL is required.",
            "rules": rules,
            "count": 0,
            "drop_pct": drop_pct,
            "price_drops": [],
            "comparison": build_session_comparison(None, []),
            "using_curl": False,
            "using_browser": False,
            "using_parallel": False,
            "engines_used": {},
            "enrich_details": enrich_details,
            "enrich_details_requested": enrich_details,
            "auto_enrich_details": False,
            "details_hydrated_from_history": 0,
            "details_enriched": 0,
            "items": [],
            "history_id": "",
            "history_public_id": "",
            "history_saved": False,
            "urls": [],
        }

    effective_browser_mode = bool(use_browser) if use_browser is not None else False

    previous_history = previous_history_override if previous_history_override is not None else db_manager.get_latest_history_for_urls(urls)
    items: List[Item] = []
    engine_used: Dict[str, str] = {}
    target_fetch_errors: List[Dict[str, str]] = []
    using_curl = False
    effective_enrich_details = bool(enrich_details)
    runtime_metadata_cache: Dict[str, dict] = {}
    runtime_metadata_cache_lock = threading.Lock()
    progress_lock = threading.Lock()
    progress_state = {
        'completed_targets': 0,
        'items_found': 0,
        'preview_items': [],
        'checkpoint_items': 0,
    }
    total_targets = len(urls)
    active_run_id_for_progress = int((automation_job or {}).get('_active_run_id') or 0)
    if active_run_id_for_progress > 0:
        try:
            active_run = db_manager.get_automation_run(active_run_id_for_progress) or {}
            active_summary = active_run.get('summary') if isinstance(active_run.get('summary'), dict) else {}
            progress_state['completed_targets'] = max(
                0,
                int(active_summary.get('completed_targets') or active_summary.get('phase1_completed') or 0),
            )
            progress_state['items_found'] = max(
                0,
                int(active_summary.get('current_items') or active_run.get('items_count') or 0),
            )
            progress_state['checkpoint_items'] = max(0, int(active_summary.get('checkpoint_items') or 0))
        except Exception as exc:
            app.logger.debug(f"[automation] Could not seed live progress from existing run: {exc}")

    def _target_label_for(url: str) -> str:
        return target_labels.get(url, '')

    def _count_valid_items(scraped_items) -> int:
        count = 0
        for item in scraped_items or []:
            title = ''
            if hasattr(item, 'title'):
                title = str(getattr(item, 'title') or '').strip()
            elif isinstance(item, dict):
                title = str(item.get('title') or '').strip()
            if title:
                count += 1
        return count

    def _enrich_target_items_if_needed(target_items, sess=None):
        if not effective_enrich_details or not target_items:
            return target_items
        valid_candidates = [it for it in target_items if is_usable_scraped_item(it)]
        if not valid_candidates:
            return target_items

        # 1. Hydrate from runtime in-memory cache first (0 ms)
        with runtime_metadata_cache_lock:
            for it in valid_candidates:
                it_url = getattr(it, 'url', '')
                if it_url in runtime_metadata_cache:
                    meta = runtime_metadata_cache[it_url]
                    if meta.get('sku') and not getattr(it, 'sku', ''):
                        it.sku = meta['sku']
                    if meta.get('description') and not getattr(it, 'description', ''):
                        it.description = meta['description']
                    if meta.get('stock_status') and not getattr(it, 'stock_status', ''):
                        it.stock_status = meta['stock_status']

        # 2. Hydrate remaining from persistent DB metadata cache (0 ms)
        uncached_urls = [
            getattr(it, 'url', '') for it in valid_candidates
            if not (str(getattr(it, 'sku', '') or '').strip() and str(getattr(it, 'description', '') or '').strip())
        ]
        if uncached_urls:
            try:
                cached = db_manager.get_product_metadata_cache(uncached_urls)
                if cached:
                    with runtime_metadata_cache_lock:
                        runtime_metadata_cache.update(cached)
                    for it in valid_candidates:
                        it_url = getattr(it, 'url', '')
                        if it_url in cached:
                            meta = cached[it_url]
                            if meta.get('sku') and not getattr(it, 'sku', ''):
                                it.sku = meta['sku']
                            if meta.get('description') and not getattr(it, 'description', ''):
                                it.description = meta['description']
                            if meta.get('stock_status') and not getattr(it, 'stock_status', ''):
                                it.stock_status = meta['stock_status']
            except Exception as cache_exc:
                app.logger.warning(f"[detail] Cache hydration error: {cache_exc}")
        return target_items

    def _report_progress(target_url: str, scraped_items) -> None:
        if not progress_callback:
            return
        valid_items_count = _count_valid_items(scraped_items)
        preview_items = [
            serialize_scraped_item(item)
            for item in (scraped_items or [])
            if is_usable_scraped_item(item)
        ]
        with progress_lock:
            if preview_items and active_run_id_for_progress > 0:
                persisted_count = db_manager.append_automation_run_items(
                    active_run_id_for_progress,
                    preview_items,
                )
                progress_state['checkpoint_items'] += persisted_count
            progress_state['completed_targets'] += 1
            progress_state['items_found'] += valid_items_count
            if preview_items and len(progress_state['preview_items']) < AUTOMATION_CHECKPOINT_ITEM_LIMIT:
                remaining = AUTOMATION_CHECKPOINT_ITEM_LIMIT - len(progress_state['preview_items'])
                progress_state['preview_items'].extend(preview_items[:remaining])
            progress_callback({
                'completed_targets': progress_state['completed_targets'],
                'total_targets': total_targets,
                'current_items': progress_state['items_found'],
                'last_target_url': str(target_url or ''),
                'last_target_items': valid_items_count,
                'preview_items': list(progress_state['preview_items']),
                'checkpoint_items': progress_state['checkpoint_items'],
            })

    def _make_failed_item(url: str, exc: Exception):
        failed_item = Item(
            url=url,
            site=urlparse(url).hostname or '',
            title='',
            price_value=None,
            price_currency=None,
            price_text=f'parallel_scrape_failed: {exc}',
            discounted_value=None,
            discounted_formatted='',
            original_formatted='',
            source='error',
            image_url='',
        )
        return [failed_item]

    def _scrape_url_batch(
        batch_urls,
        *,
        engine_name: str,
        build_session_fn,
        scrape_url_fn,
        uses_curl: bool = True,
        max_workers: int = 12,
        effective_delay_ms: int | None = None,
        force_sequential: bool = False,
        stop_on_block: bool = False,
    ):
        nonlocal using_curl
        if not batch_urls:
            return

        app.logger.info(f"[engine] Using {engine_name} for {len(batch_urls)} URL(s)")

        def _scrape_single(url: str):
            with browser_fetch_mode(effective_browser_mode):
                if uses_curl:
                    sess, local_using_curl = build_session_fn(retries=retries, verify_ssl=verify_ssl, use_curl=True)
                else:
                    sess, local_using_curl = build_session_fn(retries=retries, verify_ssl=verify_ssl)
                scraped_items = scrape_url_fn(sess, url, rules, crawl_pagination, max_pages, effective_delay_ms if effective_delay_ms is not None else delay_ms, app.logger)
            annotate_items_with_target(scraped_items, url, _target_label_for(url), automation_job)
            _enrich_target_items_if_needed(scraped_items, sess)
            blocked = bool(getattr(sess, 'xcell_blocked', False) or getattr(sess, 'gadgetfix_blocked', False))
            last_error = str(
                getattr(sess, 'xcell_last_error', '')
                or getattr(sess, 'gadgetfix_last_error', '')
                or ''
            )
            diagnostics = {
                'xcell_incomplete': bool(getattr(sess, 'xcell_incomplete', False)),
                'xcell_page_stats': list(getattr(sess, 'xcell_page_stats', []) or []),
            }
            return url, scraped_items, local_using_curl, blocked, last_error, diagnostics

        if use_parallel and not force_sequential and len(batch_urls) > 1:
            executor = ThreadPoolExecutor(max_workers=min(max_workers, len(batch_urls)))
            future_to_url = {executor.submit(_scrape_single, url): url for url in batch_urls}
            try:
                for future in as_completed(future_to_url):
                    url = future_to_url[future]
                    try:
                        source_url, scraped_items, local_using_curl, _blocked, _last_error, _diagnostics = future.result()
                        using_curl = using_curl or bool(local_using_curl)
                        items.extend(scraped_items)
                        _report_progress(source_url, scraped_items)
                        if _diagnostics.get('xcell_incomplete'):
                            target_fetch_errors.append({
                                'url': source_url,
                                'engine': engine_name,
                                'error': _last_error or 'XCell pagination did not finish cleanly.',
                                'diagnostics': _diagnostics,
                            })
                        elif _last_error and not _count_valid_items(scraped_items):
                            target_fetch_errors.append({
                                'url': source_url,
                                'engine': engine_name,
                                'error': _last_error,
                            })
                        elif not _count_valid_items(scraped_items):
                            target_fetch_errors.append({
                                'url': source_url,
                                'engine': engine_name,
                                'error': 'No usable products were returned for this target.',
                            })
                        app.logger.info(f"[engine] Completed scraping {source_url}: {_count_valid_items(scraped_items)} items")
                    except Exception as exc:
                        app.logger.error(f"[engine] Error scraping {url}: {exc}")
                        target_fetch_errors.append({
                            'url': url,
                            'engine': engine_name,
                            'error': str(exc),
                        })
                        failed_items = _make_failed_item(url, exc)
                        annotate_items_with_target(failed_items, url, _target_label_for(url), automation_job)
                        items.extend(failed_items)
                        _report_progress(url, [])
            except AutomationRunPaused:
                for pending_future in future_to_url:
                    pending_future.cancel()
                executor.shutdown(wait=False, cancel_futures=True)
                raise
            else:
                executor.shutdown(wait=True)
        else:
            for url in batch_urls:
                try:
                    source_url, scraped_items, local_using_curl, blocked, last_error, diagnostics = _scrape_single(url)
                    using_curl = using_curl or bool(local_using_curl)
                    items.extend(scraped_items)
                    _report_progress(source_url, scraped_items)
                    if diagnostics.get('xcell_incomplete'):
                        target_fetch_errors.append({
                            'url': source_url,
                            'engine': engine_name,
                            'error': last_error or 'XCell pagination did not finish cleanly.',
                            'diagnostics': diagnostics,
                        })
                    elif last_error and not _count_valid_items(scraped_items):
                        target_fetch_errors.append({
                            'url': source_url,
                            'engine': engine_name,
                            'error': last_error,
                        })
                    elif not _count_valid_items(scraped_items):
                        target_fetch_errors.append({
                            'url': source_url,
                            'engine': engine_name,
                            'error': 'No usable products were returned for this target.',
                        })
                    app.logger.info(f"[engine] Completed scraping {source_url}: {_count_valid_items(scraped_items)} items")
                    if blocked and stop_on_block:
                        app.logger.warning(
                            f"[engine] Stopping {engine_name} batch after site block: {last_error or source_url}"
                        )
                        break
                except Exception as exc:
                    app.logger.error(f"[engine] Error scraping {url}: {exc}")
                    target_fetch_errors.append({
                        'url': url,
                        'engine': engine_name,
                        'error': str(exc),
                    })
                    failed_items = _make_failed_item(url, exc)
                    annotate_items_with_target(failed_items, url, _target_label_for(url), automation_job)
                    items.extend(failed_items)
                    _report_progress(url, [])

    xcell_urls = []
    txparts_urls = []
    parts4cells_urls = []
    phonelcdparts_urls = []
    gadgetfix_urls = []
    standard_urls = []

    for url in urls:
        engine_type, _ = get_scraper_for_url(url)
        if engine_type == 'xcell':
            xcell_urls.append(url)
            engine_used[url] = 'xcell_scraper_engine'
        elif engine_type == 'txparts':
            txparts_urls.append(url)
            engine_used[url] = 'txparts_scraper_engine'
        elif engine_type == 'parts4cells':
            parts4cells_urls.append(url)
            engine_used[url] = 'parts4cells_scraper_engine'
        elif engine_type == 'phonelcdparts':
            phonelcdparts_urls.append(url)
            engine_used[url] = 'phonelcdparts_scraper_engine'
        elif engine_type == 'gadgetfix':
            gadgetfix_urls.append(url)
            engine_used[url] = 'gadgetfix_scraper_engine'
        else:
            standard_urls.append(url)
            engine_used[url] = 'scraper_engine'

    try:
        xcell_max_workers = max(1, min(24, int(os.getenv('XCELL_MAX_WORKERS') or os.getenv('SCRAPER_XCELL_MAX_WORKERS') or '16')))
    except (TypeError, ValueError):
        xcell_max_workers = 16

    _scrape_url_batch(
        xcell_urls,
        engine_name='XCellParts scraper',
        build_session_fn=xcell_scraper_engine.build_session,
        scrape_url_fn=xcell_scraper_engine.scrape_url,
        uses_curl=True,
        max_workers=xcell_max_workers,
        effective_delay_ms=delay_ms,
        force_sequential=False,
        stop_on_block=False,
    )
    _scrape_url_batch(
        txparts_urls,
        engine_name='TXParts scraper',
        build_session_fn=txparts_scraper_engine.build_session,
        scrape_url_fn=txparts_scraper_engine.scrape_url,
        uses_curl=True,
        max_workers=16,
    )
    _scrape_url_batch(
        parts4cells_urls,
        engine_name='Parts4Cells scraper',
        build_session_fn=parts4cells_scraper_engine.build_session,
        scrape_url_fn=parts4cells_scraper_engine.scrape_url,
        uses_curl=True,
        max_workers=16,
    )
    _scrape_url_batch(
        phonelcdparts_urls,
        engine_name='PhoneLCDParts scraper',
        build_session_fn=phonelcdparts_scraper_engine.build_session,
        scrape_url_fn=phonelcdparts_scraper_engine.scrape_url,
        uses_curl=True,
        max_workers=16,
    )
    _scrape_url_batch(
        gadgetfix_urls,
        engine_name='GadgetFix scraper',
        build_session_fn=gadgetfix_scraper_engine.build_session,
        scrape_url_fn=gadgetfix_scraper_engine.scrape_url,
        uses_curl=True,
        max_workers=16,
        force_sequential=False,
        stop_on_block=False,
    )
    _scrape_url_batch(
        standard_urls,
        engine_name='standard scraper',
        build_session_fn=build_session,
        scrape_url_fn=scrape_url,
        uses_curl=True,
        max_workers=16,
    )

    items = [item for item in items if is_usable_scraped_item(item)]

    guard_anomalies = detect_sparse_target_anomalies(urls, items, previous_history, engine_used)
    if guard_anomalies:
        incident = create_scraper_guard_incident(
            guard_anomalies,
            urls=urls,
            automation_job=automation_job,
            previous_history=previous_history,
            items_count=len(items),
        )
        error_text = (
            f"Scraper data-quality guard stopped this run before saving history. "
            f"{len(guard_anomalies)} target(s) returned suspiciously few products. "
            f"Incident: {incident.get('id')}"
        )
        app.logger.error(f"[scraper-guard] {error_text}")
        return {
            "error": error_text,
            "rules": rules,
            "count": len(items),
            "drop_pct": drop_pct,
            "price_drops": [],
            "comparison": build_session_comparison(
                previous_history,
                [],
                current_target_urls=urls,
                run_validation={
                    'approved': False,
                    'status': 'Rejected by Validation',
                    'reasons': ['Sparse target anomaly guard stopped the run.'],
                },
                target_errors=target_fetch_errors,
            ),
            "using_curl": using_curl,
            "using_browser": effective_browser_mode,
            "using_parallel": use_parallel and len(urls) > 1,
            "engines_used": engine_used,
            "enrich_details": enrich_details,
            "enrich_details_requested": enrich_details,
            "auto_enrich_details": False,
            "details_hydrated_from_history": 0,
            "details_enriched": 0,
            "items": [serialize_scraped_item(i) for i in items],
            "history_id": "",
            "history_public_id": "",
            "history_saved": False,
            "urls": urls,
            "target_errors": target_fetch_errors,
            "guard_incident": incident,
            "guard_anomalies": guard_anomalies,
        }

    run_validation = validate_scrape_completeness(urls, items, previous_history, target_fetch_errors)
    if not run_validation.get('approved', True):
        validation_anomalies = list(run_validation.get('target_anomalies') or [])
        if not validation_anomalies:
            validation_anomalies = [
                {
                    'type': 'run_completeness_rejected',
                    'url': urls[0] if urls else '',
                    'engine': detect_scraper_key(urls[0]) if urls else '',
                    'previous_count': run_validation.get('metrics', {}).get('previous_items', 0),
                    'current_count': run_validation.get('metrics', {}).get('current_items', 0),
                    'reason': '; '.join(run_validation.get('reasons') or []),
                }
            ]
        incident = create_scraper_guard_incident(
            validation_anomalies,
            urls=urls,
            automation_job=automation_job,
            previous_history=previous_history,
            items_count=len(items),
        )
        error_text = (
            "Scraper completeness validation stopped this run before saving history. "
            + ' '.join(str(reason) for reason in run_validation.get('reasons', []))
            + f" Incident: {incident.get('id')}"
        )
        app.logger.error(f"[scraper-guard] {error_text}")
        return {
            "error": error_text,
            "rules": rules,
            "count": len(items),
            "drop_pct": drop_pct,
            "price_drops": [],
            "comparison": build_session_comparison(
                previous_history,
                items,
                current_target_urls=urls,
                run_validation=run_validation,
                target_errors=target_fetch_errors,
            ),
            "run_validation": run_validation,
            "using_curl": using_curl,
            "using_browser": effective_browser_mode,
            "using_parallel": use_parallel and len(urls) > 1,
            "engines_used": engine_used,
            "enrich_details": enrich_details,
            "enrich_details_requested": enrich_details,
            "auto_enrich_details": False,
            "details_hydrated_from_history": 0,
            "details_enriched": 0,
            "items": [serialize_scraped_item(i) for i in items],
            "history_id": "",
            "history_public_id": "",
            "history_saved": False,
            "urls": urls,
            "target_errors": target_fetch_errors,
            "guard_incident": incident,
            "guard_anomalies": validation_anomalies,
        }

    hydrated_from_history = hydrate_items_from_previous_history(items, previous_history)
    auto_enrich_details = False
    effective_enrich_details = enrich_details or auto_enrich_details
    if auto_enrich_details:
        app.logger.info(f"[detail] Auto-enabling detail scan for {len(items)} item(s) to capture stock detail")
    items, enriched_count = enrich_scraped_items(
        items, rules, retries, verify_ssl, use_curl, enrich_details=effective_enrich_details, logger=app.logger, use_browser=use_browser, progress_callback=progress_callback
    )
    if progress_callback:
        progress_callback({
            'phase': 3,
            'phase_name': 'Phase 3: Validation & Comparison',
            'activity_label': 'Validating data',
            'status_message': 'Checking scrape quality and preparing product comparison.',
            'completed_targets': len(urls),
            'total_targets': len(urls),
            'current_items': len(items),
            'phase2_completed': enriched_count if effective_enrich_details else len(items),
            'phase2_total': len(items) if effective_enrich_details else 0,
            'progress_percent': 96.0,
        })

    previous_prices = {}
    if previous_history:
        for previous_item in previous_history.get('items', []):
            snapshot = normalize_item_snapshot(previous_item)
            url = snapshot.get('url')
            price = snapshot.get('comparison_price')
            if not url or price is None:
                continue
            previous_prices[url] = {
                'price': price,
                'title': snapshot.get('title', ''),
                'site': snapshot.get('site', ''),
                'history_id': previous_history.get('id'),
                'timestamp': previous_history.get('timestamp'),
            }
    price_drops = build_price_drop_alerts(items, previous_prices, drop_pct)
    comparison = build_session_comparison(
        previous_history,
        items,
        current_target_urls=urls,
        run_validation=run_validation,
        target_errors=target_fetch_errors,
    )
    if progress_callback:
        progress_callback({
            'phase': 4,
            'phase_name': 'Phase 4: Saving Snapshot',
            'activity_label': 'Saving snapshot',
            'status_message': 'Writing scraped products, comparison metadata, and run history to the database.',
            'completed_targets': len(urls),
            'total_targets': len(urls),
            'current_items': len(items),
            'phase2_completed': len(items),
            'phase2_total': len(items),
            'progress_percent': 99.0,
        })

    history_id = str(int(time.time() * 1000))
    history_saved = False
    history_public_id = ''
    pruned_history_ids = []
    history_rules = dict(rules)
    if automation_job:
        history_rules['_automation_job_id'] = automation_job.get('id')
        history_rules['_automation_job_name'] = automation_job.get('name')
        history_rules['_automation_category_query'] = automation_job.get('category_query')
    scraper_keys_for_log = sorted({detect_scraper_key(url) for url in urls if str(url or '').strip()})

    if items:
        try:
            history_saved = db_manager.save_fetch_history(history_id, urls, items, history_rules)
            app.logger.info(
                f"[engine] Save result history_id={history_id} saved={history_saved} "
                f"item_count={len(items)} url_count={len(urls)} scraper_keys={','.join(scraper_keys_for_log)}"
            )
            if not history_saved:
                app.logger.error("Failed to save fetch history to database")
        except Exception as exc:
            app.logger.error(f"Database error: {exc}")
    else:
        app.logger.warning(
            f"[engine] No usable products scraped; history was not saved url_count={len(urls)} "
            f"scraper_keys={','.join(scraper_keys_for_log)}"
        )

    scraper_keys = {detect_scraper_key(url) for url in urls if str(url or '').strip()}
    if history_saved and len(scraper_keys) == 1:
        history_public_id = build_public_history_id(next(iter(scraper_keys)), history_id)
    if history_saved:
        keep_histories = history_retention_keep_count()
        if keep_histories > 0:
            try:
                pruned_history_ids = db_manager.prune_histories_for_urls(urls, keep=keep_histories)
                if pruned_history_ids:
                    app.logger.info(
                        f"[engine] Pruned {len(pruned_history_ids)} old history record(s) "
                        f"after saving history_id={history_public_id or history_id}: {', '.join(pruned_history_ids[:10])}"
                    )
            except Exception as exc:
                app.logger.error(f"[engine] Failed to prune old histories after save: {exc}")

    return {
        "rules": rules,
        "count": len(items),
        "drop_pct": drop_pct,
        "price_drops": price_drops,
        "comparison": comparison,
        "run_validation": run_validation,
        "using_curl": using_curl,
        "using_browser": effective_browser_mode,
        "using_parallel": use_parallel and len(urls) > 1,
        "engines_used": engine_used,
        "enrich_details": effective_enrich_details,
        "enrich_details_requested": enrich_details,
        "auto_enrich_details": auto_enrich_details,
        "details_hydrated_from_history": hydrated_from_history,
        "details_enriched": enriched_count,
        "items": [serialize_scraped_item(i) for i in items],
        "history_id": history_id,
        "history_public_id": history_public_id,
        "history_saved": history_saved,
        "pruned_history_ids": pruned_history_ids,
        "urls": urls,
        "target_errors": target_fetch_errors,
    }


def _launch_automation_job(job_id: int, trigger_type: str = 'schedule') -> Tuple[bool, str]:
    try:
        normalized_job_id = int(job_id)
    except (TypeError, ValueError):
        return False, 'Invalid automation job id.'

    if db_manager.get_active_automation_run_for_job(normalized_job_id):
        return False, 'This automation job is already running.'

    with AUTOMATION_ACTIVE_JOBS_LOCK:
        if normalized_job_id in AUTOMATION_ACTIVE_JOBS:
            return False, 'This automation job is already running.'
        AUTOMATION_ACTIVE_JOBS.add(normalized_job_id)

    def worker():
        run_record = None
        try:
            job = db_manager.get_automation_job(normalized_job_id, include_targets=True)
            if not job:
                return

            targets = [target for target in job.get('targets', []) if target.get('active', True)]
            if job.get('auto_discover') or not targets:
                discovered = discover_category_targets_via_browser(
                    job.get('scraper_key'),
                    job.get('category_query'),
                    root_url=job.get('root_url'),
                    retries=job.get('retries', 1),
                    verify_ssl=job.get('verify_ssl', True),
                    logger=app.logger,
                )
                targets = db_manager.replace_automation_job_targets(normalized_job_id, discovered.get('targets', []))
                job = db_manager.get_automation_job(normalized_job_id, include_targets=True) or job

            active_targets = [target for target in targets if target.get('active', True)]
            target_urls = [str(target.get('url') or '').strip() for target in active_targets if str(target.get('url') or '').strip()]
            if not target_urls:
                run_record = db_manager.create_automation_run(normalized_job_id, trigger_type=trigger_type, target_urls=[])
                if run_record:
                    db_manager.complete_automation_run(
                        run_record['id'],
                        status='failed',
                        target_urls=[],
                        items_count=0,
                        summary={'target_count': 0},
                        error_text='No category links were discovered for this automation job.',
                    )
                return
            try:
                target_urls = validate_supplier_remote_urls(target_urls, job.get('scraper_key'))
            except ValueError as exc:
                run_record = db_manager.create_automation_run(
                    normalized_job_id,
                    trigger_type=trigger_type,
                    target_urls=target_urls,
                )
                if run_record:
                    db_manager.complete_automation_run(
                        run_record['id'],
                        status='failed',
                        target_urls=target_urls,
                        items_count=0,
                        summary={'target_count': len(target_urls), 'resume_available': True},
                        error_text=str(exc),
                    )
                return

            previous_history = None
            previous_history_id = ''
            last_history_ids = job.get('last_history_ids') or []
            if last_history_ids:
                previous_history_id = str(last_history_ids[0] or '').strip()
                previous_history = db_manager.get_history_detail(previous_history_id)

            run_record = db_manager.create_automation_run(
                normalized_job_id,
                trigger_type=trigger_type,
                target_urls=target_urls,
                previous_history_id=previous_history_id,
            )
            if not run_record:
                return

            total_target_count = len(target_urls)
            db_manager.update_automation_run_progress(
                run_record['id'],
                items_count=0,
                summary={
                    'target_count': total_target_count,
                    'completed_targets': 0,
                    'total_targets': total_target_count,
                    'current_items': 0,
                    'progress_percent': 0,
                    'last_target_url': '',
                    'last_target_items': 0,
                    'recent_targets_per_min': 0,
                    'recent_items_per_min': 0,
                    'recent_rate_window_seconds': 600,
                },
            )

            recent_progress = []
            recent_enrich_progress = []

            def automation_progress_callback(progress: Dict[str, object]):
                latest_run = db_manager.get_automation_run(run_record['id'])
                if str((latest_run or {}).get('status') or '').strip().lower() == 'paused':
                    raise AutomationRunPaused('Automation run paused by user.')
                latest_summary = (latest_run or {}).get('summary') if isinstance((latest_run or {}).get('summary'), dict) else {}
                now = time.time()
                current_phase = int(progress.get('phase') or (2 if progress.get('phase2_total') else 1))
                phase_name = str(progress.get('phase_name') or ('Phase 2: Product SKU & Detail Scan' if current_phase == 2 else 'Phase 1: Category Crawling'))
                completed_targets = int(progress.get('completed_targets') if progress.get('completed_targets') is not None else latest_summary.get('completed_targets') or 0)
                total_targets_local = max(1, int(progress.get('total_targets') if progress.get('total_targets') is not None else latest_summary.get('total_targets') or total_target_count or 1))
                current_items = int(progress.get('current_items') if progress.get('current_items') is not None else latest_summary.get('current_items') or 0)
                checkpoint_only_phase1 = current_phase == 1 and current_items > 0 and completed_targets <= 0
                if checkpoint_only_phase1 and phase_name == 'Phase 1: Category Crawling':
                    phase_name = 'Collecting Product Checkpoint'
                last_target_items = int(progress.get('last_target_items') or 0)
                preview_items = progress.get('preview_items') if isinstance(progress.get('preview_items'), list) else []
                if not preview_items and isinstance(latest_summary.get('preview_items'), list):
                    preview_items = latest_summary.get('preview_items') or []
                checkpoint_items = int(progress.get('checkpoint_items') if progress.get('checkpoint_items') is not None else latest_summary.get('checkpoint_items') or 0)

                # Phase 2 metrics
                phase2_completed = int(progress.get('phase2_completed') if progress.get('phase2_completed') is not None else latest_summary.get('phase2_completed') or 0)
                phase2_total = int(progress.get('phase2_total') if progress.get('phase2_total') is not None else latest_summary.get('phase2_total') or 0)
                if current_phase == 2 and phase2_completed > 0:
                    phase2_total = max(phase2_total, phase2_completed, current_items)

                cutoff = now - 10 * 60
                if current_phase == 2:
                    recent_enrich_progress.append((now, phase2_completed))
                    while len(recent_enrich_progress) > 1 and recent_enrich_progress[0][0] < cutoff:
                        recent_enrich_progress.pop(0)
                else:
                    recent_progress.append((now, completed_targets, current_items))
                    while len(recent_progress) > 1 and recent_progress[0][0] < cutoff:
                        recent_progress.pop(0)

                recent_targets_per_min = float(latest_summary.get('recent_targets_per_min') or 0.0)
                recent_items_per_min = float(latest_summary.get('recent_items_per_min') or 0.0)
                if len(recent_progress) >= 2:
                    first_time, first_targets, first_items = recent_progress[0]
                    elapsed_minutes = max((now - first_time) / 60.0, 0.001)
                    recent_targets_per_min = max(0.0, (completed_targets - first_targets) / elapsed_minutes)
                    recent_items_per_min = max(0.0, (current_items - first_items) / elapsed_minutes)
                if len(recent_enrich_progress) >= 2:
                    first_time, first_enriched = recent_enrich_progress[0]
                    elapsed_minutes = max((now - first_time) / 60.0, 0.001)
                    recent_items_per_min = max(0.0, (phase2_completed - first_enriched) / elapsed_minutes)

                phase1_eta_min = round((total_targets_local - completed_targets) / max(recent_targets_per_min, 0.5), 1) if completed_targets < total_targets_local else 0.0
                phase2_eta_min = round((phase2_total - phase2_completed) / max(recent_items_per_min, 25.0), 1) if phase2_total > phase2_completed else 0.0

                if progress.get('progress_percent') is not None:
                    try:
                        progress_percent = max(0.0, min(100.0, round(float(progress.get('progress_percent')), 1)))
                    except (TypeError, ValueError):
                        progress_percent = 100.0
                else:
                    progress_percent = round((completed_targets / total_targets_local) * 100, 1) if current_phase == 1 else (
                        round((phase2_completed / max(1, phase2_total)) * 100, 1) if current_phase == 2 else 100.0
                    )
                if checkpoint_only_phase1 and progress_percent == 0:
                    progress_percent = 1.0
                if current_phase == 2 and phase2_total > 0:
                    progress_percent = max(progress_percent, round((phase2_completed / phase2_total) * 100, 1))

                db_manager.update_automation_run_progress(
                    run_record['id'],
                    items_count=current_items,
                    summary={
                        'phase': current_phase,
                        'phase_name': phase_name,
                        'target_count': total_target_count,
                        'completed_targets': completed_targets,
                        'total_targets': total_targets_local,
                        'current_items': current_items,
                        'progress_percent': progress_percent,
                        'last_target_url': str(progress.get('last_target_url') or ''),
                        'last_target_items': last_target_items,
                        'preview_items': preview_items[:AUTOMATION_CHECKPOINT_ITEM_LIMIT],
                        'checkpoint_items': checkpoint_items,
                        'recent_targets_per_min': round(recent_targets_per_min, 2),
                        'recent_items_per_min': round(recent_items_per_min, 2),
                        'recent_rate_window_seconds': 600,
                        'phase1_completed': completed_targets,
                        'phase1_total': total_targets_local,
                        'phase1_speed': f"{recent_targets_per_min:.1f} cats/min" if recent_targets_per_min > 0 else "~45 cats/min",
                        'phase1_eta': f"{phase1_eta_min:.1f}m" if phase1_eta_min > 0 else "0m",
                        'phase2_completed': phase2_completed,
                        'phase2_total': phase2_total,
                        'phase2_speed': f"{recent_items_per_min:.0f} items/min" if recent_items_per_min > 0 else "~440 items/min",
                        'phase2_eta': f"{phase2_eta_min:.1f}m" if phase2_eta_min > 0 else "0m",
                        'status_message': str(progress.get('status_message') or ''),
                        'activity_label': str(progress.get('activity_label') or phase_name),
                    },
                )

            target_labels = {str(target.get('url') or '').strip(): str(target.get('label') or '').strip() for target in active_targets}
            workflow_job = dict(job)
            workflow_job['_active_run_id'] = run_record['id']
            result = execute_scrape_workflow(
                target_urls,
                crawl_pagination=job.get('crawl_pagination', True),
                max_pages=job.get('max_pages', 10),
                delay_ms=job.get('delay_ms', 50),
                retries=job.get('retries', 1),
                verify_ssl=job.get('verify_ssl', True),
                use_curl=True,
                use_browser=False,
                use_parallel=job.get('use_parallel', True),
                enrich_details=bool(job.get('enrich_details', False)),
                rules=job.get('rules', {}),
                drop_pct=job.get('drop_pct', 10.0),
                target_labels=target_labels,
                automation_job=workflow_job,
                previous_history_override=previous_history,
                progress_callback=automation_progress_callback,
            )
            summary = build_automation_run_summary(target_urls, result.get('comparison') or {}, result.get('price_drops') or [])
            latest_run_for_summary = db_manager.get_automation_run(run_record['id']) or {}
            latest_progress_summary = latest_run_for_summary.get('summary') if isinstance(latest_run_for_summary.get('summary'), dict) else {}
            if latest_progress_summary:
                summary = {
                    **latest_progress_summary,
                    **summary,
                    'phase': 5,
                    'phase_name': 'Completed',
                    'activity_label': 'Snapshot saved',
                    'status_message': 'Run finished and the saved snapshot is ready.',
                    'progress_percent': 100.0,
                }
            result_count = int(result.get('count') or 0)
            workflow_error = str(result.get('error') or '').strip()
            no_items_error = ''
            if target_urls and result_count == 0:
                target_errors = result.get('target_errors') or []
                error_detail = ''
                if target_errors:
                    first_error = target_errors[0] or {}
                    error_detail = f" First fetch error: {first_error.get('error') or first_error}."
                no_items_error = (
                    'No products were scraped from the selected category targets. '
                    'The supplier may be blocking requests, the categories may be empty, or the target URLs may need review.'
                    f'{error_detail}'
                )
            current_history_id = result.get('history_public_id') or ''
            previous_result_history_id = (result.get('comparison') or {}).get('previous_history_id') or ''
            error_text = workflow_error or no_items_error
            if error_text and not current_history_id and result_count > 0:
                partial_history_id, partial_count, partial_summary = save_automation_partial_history(
                    run_record['id'],
                    job,
                    target_urls,
                    previous_history=previous_history,
                    error_text=error_text,
                )
                if partial_history_id:
                    current_history_id = partial_history_id
                    result_count = partial_count
                    summary = partial_summary
            db_manager.complete_automation_run(
                run_record['id'],
                status='failed' if error_text else 'completed',
                current_history_id=current_history_id,
                previous_history_id=previous_result_history_id or previous_history_id,
                target_urls=target_urls,
                items_count=result_count,
                summary=summary,
                error_text=error_text,
            )
        except AutomationRunPaused as exc:
            app.logger.info(f"[automation] Job {normalized_job_id} paused: {exc}")
            if run_record:
                db_manager.pause_automation_run(run_record['id'], reason=str(exc) or 'Automation run paused.')
        except Exception as exc:
            app.logger.exception(f"[automation] Job {normalized_job_id} failed: {exc}")
            if run_record:
                partial_history_id, partial_count, partial_summary = save_automation_partial_history(
                    run_record['id'],
                    job if 'job' in locals() else None,
                    run_record.get('target_urls') or [],
                    previous_history=previous_history if 'previous_history' in locals() else None,
                    error_text=str(exc),
                )
                db_manager.complete_automation_run(
                    run_record['id'],
                    status='failed',
                    current_history_id=partial_history_id,
                    previous_history_id=previous_history_id if 'previous_history_id' in locals() else '',
                    target_urls=run_record.get('target_urls') or [],
                    items_count=partial_count,
                    summary=partial_summary or {'target_count': len(run_record.get('target_urls') or [])},
                    error_text=str(exc),
                )
        finally:
            with AUTOMATION_ACTIVE_JOBS_LOCK:
                AUTOMATION_ACTIVE_JOBS.discard(normalized_job_id)
            db_manager.close_connection()

    threading.Thread(target=worker, name=f'automation-job-{normalized_job_id}', daemon=True).start()
    return True, ''


def _launch_existing_automation_run(run_id: int) -> Tuple[bool, str]:
    try:
        normalized_run_id = int(run_id)
    except (TypeError, ValueError):
        return False, 'Invalid automation run id.'

    run = db_manager.get_automation_run(normalized_run_id)
    if not run:
        return False, 'Automation run not found.'

    status = str(run.get('status') or '').strip().lower()
    if status in {'running', 'resuming'}:
        return False, 'This automation run is already running.'
    if status == 'completed':
        return False, 'Completed runs already have a saved snapshot. Start a new run only when you want a fresh comparison.'

    script_path = APP_ROOT / 'scripts' / 'resume_automation_run.py'
    if not script_path.exists():
        return False, 'Resume helper script is missing.'

    tmp_dir = APP_ROOT / '.tmp'
    tmp_dir.mkdir(parents=True, exist_ok=True)
    stdout_path = tmp_dir / f'resume-run-{normalized_run_id}.out.log'
    stderr_path = tmp_dir / f'resume-run-{normalized_run_id}.err.log'
    env = os.environ.copy()
    env['RESUME_FROM_CHECKPOINT'] = '1'
    env.setdefault('AUTOMATION_SCHEDULER_DISABLED', '1')
    env.setdefault('AUTOMATION_RECOVER_RUNNING', '0')

    if str(run.get('scraper_key') or '').strip().lower() == 'xcell':
        env['XCELL_MAX_WORKERS'] = '6'
        env.setdefault('SCRAPER_USE_BROWSER', '0')

    try:
        claimed_run = db_manager.claim_automation_run_resume(normalized_run_id)
        if not claimed_run:
            return False, 'This automation run is already running or is no longer resumable.'
        with open(stdout_path, 'ab') as stdout_handle, open(stderr_path, 'ab') as stderr_handle:
            subprocess.Popen(
                [sys.executable, '-u', str(script_path), str(normalized_run_id)],
                cwd=str(APP_ROOT),
                env=env,
                stdout=stdout_handle,
                stderr=stderr_handle,
                creationflags=getattr(subprocess, 'CREATE_NO_WINDOW', 0),
            )
        return True, ''
    except Exception as exc:
        db_manager.fail_automation_run_resume_launch(normalized_run_id, str(exc))
        return False, str(exc)


def _automation_scheduler_loop():
    try:
        while not AUTOMATION_STOP_EVENT.wait(AUTOMATION_POLL_INTERVAL_SECONDS):
            try:
                for job in db_manager.get_due_automation_jobs(limit=5):
                    recent_runs = db_manager.list_automation_runs(job_id=job.get('id'), limit=1)
                    latest_status = str((recent_runs[0] if recent_runs else {}).get('status') or '').strip().lower()
                    if latest_status in {'paused', 'interrupted', 'failed', 'resuming'}:
                        app.logger.info(
                            '[automation] Skipping scheduled job %s until run %s is explicitly resumed.',
                            job.get('id'),
                            recent_runs[0].get('id'),
                        )
                        continue
                    _launch_automation_job(job.get('id'), trigger_type='schedule')
            except Exception as exc:
                app.logger.exception(f"[automation] Scheduler error: {exc}")
    finally:
        db_manager.close_connection()


def ensure_automation_scheduler_started():
    global AUTOMATION_SCHEDULER_STARTED, AUTOMATION_SCHEDULER_THREAD
    scheduler_disabled = str(os.getenv("AUTOMATION_SCHEDULER_DISABLED", "0")).strip().lower() in {"1", "true", "yes", "on"}
    if scheduler_disabled:
        app.logger.info("[automation] Scheduler disabled by AUTOMATION_SCHEDULER_DISABLED")
        return
    debug_env = str(os.getenv("FLASK_DEBUG", "0")).strip().lower()
    debug_mode = debug_env not in {"0", "false", "no", "off"}
    if debug_mode and os.environ.get('WERKZEUG_RUN_MAIN') != 'true':
        return
    if AUTOMATION_SCHEDULER_STARTED:
        return
    with AUTOMATION_SCHEDULER_LOCK:
        if AUTOMATION_SCHEDULER_STARTED:
            return
        recover_enabled = str(os.getenv("AUTOMATION_RECOVER_RUNNING", "1")).strip().lower() not in {"0", "false", "no", "off"}
        if recover_enabled:
            recovered_runs = db_manager.recover_running_automation_runs()
            if recovered_runs:
                app.logger.warning(f"[automation] Recovered {recovered_runs} interrupted automation run(s) from a previous process")
        else:
            app.logger.info("[automation] Interrupted-run recovery disabled by AUTOMATION_RECOVER_RUNNING")
        AUTOMATION_STOP_EVENT.clear()
        AUTOMATION_SCHEDULER_THREAD = threading.Thread(target=_automation_scheduler_loop, name='automation-scheduler', daemon=True)
        AUTOMATION_SCHEDULER_THREAD.start()
        AUTOMATION_SCHEDULER_STARTED = True
        app.logger.info("[automation] Scheduler started")


def shutdown_background_services():
    """Signal background workers to stop and close this thread's database connection."""
    global AUTOMATION_SCHEDULER_STARTED
    AUTOMATION_STOP_EVENT.set()
    thread = AUTOMATION_SCHEDULER_THREAD
    if thread and thread.is_alive() and thread is not threading.current_thread():
        thread.join(timeout=3)
    AUTOMATION_SCHEDULER_STARTED = False
    db_manager.close_connection()


def _handle_shutdown_signal(signum, frame):
    shutdown_background_services()
    raise SystemExit(0)


def register_shutdown_hooks():
    global SHUTDOWN_HOOKS_REGISTERED
    if SHUTDOWN_HOOKS_REGISTERED:
        return
    atexit.register(shutdown_background_services)
    for shutdown_signal in (getattr(signal, 'SIGINT', None), getattr(signal, 'SIGTERM', None)):
        if shutdown_signal is None:
            continue
        try:
            signal.signal(shutdown_signal, _handle_shutdown_signal)
        except (ValueError, OSError):
            pass
    SHUTDOWN_HOOKS_REGISTERED = True






def discover_category_targets_via_browser(
    scraper_key: str,
    category_query: str,
    *,
    root_url: str = '',
    retries: int = 1,
    verify_ssl: bool = True,
    logger=None,
) -> Dict[str, object]:
    """Discover categories through the single supported browser engine."""
    with browser_fetch_mode(True):
        discovered = discover_category_targets(
            scraper_key,
            category_query,
            root_url=root_url,
            retries=retries,
            verify_ssl=verify_ssl,
            logger=logger,
        )
    discovered['using_browser'] = True
    discovered['browser_engine'] = 'botasaurus'
    discovered['browser_fallback_used'] = False
    return discovered


def coerce_bool(value, default: bool = False) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    text = str(value).strip().lower()
    if text in {'1', 'true', 'yes', 'on'}:
        return True
    if text in {'0', 'false', 'no', 'off', ''}:
        return False
    return default if text == '' else True


def coerce_int(value, default: int = 0, min_value: int | None = None, max_value: int | None = None) -> int:
    try:
        result = int(value)
    except (TypeError, ValueError):
        result = default
    if min_value is not None:
        result = max(min_value, result)
    if max_value is not None:
        result = min(max_value, result)
    return result


def coerce_float(value, default: float = 0.0, min_value: float | None = None, max_value: float | None = None) -> float:
    try:
        result = float(value)
    except (TypeError, ValueError):
        result = default
    if min_value is not None:
        result = max(min_value, result)
    if max_value is not None:
        result = min(max_value, result)
    return result


def coerce_optional_positive_int(value) -> int:
    try:
        result = int(value)
    except (TypeError, ValueError):
        return 0
    return result if result > 0 else 0


def clean_excel_cell(value):
    if value is None:
        return ''
    if isinstance(value, (int, float, bool)):
        return value
    return ILLEGAL_CHARACTERS_RE.sub('', str(value))


def coerce_string_list(value, *, max_items: int = 20) -> List[str]:
    if not isinstance(value, list):
        return []
    normalized = []
    for item in value[:max_items]:
        text = str(item or '').strip()
        if text:
            normalized.append(text)
    return normalized


def cleanup_proxied_image_cache() -> None:
    """Drop expired proxied images to keep memory bounded."""
    now = time.time()
    with PROXIED_IMAGE_CACHE_LOCK:
        expired = [
            cache_key for cache_key, entry in PROXIED_IMAGE_CACHE.items()
            if now - float(entry.get('created_at', 0)) > PROXIED_IMAGE_TTL_SECONDS
        ]
        for cache_key in expired:
            PROXIED_IMAGE_CACHE.pop(cache_key, None)

        if len(PROXIED_IMAGE_CACHE) <= MAX_PROXIED_IMAGE_CACHE_ITEMS:
            return

        overflow = len(PROXIED_IMAGE_CACHE) - MAX_PROXIED_IMAGE_CACHE_ITEMS
        oldest = sorted(
            PROXIED_IMAGE_CACHE.items(),
            key=lambda item: float(item[1].get('created_at', 0))
        )[:overflow]
        for cache_key, _entry in oldest:
            PROXIED_IMAGE_CACHE.pop(cache_key, None)


def supplier_host_allowed(hostname: str) -> bool:
    normalized = str(hostname or '').strip().lower().rstrip('.')
    return any(
        normalized == allowed or normalized.endswith(f'.{allowed}')
        for allowed in SUPPLIER_REMOTE_HOSTS
    )


def validate_supplier_remote_url(
    raw_url: str,
    *,
    allowed_hosts: Tuple[str, ...] | None = None,
    checked_hosts: set | None = None,
) -> str:
    try:
        parsed = urlparse(str(raw_url or '').strip())
    except ValueError as exc:
        raise ValueError('A valid absolute supplier URL is required.') from exc
    if parsed.scheme not in {'http', 'https'} or not parsed.hostname:
        raise ValueError('Only absolute HTTP and HTTPS URLs are allowed.')
    if parsed.username or parsed.password:
        raise ValueError('URLs containing credentials are not allowed.')
    host_allowlist = allowed_hosts or SUPPLIER_REMOTE_HOSTS
    normalized_hostname = parsed.hostname.lower().rstrip('.')
    if not any(
        normalized_hostname == allowed or normalized_hostname.endswith(f'.{allowed}')
        for allowed in host_allowlist
    ):
        raise ValueError(f'Host is not an approved supplier domain: {parsed.hostname}')

    if checked_hosts is not None and normalized_hostname in checked_hosts:
        return parsed._replace(fragment='').geturl()
    try:
        addresses = socket.getaddrinfo(
            parsed.hostname,
            parsed.port or (443 if parsed.scheme == 'https' else 80),
            type=socket.SOCK_STREAM,
        )
    except socket.gaierror as exc:
        raise ValueError(f'Could not resolve supplier host: {parsed.hostname}') from exc
    if not addresses:
        raise ValueError(f'Could not resolve supplier host: {parsed.hostname}')
    for address in addresses:
        ip_value = ipaddress.ip_address(str(address[4][0]).split('%', 1)[0])
        if not ip_value.is_global:
            raise ValueError('Private, loopback, and link-local network addresses are not allowed.')
    if checked_hosts is not None:
        checked_hosts.add(normalized_hostname)

    return parsed._replace(fragment='').geturl()


def validate_supplier_remote_urls(raw_urls, scraper_key: str = '') -> List[str]:
    config = SCRAPER_CONFIG.get(str(scraper_key or '').strip().lower())
    allowed_hosts = tuple(config.get('domains', ())) if config else SUPPLIER_REMOTE_HOSTS
    checked_hosts = set()
    return [
        validate_supplier_remote_url(
            raw_url,
            allowed_hosts=allowed_hosts,
            checked_hosts=checked_hosts,
        )
        for raw_url in raw_urls or []
    ]


def fetch_proxied_image(image_url: str) -> Dict[str, object]:
    """Fetch and cache a remote image so the browser can load it from same-origin."""
    image_url = validate_supplier_remote_url(image_url)
    cleanup_proxied_image_cache()
    with PROXIED_IMAGE_CACHE_LOCK:
        cached = PROXIED_IMAGE_CACHE.get(image_url)
        if cached:
            return cached

    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124.0.0.0 Safari/537.36',
        'Accept': 'image/avif,image/webp,image/apng,image/png,image/jpeg,image/*;q=0.8',
        'Referer': request.host_url.rstrip('/') + '/',
    }
    response = None
    current_url = image_url
    for redirect_count in range(6):
        current_url = validate_supplier_remote_url(current_url)
        response = requests.get(
            current_url,
            timeout=30,
            headers=headers,
            allow_redirects=False,
            stream=True,
        )
        if 300 <= response.status_code < 400 and response.headers.get('location'):
            if redirect_count >= 5:
                response.close()
                raise ValueError('Remote image exceeded the redirect limit.')
            next_url = urljoin(current_url, response.headers['location'])
            response.close()
            current_url = next_url
            continue
        break
    if response is None:
        raise ValueError('Remote image request failed.')
    response.raise_for_status()

    mime_type = (response.headers.get('content-type') or '').split(';', 1)[0].strip()
    if not mime_type or mime_type == 'application/octet-stream':
        mime_type = mimetypes.guess_type(urlparse(image_url).path)[0] or 'application/octet-stream'
    if not mime_type.startswith('image/') or mime_type == 'image/svg+xml':
        response.close()
        raise ValueError('Remote URL did not return a supported raster image.')

    declared_length = int(response.headers.get('content-length') or 0)
    if declared_length > MAX_PROXIED_IMAGE_BYTES:
        response.close()
        raise ValueError('Remote image exceeds the configured size limit.')
    chunks = []
    received = 0
    for chunk in response.iter_content(chunk_size=64 * 1024):
        if not chunk:
            continue
        received += len(chunk)
        if received > MAX_PROXIED_IMAGE_BYTES:
            response.close()
            raise ValueError('Remote image exceeds the configured size limit.')
        chunks.append(chunk)
    response.close()

    entry = {
        'data': b''.join(chunks),
        'mime_type': mime_type,
        'created_at': time.time(),
    }
    with PROXIED_IMAGE_CACHE_LOCK:
        PROXIED_IMAGE_CACHE[image_url] = entry
    return entry



# -------- Flask Routes --------
@app.get('/api/health')
def api_health():
    """Detailed health check â€” unauthenticated for monitoring systems."""
    return jsonify({
        'status': 'healthy',
        'timestamp': datetime.datetime.now(datetime.timezone.utc).isoformat(),
        'browser_engine': 'botasaurus',
        'scraper_mode': 'http_first_browser_fallback',
        'active_automation_jobs': len(AUTOMATION_ACTIVE_JOBS),
        'auth_configured': is_auth_configured(),
    })


@app.get('/livez')
def livez():
    """Liveness probe â€” returns 200 if the process is alive. No auth required."""
    return jsonify({'status': 'ok'}), 200


@app.get('/readyz')
def readyz():
    """Readiness probe â€” returns 200 if the app can serve requests. No auth required."""
    try:
        # Quick DB connectivity check â€” get_history_list with limit=0 is a no-op read.
        db_manager.get_history_list(limit=1)
        db_ok = True
    except Exception:
        db_ok = False
    status = 'ready' if db_ok else 'not_ready'
    code = 200 if db_ok else 503
    return jsonify({'status': status, 'database': 'ok' if db_ok else 'error'}), code


# ---------------------------------------------------------------------------
# Auth routes â€” always accessible (no @require_login on these)
# ---------------------------------------------------------------------------
@app.route('/login', methods=['GET', 'POST'])
def auth_login():
    """Login page. Redirects authenticated users straight to / ."""
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    if not is_auth_configured():
        # Auth not configured â€” skip login and go straight to the app.
        return redirect(url_for('index'))

    error = None
    if request.method == 'POST':
        username = (request.form.get('username') or '').strip()
        password = request.form.get('password') or ''
        user = validate_credentials(username, password)
        if user:
            login_user(user, remember=bool(request.form.get('remember')))
            flask_session.permanent = True
            next_url = request.args.get('next') or request.form.get('next') or url_for('index')
            # Guard against open-redirect: only allow relative URLs.
            from urllib.parse import urlparse as _urlparse
            parsed_next = _urlparse(next_url)
            if parsed_next.netloc:
                next_url = url_for('index')
            return redirect(next_url)
        error = "Invalid username or password."
    return render_template('login.html', error=error, next=request.args.get('next', ''))


@app.get('/logout')
def auth_logout():
    logout_user()
    return redirect(url_for('auth_login'))


@app.get('/')
@require_login
def index():
    return render_template('index.html')


@app.get('/sitemap.xml')
def sitemap():
    base_url = get_public_base_url()
    urlset = ET.Element('urlset', xmlns='http://www.sitemaps.org/schemas/sitemap/0.9')

    for page in build_public_site_pages():
        url_el = ET.SubElement(urlset, 'url')
        ET.SubElement(url_el, 'loc').text = f"{base_url}{page['path']}"
        ET.SubElement(url_el, 'lastmod').text = page['lastmod']
        ET.SubElement(url_el, 'changefreq').text = page['changefreq']
        ET.SubElement(url_el, 'priority').text = page['priority']

    xml_bytes = ET.tostring(urlset, encoding='utf-8', xml_declaration=True)
    return Response(xml_bytes, mimetype='application/xml')

@app.get('/robots.txt')
def robots():
    base_url = get_public_base_url()
    content = "\n".join([
        "User-agent: *",
        "Allow: /",
        "Disallow: /api/",
        "Disallow: /test-pagination",
        f"Sitemap: {base_url}/sitemap.xml",
        "",
    ])
    return Response(content, mimetype='text/plain')

@app.get('/history')
@require_login
def history():
    return render_template('history.html')

@app.get('/automation')
@require_login
def automation():
    return render_template('automation.html')




@app.get('/menu-map')
@require_login
def menu_map():
    return render_template('menu_map.html')


def get_menu_map_output_root() -> Path:
    configured = Path(os.getenv('MENU_MAP_OUTPUT_DIR', 'output'))
    if not configured.is_absolute():
        configured = APP_ROOT / configured
    return configured.resolve()


def parse_boolish(value) -> bool:
    return str(value).strip().lower() in {'1', 'true', 'yes', 'on'}


def count_csv_rows(path: Path) -> int:
    if not path.exists():
        return 0
    try:
        with path.open('r', encoding='utf-8-sig', newline='') as handle:
            return max(0, sum(1 for _ in csv.DictReader(handle)))
    except Exception:
        return 0


def summarize_hierarchy(tree: List[Dict[str, object]]) -> Dict[str, int]:
    parent_count = len(tree)
    sub_count = 0
    child_count = 0
    for parent in tree:
        subs = parent.get('sub_children') if isinstance(parent, dict) else []
        if not isinstance(subs, list):
            continue
        sub_count += len(subs)
        for sub in subs:
            children = sub.get('children') if isinstance(sub, dict) else []
            child_count += len(children) if isinstance(children, list) else 0
    return {
        'parents': parent_count,
        'sub_children': sub_count,
        'children': child_count,
        'total_nodes': parent_count + sub_count + child_count,
    }


def read_menu_map_site(slug: str, include_tree: bool = False) -> Dict[str, object]:
    if slug not in MENU_MAP_SITES:
        raise ValueError('Unknown menu-map site')

    output_dir = get_menu_map_output_root() / slug
    categories_json = output_dir / 'categories.json'
    categories_csv = output_dir / 'categories.csv'
    duplicate_csv = output_dir / 'duplicate_urls.csv'
    errors_json = output_dir / 'scraping_errors.json'

    tree: List[Dict[str, object]] = []
    output_valid = False
    parse_error = ''
    if categories_json.exists():
        try:
            parsed_tree = json.loads(categories_json.read_text(encoding='utf-8'))
            if not isinstance(parsed_tree, list):
                raise ValueError('categories.json must contain a list')
            tree = parsed_tree
            output_valid = True
        except Exception as exc:
            tree = []
            parse_error = f'{type(exc).__name__}: {exc}'

    summary = summarize_hierarchy(tree)
    missing_urls = 0
    if categories_csv.exists():
        try:
            with categories_csv.open('r', encoding='utf-8-sig', newline='') as handle:
                for row in csv.DictReader(handle):
                    if parse_boolish(row.get('url_missing')):
                        missing_urls += 1
        except Exception:
            missing_urls = 0

    errors = []
    if errors_json.exists():
        try:
            parsed_errors = json.loads(errors_json.read_text(encoding='utf-8'))
            errors = parsed_errors if isinstance(parsed_errors, list) else []
        except Exception:
            errors = []

    files = {}
    for filename in (
        'categories.json',
        'categories.csv',
        'categories.xlsx',
        'parent_categories.csv',
        'sub_child_categories.csv',
        'child_categories.csv',
        'duplicate_urls.csv',
        'scraping_errors.json',
        'dom_inspection.html',
        'dom_inspection.json',
        'healing_profile.json',
        'healing_attempt.json',
        'scraper.log',
    ):
        file_path = output_dir / filename
        if file_path.exists():
            files[filename] = {
                'name': filename,
                'size': file_path.stat().st_size,
                'download_url': url_for('api_menu_map_file', site=slug, filename=filename),
            }

    last_modified = None
    existing_files = [output_dir / name for name in files]
    if existing_files:
        last_modified = max(path.stat().st_mtime for path in existing_files if path.exists())

    data = {
        'slug': slug,
        **MENU_MAP_SITES[slug],
        'output_dir': str(output_dir),
        'has_output': categories_json.exists(),
        'output_valid': output_valid,
        'output_empty': output_valid and not tree,
        'parse_error': parse_error,
        'last_modified': datetime.datetime.fromtimestamp(last_modified).isoformat(timespec='seconds') if last_modified else '',
        'summary': summary,
        'missing_urls': missing_urls,
        'duplicate_rows': count_csv_rows(duplicate_csv),
        'error_count': len(errors),
        'errors': errors[:20],
        'files': files,
    }
    if include_tree:
        data['tree'] = tree
    return data


def make_menu_map_export_key(parts: List[object]) -> str:
    return '::'.join(re.sub(r'\s+', ' ', str(part or '').strip()) for part in parts)


def menu_map_parent_key(parent: Dict[str, object]) -> str:
    return make_menu_map_export_key([
        'parent',
        parent.get('display_order') or parent.get('order'),
        parent.get('parent_name'),
        parent.get('parent_url'),
    ])


def menu_map_sub_key(parent: Dict[str, object], sub: Dict[str, object]) -> str:
    return make_menu_map_export_key([
        'sub',
        parent.get('display_order') or parent.get('order'),
        parent.get('parent_name'),
        sub.get('display_order') or sub.get('order'),
        sub.get('sub_child_name'),
        sub.get('sub_child_url'),
    ])


def menu_map_child_key(parent: Dict[str, object], sub: Dict[str, object], child: Dict[str, object]) -> str:
    return make_menu_map_export_key([
        'child',
        parent.get('display_order') or parent.get('order'),
        parent.get('parent_name'),
        sub.get('display_order') or sub.get('order'),
        sub.get('sub_child_name'),
        child.get('display_order') or child.get('order'),
        child.get('child_name'),
        child.get('child_url'),
    ])


def menu_map_url_is_usable(url: object) -> bool:
    value = str(url or '').strip()
    return bool(value and not re.match(r'^(#|javascript[:;]?|javascript:void\(0\))', value, re.I))


def add_menu_map_link_row(
    rows: List[Dict[str, object]],
    seen: set,
    site: Dict[str, object],
    parent: Dict[str, object],
    sub: Dict[str, object] | None,
    label: object,
    url: object,
    hierarchy_level: str,
    child_order: object = '',
) -> None:
    url_value = str(url or '').strip()
    if not menu_map_url_is_usable(url_value):
        return
    key = url_value.lower()
    if key in seen:
        return
    seen.add(key)
    rows.append({
        'website': site.get('name') or '',
        'site_slug': site.get('slug') or '',
        'parent_name': parent.get('parent_name') or '',
        'sub_child_name': (sub or {}).get('sub_child_name') or '',
        'link_label': str(label or url_value).strip(),
        'url': url_value,
        'hierarchy_level': hierarchy_level,
        'parent_display_order': parent.get('display_order') or parent.get('order') or '',
        'sub_child_display_order': (sub or {}).get('display_order') or (sub or {}).get('order') or '',
        'child_display_order': child_order or '',
    })


def build_menu_map_link_rows(
    site_slugs: List[str],
    *,
    visible_only: bool,
    excluded_by_site: Dict[str, object] | None = None,
) -> List[Dict[str, object]]:
    rows: List[Dict[str, object]] = []
    excluded_by_site = excluded_by_site or {}
    for slug in site_slugs:
        site = read_menu_map_site(slug, include_tree=True)
        tree = site.get('tree') if isinstance(site.get('tree'), list) else []
        excluded = set(excluded_by_site.get(slug) or []) if visible_only else set()
        seen: set = set()
        for parent in tree:
            if not isinstance(parent, dict):
                continue
            if menu_map_parent_key(parent) in excluded:
                continue
            subs = parent.get('sub_children') if isinstance(parent.get('sub_children'), list) else []
            for sub in subs:
                if not isinstance(sub, dict):
                    continue
                if menu_map_sub_key(parent, sub) in excluded:
                    continue
                children = sub.get('children') if isinstance(sub.get('children'), list) else []
                for child in children:
                    if not isinstance(child, dict) or menu_map_child_key(parent, sub, child) in excluded:
                        continue
                    add_menu_map_link_row(
                        rows,
                        seen,
                        site,
                        parent,
                        sub,
                        child.get('child_name'),
                        child.get('child_url'),
                        'child',
                        child.get('display_order') or child.get('order') or '',
                    )
                if not children:
                    add_menu_map_link_row(
                        rows,
                        seen,
                        site,
                        parent,
                        sub,
                        sub.get('sub_child_name'),
                        sub.get('sub_child_url'),
                        'sub_child',
                    )
            if not subs:
                add_menu_map_link_row(
                    rows,
                    seen,
                    site,
                    parent,
                    None,
                    parent.get('parent_name'),
                    parent.get('parent_url'),
                    'parent',
                )
    return rows


@app.get('/api/menu-map/sites')
@require_login
def api_menu_map_sites():
    include_tree = parse_boolish(request.args.get('include_tree'))
    try:
        return jsonify({
            'sites': [read_menu_map_site(slug, include_tree=include_tree) for slug in MENU_MAP_SITES],
            'jobs': list(MENU_MAP_JOBS.values()),
            'output_root': str(get_menu_map_output_root()),
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.post('/api/menu-map/output/clear')
@require_login
@require_role('admin')
def api_menu_map_output_clear():
    data = request.get_json(silent=True) or {}
    requested_sites = data.get('sites') or []
    if not isinstance(requested_sites, list):
        return jsonify({'error': 'sites must be a list'}), 400
    sites = [site for site in requested_sites if site in MENU_MAP_SITES]
    if not sites:
        return jsonify({'error': 'No valid sites selected'}), 400

    with MENU_MAP_JOBS_LOCK:
        active_job = next((
            job for job in MENU_MAP_JOBS.values()
            if job.get('status') in {'queued', 'running'}
            and set(sites).intersection(job.get('sites') or [])
        ), None)
        if active_job:
            return jsonify({
                'error': 'A menu-map run is already active for one or more selected websites.',
                'job': active_job,
            }), 409

    output_root = get_menu_map_output_root().resolve()
    cleared = []
    missing = []
    for site in sites:
        site_dir = (output_root / site).resolve()
        if site_dir.parent != output_root:
            return jsonify({'error': 'Invalid menu-map output path'}), 400
        if not site_dir.exists():
            missing.append(site)
            continue
        shutil.rmtree(site_dir)
        cleared.append(site)

    return jsonify({
        'success': True,
        'cleared': cleared,
        'missing': missing,
        'cleared_count': len(cleared),
    })


@app.post('/api/menu-map/links/export')
@require_login
def api_menu_map_links_export():
    data = request.get_json(silent=True) or {}
    requested_sites = data.get('sites') or []
    if not isinstance(requested_sites, list):
        return jsonify({'error': 'sites must be a list'}), 400
    sites = [site for site in requested_sites if site in MENU_MAP_SITES]
    if not sites:
        return jsonify({'error': 'No valid sites selected'}), 400

    scope = str(data.get('scope') or 'visible').strip().lower()
    if scope not in {'visible', 'full'}:
        return jsonify({'error': 'scope must be visible or full'}), 400
    export_format = str(data.get('format') or 'csv').strip().lower()
    if export_format not in {'csv', 'xlsx'}:
        return jsonify({'error': 'format must be csv or xlsx'}), 400
    excluded = data.get('excluded') if isinstance(data.get('excluded'), dict) else {}
    rows = build_menu_map_link_rows(sites, visible_only=(scope == 'visible'), excluded_by_site=excluded)
    if not rows:
        return jsonify({'error': 'No menu-map links are available to export'}), 404

    stamp = datetime.datetime.now().strftime('%Y%m%d-%H%M%S')
    site_part = sites[0] if len(sites) == 1 else f'{len(sites)}-sites'
    filename = f'menu-map-links-{site_part}-{scope}-{stamp}.{export_format}'
    headers = [
        'website',
        'site_slug',
        'parent_name',
        'sub_child_name',
        'link_label',
        'url',
        'hierarchy_level',
        'parent_display_order',
        'sub_child_display_order',
        'child_display_order',
    ]

    if export_format == 'csv':
        buffer = io.StringIO()
        writer = csv.DictWriter(buffer, fieldnames=headers, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(rows)
        return Response(
            buffer.getvalue(),
            mimetype='text/csv; charset=utf-8',
            headers={'Content-Disposition': f'attachment; filename="{filename}"'},
        )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Menu Links'
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, '') for header in headers])
    for cell in sheet[1]:
        header_font = copy(cell.font)
        header_font.bold = True
        cell.font = header_font
    sheet.freeze_panes = 'A2'
    sheet.auto_filter.ref = sheet.dimensions
    for column_cells in sheet.columns:
        width = min(max(len(str(cell.value or '')) for cell in column_cells[:200]) + 2, 70)
        sheet.column_dimensions[column_cells[0].column_letter].width = width
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )


def run_menu_map_job(job_id: str, sites: List[str], options: Dict[str, object]) -> None:
    with MENU_MAP_JOBS_LOCK:
        job = MENU_MAP_JOBS[job_id]
        job['status'] = 'running'
        job['started_at'] = datetime.datetime.now().isoformat(timespec='seconds')

    for site in sites:
        site_config = MENU_MAP_SITES.get(site)
        if not site_config:
            continue

        cmd = [
            sys.executable,
            '-m',
            site_config['module'],
            '--output-dir',
            str(options.get('output_dir') or get_menu_map_output_root()),
            '--timeout',
            str(site_config.get('timeout') or options.get('timeout') or 60000),
            '--interaction-delay',
            str(options.get('interaction_delay') or 600),
            '--log-level',
            str(options.get('log_level') or 'INFO'),
        ]
        cmd.append('--visible' if options.get('visible') else '--headless')
        if options.get('inspect_only'):
            cmd.append('--inspect-only')
        if options.get('validate_urls'):
            cmd.append('--validate-urls')

        with MENU_MAP_JOBS_LOCK:
            job = MENU_MAP_JOBS[job_id]
            job['current_site'] = site
            job['events'].append({'site': site, 'message': 'Starting scraper', 'command': ' '.join(cmd)})

        try:
            completed = subprocess.run(
                cmd,
                cwd=str(APP_ROOT),
                text=True,
                capture_output=True,
                timeout=int(options.get('process_timeout') or 900),
            )
            event = {
                'site': site,
                'returncode': completed.returncode,
                'stdout': completed.stdout[-4000:],
                'stderr': completed.stderr[-4000:],
            }
            with MENU_MAP_JOBS_LOCK:
                job = MENU_MAP_JOBS[job_id]
                job['events'].append(event)
                job['site_status'][site] = 'success' if completed.returncode == 0 else 'failed'
        except Exception as exc:
            with MENU_MAP_JOBS_LOCK:
                job = MENU_MAP_JOBS[job_id]
                job['events'].append({'site': site, 'returncode': -1, 'stderr': f'{type(exc).__name__}: {exc}'})
                job['site_status'][site] = 'failed'

    with MENU_MAP_JOBS_LOCK:
        job = MENU_MAP_JOBS[job_id]
        failed = [site for site, status in job['site_status'].items() if status == 'failed']
        job['status'] = 'failed' if failed else 'complete'
        job['current_site'] = ''
        job['completed_at'] = datetime.datetime.now().isoformat(timespec='seconds')


@app.post('/api/menu-map/run')
@require_login
def api_menu_map_run():
    data = request.get_json(silent=True) or {}
    requested_sites = data.get('sites') or list(MENU_MAP_SITES)
    if not isinstance(requested_sites, list):
        return jsonify({'error': 'sites must be a list'}), 400
    sites = [site for site in requested_sites if site in MENU_MAP_SITES]
    if not sites:
        return jsonify({'error': 'No valid sites selected'}), 400

    job_id = str(uuid.uuid4())
    options = {
        'visible': False,
        'inspect_only': coerce_bool(data.get('inspect_only'), default=False),
        'validate_urls': coerce_bool(data.get('validate_urls'), default=False),
        'output_dir': str(data.get('output_dir') or get_menu_map_output_root()),
        'timeout': coerce_int(data.get('timeout'), 60000, min_value=10000, max_value=180000),
        'interaction_delay': coerce_int(data.get('interaction_delay'), 600, min_value=0, max_value=5000),
        'log_level': str(data.get('log_level') or 'INFO'),
        'process_timeout': coerce_int(data.get('process_timeout'), 900, min_value=60, max_value=3600),
    }
    with MENU_MAP_JOBS_LOCK:
        requested_site_set = set(sites)
        overlapping_job = next((
            job for job in MENU_MAP_JOBS.values()
            if job.get('status') in {'queued', 'running'}
            and requested_site_set.intersection(job.get('sites') or [])
        ), None)
        if overlapping_job:
            return jsonify({
                'error': 'A menu-map run is already active for one or more selected websites.',
                'job': overlapping_job,
            }), 409
        MENU_MAP_JOBS[job_id] = {
            'id': job_id,
            'status': 'queued',
            'sites': sites,
            'site_status': {},
            'current_site': '',
            'created_at': datetime.datetime.now().isoformat(timespec='seconds'),
            'started_at': '',
            'completed_at': '',
            'options': options,
            'events': [],
        }
    thread = threading.Thread(target=run_menu_map_job, args=(job_id, sites, options), daemon=True)
    thread.start()
    return jsonify({'success': True, 'job': MENU_MAP_JOBS[job_id]})


@app.get('/api/menu-map/jobs/<job_id>')
@require_login
def api_menu_map_job(job_id):
    with MENU_MAP_JOBS_LOCK:
        job = MENU_MAP_JOBS.get(job_id)
        if not job:
            return jsonify({'error': 'Job not found'}), 404
        return jsonify({'job': job})


@app.get('/api/menu-map/file/<site>/<path:filename>')
@require_login
def api_menu_map_file(site, filename):
    if site not in MENU_MAP_SITES:
        return jsonify({'error': 'Unknown site'}), 404
    allowed = {
        'categories.json',
        'categories.csv',
        'categories.xlsx',
        'parent_categories.csv',
        'sub_child_categories.csv',
        'child_categories.csv',
        'duplicate_urls.csv',
        'scraping_errors.json',
        'dom_inspection.html',
        'dom_inspection.json',
        'healing_profile.json',
        'healing_attempt.json',
        'scraper.log',
    }
    safe_name = os.path.basename(filename)
    if safe_name not in allowed:
        return jsonify({'error': 'File is not available for download'}), 403
    file_path = get_menu_map_output_root() / site / safe_name
    if not file_path.exists():
        return jsonify({'error': 'File not found'}), 404
    return send_file(file_path, as_attachment=True)



@app.get('/api/history')
@require_login
def api_history():
    """Return history data from database"""
    try:
        page = coerce_int(request.args.get('page', 1), 1, min_value=1)
        limit = coerce_int(request.args.get('limit', 20), 20, min_value=1, max_value=200)
        offset = (page - 1) * limit

        histories = db_manager.get_history_list(limit=limit, offset=offset)
        return jsonify({
            'histories': histories,
            'page': page,
            'limit': limit,
            'total': len(histories)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.get('/api/history/<history_id>')
@require_login
def api_history_detail(history_id):
    """Return specific history entry from database"""
    try:
        history = db_manager.get_history_detail(history_id)
        if not history:
            return jsonify({'error': 'History entry not found'}), 404
        return jsonify(history)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.post('/api/history/<history_id>/export/xlsx')
@require_login
def api_history_export(history_id):
    """Export a specific history entry to XLSX"""
    try:
        history = db_manager.get_history_detail(history_id)
        if not history:
            return jsonify({'error': 'History entry not found'}), 404

        items = history.get('items', [])

        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["Product", "SKU", "Price", "Stock", "Description", "Link"])
        ws.freeze_panes = "A2"

        for item in items:
            extra = item.get('extra') if isinstance(item.get('extra'), dict) else {}
            price = (
                item.get('original_formatted')
                or item.get('price_text')
                or item.get('discounted_formatted')
                or item.get('price_value')
                or item.get('original')
                or ''
            )
            ws.append([
                clean_excel_cell(item.get('title', '')),
                clean_excel_cell(item.get('sku') or extra.get('sku') or ''),
                clean_excel_cell(price),
                clean_excel_cell(item.get('stock_status') or extra.get('stock_status') or ''),
                clean_excel_cell(item.get('description') or extra.get('description') or ''),
                clean_excel_cell(item.get('url', '')),
            ])

        for column_cells in ws.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                max_length = max(max_length, len(str(cell.value or '')))
            ws.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 70)

        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        filename = f"history_{history_id}.xlsx"
        return send_file(
            bio,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

def _delete_history_response(history_id):
    normalized_history_id = str(history_id).strip()
    if not normalized_history_id:
        return jsonify({'error': 'History ID is required'}), 400

    try:
        success = db_manager.delete_history(normalized_history_id)
        if success:
            return jsonify({'success': True, 'history_id': normalized_history_id})
        return jsonify({'error': 'History entry not found'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.delete('/api/history/<history_id>')
@require_login
@require_role('admin')
@require_destructive_confirmation
def api_delete_history(history_id):
    """Delete history entry from database"""
    return _delete_history_response(history_id)

@app.post('/api/history/<history_id>/delete')
@require_login
@require_role('admin')
@require_destructive_confirmation
def api_delete_history_fallback(history_id):
    """Delete history entry from database (POST fallback when DELETE is blocked)"""
    return _delete_history_response(history_id)

@app.get('/api/statistics')
@require_login
def api_statistics():
    """Get database statistics"""
    try:
        stats = db_manager.get_statistics()
        return jsonify(stats)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.post('/api/search')
@require_login
def api_search():
    """Search items in database"""
    try:
        data = request.get_json(silent=True) or {}
        query = str(data.get('query') or '').strip()
        limit = coerce_int(data.get('limit', 100), 100, min_value=1, max_value=500)

        if not query:
            return jsonify({'error': 'Search query is required'}), 400

        items = db_manager.search_items(query, limit)
        return jsonify({
            'query': query,
            'results': items,
            'count': len(items)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.get('/api/automation/overview')
@require_login
def api_automation_overview():
    try:
        overview = db_manager.get_automation_overview()
        recent_runs = db_manager.list_automation_runs(limit=10)
        return jsonify({
            'overview': overview,
            'recent_runs': recent_runs,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.post('/api/automation/discover')
@require_login
def api_automation_discover():
    try:
        data = request.get_json(silent=True) or {}
        raw_scraper_value = data.get('scraper_key') or data.get('site') or data.get('site_key') or ''
        root_url = str(data.get('root_url') or '').strip()
        scraper_key = str(raw_scraper_value or '').strip().lower()
        if scraper_key not in SCRAPER_CONFIG:
            scraper_key = detect_scraper_key(str(raw_scraper_value or root_url or '').strip())
        if scraper_key not in SCRAPER_CONFIG:
            scraper_key = 'standard'
        category_query = str(data.get('category_query') or '').strip()
        retries = coerce_int(data.get('retries') or 1, 1, min_value=1, max_value=5)
        verify_ssl = coerce_bool(data.get('verify_ssl'), default=True)

        if not category_query:
            return jsonify({'error': 'Category query is required.'}), 400
        if root_url:
            try:
                root_url = validate_supplier_remote_urls([root_url], scraper_key)[0]
            except ValueError as exc:
                return jsonify({'error': str(exc)}), 400

        discovered = discover_category_targets_via_browser(
            scraper_key,
            category_query,
            root_url=root_url,
            retries=retries,
            verify_ssl=verify_ssl,
            logger=app.logger,
        )
        return jsonify(discovered)
    except Exception as e:
        app.logger.exception("[automation] Discovery failed")
        return jsonify({'error': str(e)}), 500


@app.get('/api/automation/jobs')
@require_login
def api_automation_jobs():
    try:
        include_targets = coerce_bool(request.args.get('include_targets'), default=False)
        jobs = db_manager.list_automation_jobs(include_targets=include_targets, limit=200)
        return jsonify({
            'jobs': jobs,
            'overview': db_manager.get_automation_overview(),
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.post('/api/automation/jobs')
@require_login
def api_automation_jobs_save():
    try:
        data = request.get_json(silent=True) or {}
        targets = data.get('targets') if isinstance(data.get('targets'), list) else None
        raw_scraper_key = data.get('scraper_key') or data.get('site') or data.get('site_key') or ''
        scraper_key = str(raw_scraper_key or '').strip().lower()
        if scraper_key not in SCRAPER_CONFIG:
            scraper_key = detect_scraper_key(data.get('root_url') or raw_scraper_key)
        try:
            if str(data.get('root_url') or '').strip():
                data['root_url'] = validate_supplier_remote_urls([data['root_url']], scraper_key)[0]
            if targets is not None:
                validated_urls = validate_supplier_remote_urls(
                    [target.get('url') for target in targets if isinstance(target, dict)],
                    scraper_key,
                )
                validated_iter = iter(validated_urls)
                targets = [
                    {**target, 'url': next(validated_iter)}
                    for target in targets
                    if isinstance(target, dict)
                ]
        except ValueError as exc:
            return jsonify({'error': str(exc)}), 400
        job = db_manager.save_automation_job(data, targets=targets)
        if not job:
            return jsonify({'error': 'Failed to save automation job. Check the site and category query.'}), 400
        return jsonify({
            'success': True,
            'job': job,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.get('/api/automation/jobs/<int:job_id>')
@require_login
def api_automation_job_detail(job_id):
    try:
        job = db_manager.get_automation_job(job_id, include_targets=True)
        if not job:
            return jsonify({'error': 'Automation job not found.'}), 404
        return jsonify({
            'job': job,
            'runs': db_manager.list_automation_runs(job_id=job_id, limit=15),
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.delete('/api/automation/jobs/<int:job_id>')
@require_login
@require_role('admin')
@require_destructive_confirmation
def api_automation_job_delete(job_id):
    try:
        if db_manager.delete_automation_job(job_id):
            return jsonify({'success': True, 'job_id': job_id})
        return jsonify({'error': 'Automation job not found.'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.post('/api/automation/jobs/<int:job_id>/toggle')
@require_login
def api_automation_job_toggle(job_id):
    try:
        data = request.get_json(silent=True) or {}
        enabled = data.get('enabled', True)
        if isinstance(enabled, str):
            enabled = enabled.strip().lower() in {'1', 'true', 'yes', 'on'}
        else:
            enabled = bool(enabled)
        job = db_manager.set_automation_job_enabled(job_id, enabled)
        if not job:
            return jsonify({'error': 'Automation job not found.'}), 404
        return jsonify({'success': True, 'job': job})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.post('/api/automation/jobs/<int:job_id>/refresh-targets')
@require_login
def api_automation_job_refresh_targets(job_id):
    try:
        job = db_manager.get_automation_job(job_id, include_targets=True)
        if not job:
            return jsonify({'error': 'Automation job not found.'}), 404

        try:
            safe_root_url = validate_supplier_remote_urls(
                [job.get('root_url')],
                job.get('scraper_key'),
            )[0]
        except ValueError as exc:
            return jsonify({'error': str(exc)}), 400

        discovered = discover_category_targets_via_browser(
            job.get('scraper_key'),
            job.get('category_query'),
            root_url=safe_root_url,
            retries=job.get('retries', 1),
            verify_ssl=job.get('verify_ssl', True),
            logger=app.logger,
        )
        targets = db_manager.replace_automation_job_targets(job_id, discovered.get('targets', []))
        return jsonify({
            'success': True,
            'job': db_manager.get_automation_job(job_id, include_targets=True),
            'discovery': discovered,
            'targets': targets,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.post('/api/automation/jobs/<int:job_id>/targets')
@require_login
def api_automation_job_targets(job_id):
    try:
        job = db_manager.get_automation_job(job_id, include_targets=False)
        if not job:
            return jsonify({'error': 'Automation job not found.'}), 404

        data = request.get_json(silent=True) or {}
        targets = data.get('targets')
        if not isinstance(targets, list):
            return jsonify({'error': 'A targets list is required.'}), 400
        try:
            validated_urls = validate_supplier_remote_urls(
                [target.get('url') for target in targets if isinstance(target, dict)],
                job.get('scraper_key'),
            )
            validated_iter = iter(validated_urls)
            targets = [
                {**target, 'url': next(validated_iter)}
                for target in targets
                if isinstance(target, dict)
            ]
        except ValueError as exc:
            return jsonify({'error': str(exc)}), 400

        saved_targets = db_manager.replace_automation_job_targets(job_id, targets)
        updated_job = db_manager.get_automation_job(job_id, include_targets=True)
        return jsonify({
            'success': True,
            'job': updated_job,
            'targets': saved_targets,
            'overview': db_manager.get_automation_overview(),
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.post('/api/automation/jobs/<int:job_id>/run')
@require_login
def api_automation_job_run(job_id):
    try:
        if not db_manager.get_automation_job(job_id, include_targets=False):
            return jsonify({'error': 'Automation job not found.'}), 404

        active_run = db_manager.get_active_automation_run_for_job(job_id)
        if active_run:
            return jsonify({
                'error': 'Automation job is already running.',
                'run_id': active_run.get('id'),
                'status': active_run.get('status'),
            }), 409

        recent_runs = db_manager.list_automation_runs(job_id=job_id, limit=1)
        latest_run = recent_runs[0] if recent_runs else None
        latest_status = str((latest_run or {}).get('status') or '').strip().lower()
        if latest_status in {'paused', 'interrupted', 'failed'}:
            queued, message = _launch_existing_automation_run(latest_run['id'])
            if not queued:
                return jsonify({'error': message or 'Failed to resume automation run.'}), 409
            return jsonify({
                'success': True,
                'queued': True,
                'resumed': True,
                'job_id': job_id,
                'run_id': latest_run['id'],
            })

        queued, message = _launch_automation_job(job_id, trigger_type='manual')
        if not queued:
            return jsonify({'error': message or 'Automation job is already running.'}), 409
        return jsonify({
            'success': True,
            'queued': True,
            'resumed': False,
            'job_id': job_id,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.get('/api/automation/runs')
@require_login
def api_automation_runs():
    try:
        job_id = request.args.get('job_id')
        scraper_key = request.args.get('scraper_key')
        limit = coerce_int(request.args.get('limit', 25), 25, min_value=1, max_value=100)
        runs = db_manager.list_automation_runs(job_id=job_id, scraper_key=scraper_key, limit=limit)
        include_details = coerce_bool(request.args.get('include_details'), default=False)
        include_models = coerce_bool(request.args.get('include_models'), default=True)
        if not include_details:
            compact_runs = []
            for run in runs:
                compact_run = dict(run)
                summary = compact_run.get('summary')
                summary = dict(summary) if isinstance(summary, dict) else {}
                preview_items = summary.pop('preview_items', None)
                if isinstance(preview_items, list):
                    summary['preview_item_count'] = len(preview_items)
                summary.pop('checkpoint_items', None)
                if not include_models:
                    models = summary.pop('models', None)
                    if isinstance(models, list):
                        summary['model_count'] = len(models)
                target_urls = compact_run.pop('target_urls', None)
                if isinstance(target_urls, list):
                    summary.setdefault('total_targets', len(target_urls))
                    summary.setdefault('target_count', len(target_urls))
                compact_run['summary'] = summary
                compact_runs.append(compact_run)
            runs = compact_runs
        return jsonify({
            'runs': runs,
            'count': len(runs),
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.delete('/api/automation/runs/<int:run_id>')
@require_login
@require_role('admin')
@require_destructive_confirmation
def api_delete_automation_run(run_id):
    try:
        success = db_manager.delete_automation_run(run_id)
        if not success:
            return jsonify({'error': 'Failed to delete automation run or run not found.'}), 404
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.post('/api/automation/runs/<int:run_id>/delete')
@require_login
@require_role('admin')
@require_destructive_confirmation
def api_delete_automation_run_fallback(run_id):
    try:
        success = db_manager.delete_automation_run(run_id)
        if not success:
            return jsonify({'error': 'Failed to delete automation run or run not found.'}), 404
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.post('/api/automation/runs/<int:run_id>/pause')
@require_login
def api_pause_automation_run(run_id):
    try:
        run = db_manager.get_automation_run(run_id)
        if not run:
            return jsonify({'error': 'Automation run not found.'}), 404
        if str(run.get('status') or '').strip().lower() != 'running':
            return jsonify({'error': 'Only a running automation run can be paused.'}), 409
        paused = db_manager.pause_automation_run(run_id, reason='Automation run paused by user.')
        if not paused:
            return jsonify({'error': 'Failed to pause automation run.'}), 500
        return jsonify({'success': True, 'run': paused})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.post('/api/automation/runs/<int:run_id>/resume')
@require_login
def api_resume_automation_run(run_id):
    try:
        queued, message = _launch_existing_automation_run(run_id)
        if not queued:
            return jsonify({'error': message or 'Failed to resume automation run.'}), 409
        return jsonify({'success': True, 'queued': True, 'run_id': run_id})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.get('/api/automation/runs/<int:run_id>')
@require_login
def api_automation_run_detail(run_id):
    try:
        run = db_manager.get_automation_run(run_id)
        if not run:
            return jsonify({'error': 'Automation run not found.'}), 404

        current_history = db_manager.get_history_detail(run.get('current_history_id')) if run.get('current_history_id') else None
        previous_history = db_manager.get_history_detail(run.get('previous_history_id')) if run.get('previous_history_id') else None
        live_preview_items = []
        if not current_history:
            run_summary = run.get('summary') if isinstance(run.get('summary'), dict) else {}
            # Populate live_preview_items from the in-progress run summary so that
            # running automation runs expose their current products to the frontend.
            persisted_preview = db_manager.get_automation_run_items(run_id, limit=AUTOMATION_LIVE_DETAIL_ITEM_LIMIT)
            if persisted_preview:
                live_preview_items = [item for item in persisted_preview if isinstance(item, dict)]
            else:
                raw_preview = run_summary.get('preview_items')
                if isinstance(raw_preview, list):
                    live_preview_items = [item for item in raw_preview if isinstance(item, dict)]

        current_history_items = (current_history or {}).get('items', [])
        current_product_items = deduplicate_comparable_items(current_history_items)
        curr_hid = str(run.get('current_history_id') or '')
        prev_hid = str(run.get('previous_history_id') or '')
        cache_key = f"{curr_hid}:{prev_hid}"

        global _COMPARISON_CACHE
        if '_COMPARISON_CACHE' not in globals():
            _COMPARISON_CACHE = {}

        if cache_key in _COMPARISON_CACHE and curr_hid:
            comparison = _COMPARISON_CACHE[cache_key]
        elif current_history:
            comparison = build_session_comparison(
                previous_history,
                current_history_items,
                current_target_urls=(current_history or {}).get('urls', []),
                run_validation=validation_from_history_rejection(current_history),
            )
            if curr_hid and len(_COMPARISON_CACHE) < 50:
                _COMPARISON_CACHE[cache_key] = comparison
        else:
            comparison = {
                'has_previous_run': False,
                'summary': {},
                'added': [],
                'removed': [],
                'verification_required': [],
                'review_required': [],
                'changed': [],
            }
        models = build_automation_model_summary(comparison)
        comparison_summary = comparison.get('summary', {}) if isinstance(comparison.get('summary'), dict) else {}
        public_run = dict(run)
        public_run.pop('target_urls', None)
        public_summary = public_run.get('summary')
        public_summary = dict(public_summary) if isinstance(public_summary, dict) else {}
        preview_items = public_summary.pop('preview_items', None)
        if isinstance(preview_items, list):
            public_summary['preview_item_count'] = len(preview_items)
        public_summary.pop('checkpoint_items', None)
        is_running_run = str(run.get('status') or '').lower() in {'running', 'resuming'}
        if is_running_run:
            raw_summary = run.get('summary') if isinstance(run.get('summary'), dict) else {}
            raw_harvested = raw_summary.get('current_items') or run.get('items_count') or 0
            public_summary['current_items'] = raw_harvested
            public_run['items_count'] = raw_harvested
        elif current_history and comparison_summary:
            for key in (
                'previous_items',
                'current_items',
                'previous_rows',
                'current_rows',
                'excluded_previous_non_products',
                'excluded_current_non_products',
                'duplicate_previous_rows',
                'duplicate_current_rows',
                'out_of_scope_previous_products',
                'added',
                'removed',
                'changed',
                'price_changes',
                'stock_changes',
                'title_changes',
                'sku_changes',
                'description_changes',
                'url_changes',
                'category_changes',
                'temporarily_missing',
                'verification_required',
                'scrape_failures',
                'removed_confirmed',
                'review_required',
            ):
                public_summary[key] = comparison_summary.get(key, 0)
            public_run['stored_items_count'] = public_run.get('items_count')
            public_run['items_count'] = comparison_summary.get('current_items', len(current_product_items))
        public_run['summary'] = public_summary

        compact_product_items = [
            {
                'title': i.get('title') or '',
                'url': i.get('url') or '',
                'image_url': i.get('image_url') or '',
                'original_formatted': i.get('original_formatted') or '',
                'discounted_formatted': i.get('discounted_formatted') or '',
                'sku': i.get('sku') or '',
                'category': i.get('category') or (i.get('extra') or {}).get('model_label') or (i.get('extra') or {}).get('target_label') or '',
                'model_label': (i.get('extra') or {}).get('model_label') or '',
                'target_label': (i.get('extra') or {}).get('target_label') or '',
                'duplicate_categories': i.get('duplicate_categories') or [],
                'duplicate_count': i.get('duplicate_count', 1),
                'is_duplicate': bool(i.get('is_duplicate')),
                'stock_status': i.get('stock_status') or 'In Stock',
                'description': i.get('description') or '',
                'source': i.get('site') or i.get('source') or (str(run.get('current_history_id') or '').split(':')[0] if ':' in str(run.get('current_history_id') or '') else 'parts4cells'),
                'extra': i.get('extra') or {},
            }
            for i in current_product_items
        ]

        return jsonify({
            'run': public_run,
            'job': db_manager.get_automation_job(run.get('job_id'), include_targets=False),
            'current_history': {
                'id': (current_history or {}).get('id'),
                'timestamp': (current_history or {}).get('timestamp'),
                'items_count': comparison_summary.get('current_items', len(current_product_items)),
                'rows_count': (current_history or {}).get('items_count'),
                'urls': (current_history or {}).get('urls', []),
                'items': compact_product_items,
            } if current_history else ({
                'id': '',
                'timestamp': run.get('started_at'),
                'items_count': run.get('items_count') or len(live_preview_items),
                'urls': [],
                'items': live_preview_items[:50],
                'is_live_preview': True,
            } if live_preview_items else None),
            'previous_history': {
                'id': (previous_history or {}).get('id'),
                'timestamp': (previous_history or {}).get('timestamp'),
                'items_count': comparison_summary.get('previous_items', 0),
                'rows_count': (previous_history or {}).get('items_count'),
                'urls': (previous_history or {}).get('urls', []),
            } if previous_history else None,
            'comparison': comparison,
            'models': models,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.get('/api/automation/verification-products')
@require_login
def api_automation_verification_products():
    """Return products from the latest one-child verification scrape for dashboard preview."""
    try:
        verification_path = APP_ROOT / 'output' / 'single_child_scrape_verification.json'
        if not verification_path.exists():
            return jsonify({'groups': [], 'items_count': 0})

        try:
            verification_rows = json.loads(verification_path.read_text(encoding='utf-8'))
        except Exception:
            verification_rows = []

        groups = []
        total_items = 0
        ignored_titles = {'forbidden', 'access denied'}
        for row in verification_rows if isinstance(verification_rows, list) else []:
            history_id = str(row.get('history_public_id') or '').strip()
            if not history_id:
                continue
            history = db_manager.get_history_detail(history_id)
            items = (history or {}).get('items', []) if history else []
            cleaned_items = []
            for item in items:
                title = str((item or {}).get('title') or '').strip()
                if not title or title.lower() in ignored_titles:
                    continue
                lowered_title = title.lower()
                if lowered_title.startswith('www.') and lowered_title.endswith('.com'):
                    continue
                cleaned_items.append(item)
            if not cleaned_items:
                continue
            total_items += len(cleaned_items)
            groups.append({
                'site': row.get('site') or '',
                'parent': row.get('parent') or '',
                'sub_child': row.get('sub_child') or '',
                'child': row.get('name') or '',
                'category_url': row.get('url') or '',
                'history_id': history_id,
                'items_count': len(cleaned_items),
                'items': cleaned_items[:500],
                'first_title': row.get('first_title') or '',
            })

        return jsonify({
            'groups': groups,
            'items_count': total_items,
            'source': str(verification_path),
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.get('/api/watchlist')
@require_login
def api_watchlist():
    """Return all saved watchlist items across site databases."""
    try:
        limit_value = request.args.get('limit')
        limit = coerce_int(limit_value, 100, min_value=1, max_value=1000) if limit_value else None
        items = db_manager.get_watchlist_items(limit=limit)
        return jsonify({
            'items': items,
            'count': len(items),
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.post('/api/watchlist')
@require_login
def api_watchlist_save():
    """Save or update a watchlist item snapshot."""
    try:
        data = request.get_json(silent=True) or {}
        if not isinstance(data, dict):
            return jsonify({'error': 'Invalid watchlist payload'}), 400

        url = str(data.get('url') or '').strip()
        if not url:
            return jsonify({'error': 'Item URL is required'}), 400

        saved_item = db_manager.save_watchlist_item(data)
        if not saved_item:
            return jsonify({'error': 'Failed to save watchlist item'}), 500

        return jsonify({
            'success': True,
            'item': saved_item,
            'count': len(db_manager.get_watchlist_urls()),
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.delete('/api/watchlist')
@require_login
def api_watchlist_delete():
    """Remove one watchlist item by URL."""
    try:
        data = request.get_json(silent=True) or {}
        url = str(request.args.get('url') or data.get('url') or '').strip()
        if not url:
            return jsonify({'error': 'Item URL is required'}), 400

        removed = db_manager.remove_watchlist_item(url)
        if not removed:
            return jsonify({'error': 'Watchlist item not found', 'count': len(db_manager.get_watchlist_urls())}), 404

        return jsonify({
            'success': True,
            'url': url,
            'count': len(db_manager.get_watchlist_urls()),
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.post('/api/watchlist/clear')
@require_login
@require_role('admin')
@require_destructive_confirmation
def api_watchlist_clear():
    """Clear the entire shared watchlist."""
    try:
        cleared = db_manager.clear_watchlist()
        return jsonify({
            'success': True,
            'cleared': cleared,
            'count': 0,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.post('/api/cleanup')
@require_login
@require_role('admin')
@require_destructive_confirmation
def api_cleanup():
    """Cleanup old database entries"""
    try:
        data = request.get_json(silent=True) or {}
        delete_all = bool(data.get('delete_all')) or str(data.get('mode') or '').strip().lower() in {'all', 'delete_all'}
        raw_days = data.get('days', 90)
        if not delete_all:
            try:
                delete_all = int(raw_days) >= 99999
            except (TypeError, ValueError):
                delete_all = False

        days = 99999 if delete_all else coerce_int(raw_days, 90, min_value=1, max_value=3650)

        deleted_count = db_manager.cleanup_old_entries(days)
        return jsonify({
            'success': True,
            'deleted_entries': deleted_count,
            'days': days,
            'delete_all': delete_all
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500



@app.get('/api/image-proxy')
@require_login
def proxy_remote_image():
    """Proxy remote images through the app to avoid browser-side hotlink failures."""
    image_url = (request.args.get('url') or '').strip()
    if not image_url:
        return jsonify({'error': 'Image URL is required'}), 400

    try:
        entry = fetch_proxied_image(image_url)
    except ValueError as exc:
        return jsonify({'error': str(exc)}), 400
    except Exception as exc:
        app.logger.warning(f"[image-proxy] Failed to fetch {image_url}: {exc}")
        return jsonify({'error': f'Failed to load image: {exc}'}), 502

    return send_file(
        io.BytesIO(entry['data']),
        mimetype=entry['mime_type'],
        max_age=PROXIED_IMAGE_TTL_SECONDS,
    )

@app.post('/api/scrape')
@require_login
def api_scrape():
    data = request.get_json(silent=True) or {}
    urls_raw = data.get('urls') or ''
    requested_urls = [
        value.strip()
        for value in (urls_raw.splitlines() if isinstance(urls_raw, str) else urls_raw or [])
        if str(value or '').strip()
    ]
    try:
        validate_supplier_remote_urls(requested_urls)
    except ValueError as exc:
        return jsonify({'error': str(exc), 'history_saved': False, 'count': 0}), 400
    crawl_pagination = coerce_bool(data.get('crawl_pagination'), default=True)
    max_pages = coerce_int(data.get('max_pages') or 10, 10, min_value=1, max_value=20)

    delay_ms = coerce_int(data.get('delay_ms') or 50, 50, min_value=0, max_value=5000)
    retries = coerce_int(data.get('retries') or 1, 1, min_value=1, max_value=5)
    verify_ssl = coerce_bool(data.get('verify_ssl'), default=True)
    use_curl = coerce_bool(data.get('use_curl'), default=True)
    use_browser = coerce_bool(data.get('use_browser'), default=False)
    use_parallel = coerce_bool(data.get('use_parallel'), default=True)
    enrich_details = True

    rules = {
        "add_percent": coerce_float(data.get('add_percent') or 0.0, 0.0),
        "percent_off": coerce_float(data.get('percent_off') or 0.0, 0.0),
        "absolute_off": coerce_float(data.get('absolute_off') or 0.0, 0.0),
    }
    drop_pct = coerce_float(data.get('drop_pct') or 10.0, 10.0, min_value=1.0, max_value=90.0)
    result = execute_scrape_workflow(
        urls_raw,
        crawl_pagination=crawl_pagination,
        max_pages=max_pages,
        delay_ms=delay_ms,
        retries=retries,
        verify_ssl=verify_ssl,
        use_curl=use_curl,
        use_browser=use_browser,
        use_parallel=use_parallel,
        enrich_details=enrich_details,
        rules=rules,
        drop_pct=drop_pct,
    )
    if result.get('error') and not result.get('items'):
        return jsonify(result), 400
    return jsonify(result), 200

@app.post('/api/export/xlsx')
@require_login
def export_xlsx():
    data = request.get_json(silent=True) or {}
    rows = data.get('rows') or []
    wb = Workbook()
    ws = wb.active
    ws.title = "Extract"
    headers = []
    # dynamic headers from keys (preserve a friendly order if present)
    preferred = ["image_url","title","price","original","percent_off","absolute_off","url","source","model"]
    excluded_export_fields = {"adjusted_price", "final"}
    if rows:
        keys = list(rows[0].keys())
        for k in preferred:
            if k in keys and k not in headers and k not in excluded_export_fields:
                headers.append(k)
        for k in keys:
            if k not in headers and k not in excluded_export_fields:
                headers.append(k)
    else:
        headers = preferred
    ws.append([k for k in headers])
    for r in rows:
        row_data = r if isinstance(r, dict) else {}
        ws.append([clean_excel_cell(row_data.get(k, "")) for k in headers])
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(bio, as_attachment=True, download_name="export.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.post('/api/comparison/upload')
@require_login
def upload_comparison_file():
    """Accept a CSV/XLSX file and return normalized comparison rows."""
    uploaded = request.files.get('file')
    if not uploaded or not uploaded.filename:
        return jsonify({'status': 'error', 'error': 'No file uploaded.'}), 400

    filename = uploaded.filename
    ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
    allowed = {'csv', 'txt', 'xlsx', 'xlsm', 'xltx', 'xltm'}
    if ext not in allowed:
        return jsonify({'status': 'error', 'error': 'Unsupported file type. Please upload a CSV or XLSX file.'}), 400

    raw = uploaded.read()
    if not raw:
        return jsonify({'status': 'error', 'error': 'Uploaded file is empty.'}), 400

    title_fields = ('title', 'name', 'product', 'product_name', 'clean_title', 'model')
    price_fields = ('final', 'price', 'compare_price', 'original', 'cost', 'amount', 'my_price', 'list_price', 'sale_price')
    site_fields = ('site', 'source', 'store', 'market', 'domain')
    url_fields = ('url', 'product_url', 'link')

    def extract_row(row: Dict[str, object]) -> Dict[str, object] | None:
        title_value = ''
        for field in title_fields:
            value = row.get(field)
            if value is None:
                continue
            text = str(value).strip()
            if text:
                title_value = text
                break
        if not title_value:
            return None

        price_value = None
        for field in price_fields:
            value = row.get(field)
            if value in (None, ''):
                continue
            if isinstance(value, (int, float)):
                price_value = float(value)
            else:
                price_value = parse_price_number(str(value))
            if price_value is not None:
                break
        if price_value is None:
            return None

        site_value = ''
        for field in site_fields:
            value = row.get(field)
            if value in (None, ''):
                continue
            text = str(value).strip()
            if text:
                site_value = text
                break

        url_value = ''
        for field in url_fields:
            value = row.get(field)
            if value in (None, ''):
                continue
            text = str(value).strip()
            if text:
                url_value = text
                break

        return {
            'title': title_value,
            'price': float(round(price_value, 4)),
            'site': site_value,
            'url': url_value
        }

    extracted: List[Dict[str, object]] = []
    skipped = 0

    try:
        if ext in {'csv', 'txt'}:
            text = raw.decode('utf-8-sig', errors='ignore')
            reader = csv.DictReader(io.StringIO(text))
            if not reader.fieldnames:
                raise ValueError('No headers found in CSV file.')
            for row in reader:
                result = extract_row(row)
                if result is None:
                    skipped += 1
                    continue
                extracted.append(result)
        else:
            workbook = load_workbook(io.BytesIO(raw), data_only=True)
            sheet = workbook.active
            rows_iter = list(sheet.iter_rows(values_only=True))
            if not rows_iter:
                raise ValueError('Spreadsheet is empty.')

            headers_raw = rows_iter[0]
            headers = [str(h or '').strip().lower() for h in headers_raw]
            if not any(headers):
                raise ValueError('Header row is missing in the spreadsheet.')

            for row_values in rows_iter[1:]:
                row_dict = {}
                for idx, header in enumerate(headers):
                    if not header:
                        continue
                    value = row_values[idx] if idx < len(row_values) else None
                    row_dict[header] = value
                result = extract_row(row_dict)
                if result is None:
                    skipped += 1
                    continue
                extracted.append(result)
    except ValueError as ve:
        return jsonify({'status': 'error', 'error': str(ve)}), 400
    except Exception as exc:
        return jsonify({'status': 'error', 'error': f'Failed to process file: {exc}'}), 400

    if not extracted:
        return jsonify({'status': 'error', 'error': 'No valid rows found. Ensure the file includes both title and price columns.'}), 400

    message = f"Loaded {len(extracted)} comparison rows"
    if skipped:
        message += f" (skipped {skipped} rows without title or price)"

    return jsonify({'status': 'success', 'message': message, 'rows': extracted})

# -------- Main --------
def find_free_port(start=5000, end=5050):
    for p in range(start, end + 1):
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            s.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
            try:
                s.bind(('0.0.0.0', p))
                return p
            except OSError:
                continue
    return 0

# ===== USERS MANAGEMENT ROUTES =====

@app.route('/users')
@require_role('admin')
def users_page():
    """Render the Users management UI."""
    return render_template('users.html', current_route='users')

@app.route('/api/users', methods=['GET'])
@require_role('admin')
def get_users():
    """List all users."""
    from database import db_manager
    users = db_manager.get_all_users()
    # Don't send password_hash!
    for u in users:
        u.pop('password_hash', None)
    return jsonify(users)

@app.route('/api/users', methods=['POST'])
@require_role('admin')
def create_user():
    """Create a new user."""
    data = request.json or {}
    username = data.get('username', '').strip()
    password = data.get('password', '')
    role = data.get('role', 'viewer')

    if not username or not password:
        return jsonify({'error': 'Username and password are required'}), 400

    from werkzeug.security import generate_password_hash
    password_hash = generate_password_hash(password)

    from database import db_manager
    db = db_manager
    if db.get_user_by_username(username):
        return jsonify({'error': 'Username already exists'}), 409

    user_id = db.add_user(username, password_hash, role)
    if not user_id:
        return jsonify({'error': 'Database error'}), 500

    return jsonify({'id': user_id, 'username': username, 'role': role}), 201

@app.route('/api/users/<int:user_id>', methods=['PUT'])
@require_role('admin')
def update_user(user_id):
    """Update a user's role or password."""
    data = request.json or {}
    role = data.get('role')
    password = data.get('password')

    password_hash = None
    if password:
        from werkzeug.security import generate_password_hash
        password_hash = generate_password_hash(password)

    from database import db_manager
    success = db_manager.update_user(user_id, role=role, password_hash=password_hash)
    if success:
        # Prevent demoting the superadmin entirely from UI
        # But this is just basic protection, let's keep it simple.
        return jsonify({'success': True})
    return jsonify({'error': 'Failed to update'}), 500

@app.route('/api/users/<int:user_id>', methods=['DELETE'])
@require_role('admin')
def delete_user(user_id):
    """Delete a user."""
    # Prevent self-deletion
    from flask_login import current_user
    if str(user_id) == str(current_user.id):
        return jsonify({'error': 'Cannot delete yourself'}), 400

    from database import db_manager
    success = db_manager.delete_user(user_id)
    if success:
        return jsonify({'success': True})
    return jsonify({'error': 'Failed to delete'}), 500


if __name__ == '__main__':
    # Use PORT when provided, otherwise keep the app on the default local port.
    port = coerce_int(os.getenv("PORT", "5000"), 5000, min_value=1, max_value=65535)
    if not port:
        raise SystemExit("PORT must be between 1 and 65535.")

    debug_env = str(os.getenv("FLASK_DEBUG", "0")).strip().lower()
    debug_mode = debug_env not in {"0", "false", "no", "off"}
    register_shutdown_hooks()
    ensure_automation_scheduler_started()

    try:
        app.run(host='0.0.0.0', port=port, debug=debug_mode, threaded=True)
    finally:
        shutdown_background_services()
