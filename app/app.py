from flask import Flask, request, jsonify, render_template, send_file, url_for, Response
import requests, re, json, io, time, os, socket, datetime, csv
from urllib.parse import urlparse
from dataclasses import asdict
from typing import List, Dict, Tuple
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook, load_workbook
from PIL import Image, ImageOps, ImageDraw, ImageEnhance, ImageFont, PngImagePlugin
import mimetypes
import pytz
import threading
import uuid
import html
from xml.etree import ElementTree as ET
from database import db_manager

# Import scraper engines (separated for maintainability)
from scraper_engine import (
    Item, build_session, scrape_url, scrape_urls_parallel,
    clean_text, parse_price_number, fmt_price, host_currency,
    scrape_category_all_pages, enrich_item_details as enrich_standard_item_details
)

# Import XCellParts specialized scraper
import xcell_scraper_engine

# Import TXParts specialized scraper
import txparts_scraper_engine

# Import Parts4Cells specialized scraper
import parts4cells_scraper_engine

app = Flask(__name__)

IMAGE_OUTPUT_MIME_TYPES = {
    'JPEG': 'image/jpeg',
    'PNG': 'image/png',
    'WEBP': 'image/webp',
    'GIF': 'image/gif',
    'BMP': 'image/bmp',
    'TIFF': 'image/tiff',
}
INLINE_PREVIEW_FORMATS = {'JPEG', 'PNG', 'WEBP', 'GIF'}
CONVERTED_FILE_TTL_SECONDS = 30 * 60
MAX_CONVERTED_FILE_CACHE_ITEMS = 256
CONVERTED_FILE_CACHE: Dict[str, Dict[str, object]] = {}
CONVERTED_FILE_CACHE_LOCK = threading.Lock()
PROXIED_IMAGE_TTL_SECONDS = 15 * 60
MAX_PROXIED_IMAGE_CACHE_ITEMS = 512
PROXIED_IMAGE_CACHE: Dict[str, Dict[str, object]] = {}
PROXIED_IMAGE_CACHE_LOCK = threading.Lock()
AUTO_DETAIL_SCAN_MAX_ITEMS = 20
RESAMPLE_LANCZOS = Image.Resampling.LANCZOS if hasattr(Image, 'Resampling') else Image.LANCZOS
WATERMARK_FONT_CANDIDATES = (
    r'C:\Windows\Fonts\arial.ttf',
    r'C:\Windows\Fonts\segoeui.ttf',
    'arial.ttf',
    'DejaVuSans.ttf',
)


def asset_url(filename: str) -> str:
    """Generate cache-busted static asset URLs so frontend fixes load immediately."""
    version = None
    try:
        static_path = os.path.join(app.static_folder or '', filename)
        if static_path and os.path.exists(static_path):
            version = int(os.path.getmtime(static_path))
    except Exception:
        version = None
    return url_for('static', filename=filename, v=version)


@app.context_processor
def inject_asset_url():
    return {'asset_url': asset_url}


def get_public_base_url() -> str:
    """Resolve the canonical external base URL when available, otherwise use the active request."""
    configured_url = os.getenv('PUBLIC_BASE_URL') or os.getenv('SITE_URL')
    if configured_url:
        return configured_url.rstrip('/')

    fly_app_name = os.getenv('FLY_APP_NAME')
    if fly_app_name and fly_app_name != 'local':
        return f"https://{fly_app_name}.fly.dev"

    return request.url_root.rstrip('/')


def get_template_lastmod(template_name: str) -> str:
    template_path = os.path.join(app.template_folder or '', template_name)
    try:
        modified_at = datetime.datetime.fromtimestamp(os.path.getmtime(template_path), tz=datetime.timezone.utc)
    except OSError:
        modified_at = datetime.datetime.now(datetime.timezone.utc)
    return modified_at.replace(microsecond=0).isoformat().replace('+00:00', 'Z')


def build_public_site_pages() -> List[Dict[str, str]]:
    return [
        {
            'path': '/',
            'endpoint': 'index',
            'changefreq': 'daily',
            'priority': '1.0',
            'lastmod': get_template_lastmod('index.html'),
        },
        {
            'path': '/history',
            'endpoint': 'history',
            'changefreq': 'daily',
            'priority': '0.8',
            'lastmod': get_template_lastmod('history.html'),
        },
        {
            'path': '/image-converter',
            'endpoint': 'image_converter',
            'changefreq': 'weekly',
            'priority': '0.8',
            'lastmod': get_template_lastmod('image_converter.html'),
        },
    ]

@app.teardown_appcontext
def close_db_connection(_exception=None):
    """Release SQLite connections after each request to avoid stale thread-local snapshots."""
    db_manager.close_connection()

def normalize_compare_text(value) -> str:
    return re.sub(r'\s+', ' ', str(value or '')).strip()


def normalize_stock_status_for_compare(value) -> str:
    """Collapse stock quantity variants into stable compare values."""
    text = normalize_compare_text(value).lower()
    if not text:
        return ''
    if 'outofstock' in text or 'out of stock' in text:
        return 'out of stock'
    if re.search(r'\b\d+\s+in stock\b', text) or 'in stock' in text:
        return 'in stock'
    if 'backorder' in text or 'back order' in text:
        return 'backorder'
    if 'preorder' in text or 'pre-order' in text:
        return 'preorder'
    return text


def has_specific_stock_detail(value) -> bool:
    """Detect whether a stock label carries quantity/low-stock detail instead of a generic state."""
    text = normalize_compare_text(value).lower()
    if not text:
        return False
    return bool(
        re.search(r'\b\d+\s+in stock\b', text)
        or 'left in stock' in text
        or re.search(r'\bonly\b.*\bin stock\b', text)
    )


def needs_specific_stock_refresh(items) -> bool:
    """Detect when a fast listing scrape only has generic stock values."""
    for item in items:
        item_dict = asdict(item) if hasattr(item, '__dict__') else dict(item or {})
        extra = item_dict.get('extra') if isinstance(item_dict.get('extra'), dict) else {}
        stock_status = normalize_compare_text(item_dict.get('stock_status') or extra.get('stock_status'))
        if not stock_status:
            return True
        if normalize_stock_status_for_compare(stock_status) == 'in stock' and not has_specific_stock_detail(stock_status):
            return True
    return False


def get_scraper_for_url(url: str):
    """Determine which scraper engine to use based on domain."""
    domain = urlparse(url).netloc.lower()

    if 'xcellparts.com' in domain:
        return 'xcell', xcell_scraper_engine
    if 'txparts.com' in domain:
        return 'txparts', txparts_scraper_engine
    if 'parts4cells.com' in domain:
        return 'parts4cells', parts4cells_scraper_engine
    if 'mobilesentrix' in domain:
        return 'standard', None
    return 'standard', None


def apply_enriched_item_data(item, enriched_data: Dict[str, object]):
    """Copy dataclass-exported enriched fields into another item object."""
    for key, value in (enriched_data or {}).items():
        if key == 'extra':
            current_extra = getattr(item, 'extra', None)
            if isinstance(current_extra, dict) and isinstance(value, dict):
                current_extra.update(value)
            elif hasattr(item, 'extra') and isinstance(value, dict):
                setattr(item, 'extra', dict(value))
            continue
        if not hasattr(item, key):
            continue
        if value is None:
            continue
        if isinstance(value, str) and not value.strip():
            continue
        setattr(item, key, value)
    return item


def hydrate_items_from_previous_history(items, previous_history: Dict | None) -> int:
    """Reuse saved metadata to avoid false diffs when deep detail scan is disabled."""
    if not previous_history:
        return 0

    previous_by_url = {}
    for previous_item in previous_history.get('items', []):
        url = normalize_compare_text(previous_item.get('url'))
        if url:
            previous_by_url[url] = previous_item

    hydrated = 0
    for item in items:
        item_url = normalize_compare_text(getattr(item, 'url', ''))
        previous_item = previous_by_url.get(item_url)
        if not previous_item:
            continue

        item_dict = asdict(item) if hasattr(item, '__dict__') else dict(item or {})
        extra = item_dict.get('extra') if isinstance(item_dict.get('extra'), dict) else {}
        patch = {}

        if not normalize_compare_text(item_dict.get('sku') or extra.get('sku')) and previous_item.get('sku'):
            patch['sku'] = previous_item.get('sku')
        current_stock = normalize_compare_text(item_dict.get('stock_status') or extra.get('stock_status'))
        previous_stock = normalize_compare_text(previous_item.get('stock_status'))
        if previous_stock and (
            not current_stock or
            (not has_specific_stock_detail(current_stock) and has_specific_stock_detail(previous_stock))
        ):
            patch['stock_status'] = previous_item.get('stock_status')
        if not normalize_compare_text(item_dict.get('description') or extra.get('description')) and previous_item.get('description'):
            patch['description'] = previous_item.get('description')

        if patch:
            apply_enriched_item_data(item, patch)
            if hasattr(item, 'extra') and isinstance(item.extra, dict):
                item.extra.update(patch)
            hydrated += 1

    return hydrated


def enrich_scraped_items(items, rules: Dict, retries: int, verify_ssl: bool, use_curl: bool, enrich_details: bool = False, logger=None):
    """Open each unique product detail page and merge richer metadata into scrape results."""
    if not enrich_details:
        return items, 0

    url_to_indexes: Dict[str, List[int]] = {}
    for idx, item in enumerate(items):
        item_url = normalize_compare_text(getattr(item, 'url', ''))
        if item_url:
            url_to_indexes.setdefault(item_url, []).append(idx)

    if not url_to_indexes:
        return items, 0

    thread_state = threading.local()
    created_sessions = []
    created_sessions_lock = threading.Lock()

    def get_thread_session(engine_type: str):
        if engine_type == 'parts4cells':
            return None

        sessions = getattr(thread_state, 'sessions', None)
        if sessions is None:
            sessions = {}
            thread_state.sessions = sessions

        if engine_type in sessions:
            return sessions[engine_type]

        if engine_type == 'xcell':
            session, _ = xcell_scraper_engine.build_session(retries=retries, verify_ssl=verify_ssl)
        elif engine_type == 'txparts':
            session, _ = txparts_scraper_engine.build_session(retries=retries, verify_ssl=verify_ssl)
        else:
            session, _ = build_session(retries=retries, verify_ssl=verify_ssl, use_curl=use_curl)

        sessions[engine_type] = session
        with created_sessions_lock:
            created_sessions.append(session)
        return session

    def enrich_one(item_url: str, item):
        engine_type, _ = get_scraper_for_url(item_url)
        if engine_type == 'xcell':
            enriched = xcell_scraper_engine.enrich_item_details(get_thread_session(engine_type), item, rules, logger)
        elif engine_type == 'txparts':
            enriched = txparts_scraper_engine.enrich_item_details(get_thread_session(engine_type), item, rules, logger)
        elif engine_type == 'parts4cells':
            enriched = parts4cells_scraper_engine.enrich_item_details(None, item, rules, logger)
        else:
            enriched = enrich_standard_item_details(get_thread_session(engine_type), item, rules, logger)
        return item_url, asdict(enriched)

    enriched_count = 0
    max_workers = min(10, max(1, len(url_to_indexes)))
    try:
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = {
                executor.submit(enrich_one, item_url, items[indexes[0]]): item_url
                for item_url, indexes in url_to_indexes.items()
            }
            for future in as_completed(futures):
                item_url = futures[future]
                try:
                    _, enriched_data = future.result()
                except Exception as exc:
                    if logger:
                        logger.warning(f"[detail] Failed to enrich {item_url}: {exc}")
                    continue

                for idx in url_to_indexes[item_url]:
                    apply_enriched_item_data(items[idx], enriched_data)
                enriched_count += len(url_to_indexes[item_url])
    finally:
        for session in created_sessions:
            if session and hasattr(session, 'close'):
                try:
                    session.close()
                except Exception:
                    pass

    if logger:
        logger.info(f"[detail] Enriched {enriched_count} item(s) across {len(url_to_indexes)} unique product URLs")
    return items, enriched_count

def get_effective_item_price(item) -> float | None:
    """Extract the current comparable price from any scraper item shape."""
    item_dict = asdict(item) if hasattr(item, '__dict__') else dict(item or {})

    for key in ('discounted_value', 'price_value', 'discounted', 'original'):
        value = item_dict.get(key)
        if isinstance(value, (int, float)) and float(value) > 0:
            return round(float(value), 2)

    for key in ('discounted_formatted', 'original_formatted', 'price_text'):
        value = parse_price_number(str(item_dict.get(key) or ''))
        if value is not None and float(value) > 0:
            return round(float(value), 2)

    return None

def build_price_drop_alerts(items, previous_prices: Dict[str, Dict], drop_pct: float) -> List[Dict]:
    """Compare current items against the latest saved DB prices for the same URL."""
    alerts = []

    for item in items:
        item_dict = asdict(item) if hasattr(item, '__dict__') else dict(item or {})
        url = str(item_dict.get('url') or '').strip()
        if not url:
            continue

        previous = previous_prices.get(url)
        if not previous or previous.get('price') is None:
            continue

        current_price = get_effective_item_price(item)
        previous_price = float(previous['price'])
        if current_price is None or previous_price <= 0 or current_price >= previous_price:
            continue

        percent_drop = round(((previous_price - current_price) / previous_price) * 100, 2)
        if percent_drop < drop_pct:
            continue

        alerts.append({
            'url': url,
            'title': item_dict.get('title', ''),
            'site': item_dict.get('site', ''),
            'previous_price': previous_price,
            'current_price': round(current_price, 2),
            'drop_percent': percent_drop,
            'previous_history_id': previous.get('history_id'),
            'previous_timestamp': previous.get('timestamp'),
        })

    alerts.sort(key=lambda alert: alert['drop_percent'], reverse=True)
    return alerts

def normalize_item_snapshot(item) -> Dict[str, object]:
    """Create a consistent comparison snapshot from current or historical item data."""
    item_dict = asdict(item) if hasattr(item, '__dict__') else dict(item or {})
    extra = item_dict.get('extra') if isinstance(item_dict.get('extra'), dict) else {}
    effective_price = get_effective_item_price(item_dict)
    return {
        'url': normalize_compare_text(item_dict.get('url')),
        'site': normalize_compare_text(item_dict.get('site')),
        'title': normalize_compare_text(item_dict.get('title')),
        'sku': normalize_compare_text(item_dict.get('sku') or extra.get('sku')),
        'stock_status': normalize_compare_text(item_dict.get('stock_status') or extra.get('stock_status')),
        'stock_status_compare': normalize_stock_status_for_compare(item_dict.get('stock_status') or extra.get('stock_status')),
        'description': normalize_compare_text(item_dict.get('description') or extra.get('description')),
        'effective_price': effective_price,
        'price_formatted': normalize_compare_text(
            item_dict.get('discounted_formatted')
            or item_dict.get('original_formatted')
            or item_dict.get('price_text')
        ),
    }

def build_session_comparison(previous_history: Dict | None, current_items) -> Dict:
    """Compare current scrape results against the latest previous run for the same target URLs."""
    current_snapshots = [normalize_item_snapshot(item) for item in current_items]
    previous_items = (previous_history or {}).get('items', []) if previous_history else []
    previous_snapshots = [normalize_item_snapshot(item) for item in previous_items]

    comparison = {
        'has_previous_run': previous_history is not None,
        'previous_history_id': previous_history.get('id') if previous_history else None,
        'previous_timestamp': previous_history.get('timestamp') if previous_history else None,
        'summary': {
            'previous_items': len(previous_snapshots),
            'current_items': len(current_snapshots),
            'added': 0,
            'removed': 0,
            'changed': 0,
            'price_changes': 0,
            'stock_changes': 0,
            'title_changes': 0,
            'sku_changes': 0,
            'description_changes': 0,
            'url_changes': 0,
        },
        'added': [],
        'removed': [],
        'changed': [],
    }

    if not previous_history:
        return comparison

    previous_by_url = {item['url']: item for item in previous_snapshots if item.get('url')}
    current_by_url = {item['url']: item for item in current_snapshots if item.get('url')}
    matched_previous = set()
    matched_current = set()
    matched_pairs = []

    for url in sorted(set(previous_by_url) & set(current_by_url)):
        matched_pairs.append((previous_by_url[url], current_by_url[url]))
        matched_previous.add(url)
        matched_current.add(url)

    remaining_previous = [item for item in previous_snapshots if item.get('url') not in matched_previous]
    remaining_current = [item for item in current_snapshots if item.get('url') not in matched_current]

    def match_by_key(previous_items_left, current_items_left, key_name: str):
        current_lookup = {}
        for item in current_items_left:
            key = normalize_compare_text(item.get(key_name))
            if key and key not in current_lookup:
                current_lookup[key] = item
        pairs = []
        used_current_urls = set()
        for prev_item in previous_items_left:
            key = normalize_compare_text(prev_item.get(key_name))
            current_item = current_lookup.get(key)
            if not key or not current_item or current_item.get('url') in used_current_urls:
                continue
            pairs.append((prev_item, current_item))
            used_current_urls.add(current_item.get('url'))
        return pairs

    for key_name in ('sku', 'title'):
        pairs = match_by_key(remaining_previous, remaining_current, key_name)
        if not pairs:
            continue
        for prev_item, current_item in pairs:
            matched_pairs.append((prev_item, current_item))
            matched_previous.add(prev_item.get('url'))
            matched_current.add(current_item.get('url'))
        remaining_previous = [item for item in remaining_previous if item.get('url') not in matched_previous]
        remaining_current = [item for item in remaining_current if item.get('url') not in matched_current]

    for item in remaining_current:
        comparison['added'].append(item)
    for item in remaining_previous:
        comparison['removed'].append(item)

    for prev_item, current_item in matched_pairs:
        field_changes = {}

        if prev_item['url'] != current_item['url']:
            field_changes['url'] = {'before': prev_item['url'], 'after': current_item['url']}
        if prev_item['title'] != current_item['title']:
            field_changes['title'] = {'before': prev_item['title'], 'after': current_item['title']}
        if prev_item['sku'] != current_item['sku']:
            field_changes['sku'] = {'before': prev_item['sku'], 'after': current_item['sku']}
        if prev_item['description'] != current_item['description']:
            field_changes['description'] = {'before': prev_item['description'], 'after': current_item['description']}
        if prev_item.get('stock_status_compare', prev_item['stock_status']) != current_item.get('stock_status_compare', current_item['stock_status']):
            field_changes['stock_status'] = {'before': prev_item['stock_status'], 'after': current_item['stock_status']}

        prev_price = prev_item.get('effective_price')
        current_price = current_item.get('effective_price')
        if prev_price is not None and current_price is not None and abs(prev_price - current_price) > 0.009:
            field_changes['price'] = {
                'before': prev_price,
                'after': current_price,
                'before_formatted': prev_item.get('price_formatted', ''),
                'after_formatted': current_item.get('price_formatted', ''),
            }

        if field_changes:
            comparison['changed'].append({
                'before': prev_item,
                'after': current_item,
                'changes': field_changes,
            })

    comparison['summary']['added'] = len(comparison['added'])
    comparison['summary']['removed'] = len(comparison['removed'])
    comparison['summary']['changed'] = len(comparison['changed'])
    comparison['summary']['price_changes'] = sum(1 for change in comparison['changed'] if 'price' in change['changes'])
    comparison['summary']['stock_changes'] = sum(1 for change in comparison['changed'] if 'stock_status' in change['changes'])
    comparison['summary']['title_changes'] = sum(1 for change in comparison['changed'] if 'title' in change['changes'])
    comparison['summary']['sku_changes'] = sum(1 for change in comparison['changed'] if 'sku' in change['changes'])
    comparison['summary']['description_changes'] = sum(1 for change in comparison['changed'] if 'description' in change['changes'])
    comparison['summary']['url_changes'] = sum(1 for change in comparison['changed'] if 'url' in change['changes'])
    return comparison

# -------- Image Processing --------
def normalize_target_format(target_format: str) -> str:
    target = str(target_format or 'JPEG').upper()
    return 'JPEG' if target == 'JPG' else target


def detect_source_format(image_data: bytes) -> str:
    """Detect image source format from Pillow or magic bytes."""
    try:
        with Image.open(io.BytesIO(image_data)) as img:
            return img.format if img.format else 'UNKNOWN'
    except Exception:
        if image_data.startswith(b'\x89PNG'):
            return 'PNG'
        if image_data.startswith(b'\xff\xd8\xff'):
            return 'JPEG'
        if image_data.startswith(b'RIFF') and b'WEBP' in image_data[:12]:
            return 'WEBP'
        if image_data.startswith(b'GIF8'):
            return 'GIF'
        if image_data.startswith(b'BM'):
            return 'BMP'
        if image_data[:4] in (b'II*\x00', b'MM\x00*'):
            return 'TIFF'
    return 'UNKNOWN'


def parse_converter_option(raw_value) -> Dict[str, object]:
    """Accept either a dict or a JSON string for image converter options."""
    if isinstance(raw_value, dict):
        return raw_value
    if raw_value in (None, '', b''):
        return {}
    try:
        parsed = json.loads(raw_value)
        return parsed if isinstance(parsed, dict) else {}
    except Exception:
        return {}


def parse_converter_options(raw_source) -> Dict[str, Dict[str, object]]:
    return {
        'resize': parse_converter_option(raw_source.get('resize')),
        'watermark': parse_converter_option(raw_source.get('watermark')),
        'transform': parse_converter_option(raw_source.get('transform')),
        'optimization': parse_converter_option(raw_source.get('optimization')),
        'naming': parse_converter_option(raw_source.get('naming')),
        'color_adjust': parse_converter_option(raw_source.get('color_adjust')),
    }


def sanitize_filename_stem(name: str) -> str:
    stem = re.sub(r'\.[A-Za-z0-9]+$', '', str(name or '').strip())
    stem = re.sub(r'[^A-Za-z0-9._-]+', '_', stem).strip('._')
    return stem or 'converted_image'


def build_output_filename(name_hint: str, target_format: str, naming: Dict[str, object] | None = None) -> str:
    naming = naming or {}
    stem = sanitize_filename_stem(name_hint)
    prefix = str(naming.get('prefix') or '')
    suffix = str(naming.get('suffix') or '')
    timestamp = ''
    if naming.get('add_timestamp'):
        timestamp = datetime.datetime.utcnow().strftime('_%Y%m%d_%H%M%S')
    ext = normalize_target_format(target_format).lower()
    return f"{prefix}{stem}{suffix}{timestamp}.{ext}"


def coerce_bool(value, default: bool = False) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    text = str(value).strip().lower()
    if text in {'1', 'true', 'yes', 'on'}:
        return True
    if text in {'0', 'false', 'no', 'off', ''}:
        return False
    return default if text == '' else True


def coerce_int(value, default: int = 0, min_value: int | None = None, max_value: int | None = None) -> int:
    try:
        result = int(value)
    except (TypeError, ValueError):
        result = default
    if min_value is not None:
        result = max(min_value, result)
    if max_value is not None:
        result = min(max_value, result)
    return result


def coerce_optional_positive_int(value) -> int:
    try:
        result = int(value)
    except (TypeError, ValueError):
        return 0
    return result if result > 0 else 0


def is_active_converter_option(option: Dict[str, object] | None) -> bool:
    if not option:
        return False
    if 'enabled' not in option:
        return True
    return coerce_bool(option.get('enabled'))


def normalize_resize_dimensions(img: Image.Image, width_value, height_value) -> Tuple[int, int]:
    width = coerce_optional_positive_int(width_value)
    height = coerce_optional_positive_int(height_value)
    if not width and not height:
        return img.width, img.height
    if not width:
        width = max(1, round(img.width * (height / max(1, img.height))))
    if not height:
        height = max(1, round(img.height * (width / max(1, img.width))))
    return width, height


def apply_resize_options(img: Image.Image, resize_options: Dict[str, object] | None) -> Image.Image:
    if not is_active_converter_option(resize_options):
        return img

    mode = str((resize_options or {}).get('mode') or 'fit').strip().lower()
    if mode == 'scale':
        scale_percent = coerce_int((resize_options or {}).get('scale_percent'), 100, min_value=1, max_value=500)
        target_width = max(1, round(img.width * (scale_percent / 100.0)))
        target_height = max(1, round(img.height * (scale_percent / 100.0)))
        return img.resize((target_width, target_height), RESAMPLE_LANCZOS)

    target_width, target_height = normalize_resize_dimensions(
        img,
        (resize_options or {}).get('width'),
        (resize_options or {}).get('height'),
    )
    if (target_width, target_height) == img.size:
        return img

    if mode == 'fill' and coerce_optional_positive_int((resize_options or {}).get('width')) and coerce_optional_positive_int((resize_options or {}).get('height')):
        return ImageOps.fit(img, (target_width, target_height), method=RESAMPLE_LANCZOS)
    if mode == 'fit' or mode == 'fill':
        return ImageOps.contain(img, (target_width, target_height), method=RESAMPLE_LANCZOS)
    return img.resize((target_width, target_height), RESAMPLE_LANCZOS)


def apply_transform_options(img: Image.Image, transform_options: Dict[str, object] | None) -> Image.Image:
    if not transform_options:
        return img

    rotation = coerce_int((transform_options or {}).get('rotation'), 0) % 360
    if rotation:
        img = img.rotate(-rotation, expand=True)
    if coerce_bool((transform_options or {}).get('flip_horizontal')):
        img = ImageOps.mirror(img)
    if coerce_bool((transform_options or {}).get('flip_vertical')):
        img = ImageOps.flip(img)
    return img


def load_watermark_font(font_size: int) -> ImageFont.ImageFont:
    for font_path in WATERMARK_FONT_CANDIDATES:
        try:
            return ImageFont.truetype(font_path, font_size)
        except Exception:
            continue
    return ImageFont.load_default()


def apply_watermark_options(img: Image.Image, watermark_options: Dict[str, object] | None) -> Image.Image:
    if not is_active_converter_option(watermark_options):
        return img

    text = str((watermark_options or {}).get('text') or '').strip()
    if not text:
        return img

    opacity = coerce_int((watermark_options or {}).get('opacity'), 50, min_value=0, max_value=100)
    if opacity <= 0:
        return img

    font_size = coerce_int((watermark_options or {}).get('font_size'), 24, min_value=8, max_value=200)
    font = load_watermark_font(font_size)
    position = str((watermark_options or {}).get('position') or 'bottom-right').strip().lower()

    base = img.convert('RGBA')
    overlay = Image.new('RGBA', base.size, (255, 255, 255, 0))
    draw = ImageDraw.Draw(overlay)
    bbox = draw.textbbox((0, 0), text, font=font)
    text_width = max(1, bbox[2] - bbox[0])
    text_height = max(1, bbox[3] - bbox[1])
    padding = max(8, round(min(base.size) * 0.04))

    positions = {
        'top-left': (padding, padding),
        'top-right': (max(padding, base.width - text_width - padding), padding),
        'bottom-left': (padding, max(padding, base.height - text_height - padding)),
        'bottom-right': (
            max(padding, base.width - text_width - padding),
            max(padding, base.height - text_height - padding),
        ),
        'center': (
            max(padding, (base.width - text_width) // 2),
            max(padding, (base.height - text_height) // 2),
        ),
    }
    x, y = positions.get(position, positions['bottom-right'])
    alpha = round(255 * (opacity / 100.0))
    shadow_alpha = max(32, round(alpha * 0.45))

    draw.text((x + 2, y + 2), text, font=font, fill=(0, 0, 0, shadow_alpha))
    draw.text((x, y), text, font=font, fill=(255, 255, 255, alpha))
    return Image.alpha_composite(base, overlay)


def split_alpha_channel(img: Image.Image) -> Tuple[Image.Image, Image.Image | None]:
    if img.mode in ('RGBA', 'LA'):
        return img.convert('RGB'), img.getchannel('A')
    if img.mode == 'P' and 'transparency' in img.info:
        rgba_img = img.convert('RGBA')
        return rgba_img.convert('RGB'), rgba_img.getchannel('A')
    if img.mode == 'RGB':
        return img.copy(), None
    return img.convert('RGB'), None


def apply_color_adjust_options(img: Image.Image, color_options: Dict[str, object] | None) -> Image.Image:
    if not is_active_converter_option(color_options):
        return img

    brightness = coerce_int((color_options or {}).get('brightness'), 0, min_value=-100, max_value=100)
    contrast = coerce_int((color_options or {}).get('contrast'), 0, min_value=-100, max_value=100)
    saturation = coerce_int((color_options or {}).get('saturation'), 0, min_value=-100, max_value=100)
    grayscale = coerce_bool((color_options or {}).get('grayscale'))
    if brightness == 0 and contrast == 0 and saturation == 0 and not grayscale:
        return img

    working, alpha = split_alpha_channel(img)
    if brightness != 0:
        working = ImageEnhance.Brightness(working).enhance(max(0.0, 1.0 + (brightness / 100.0)))
    if contrast != 0:
        working = ImageEnhance.Contrast(working).enhance(max(0.0, 1.0 + (contrast / 100.0)))
    if saturation != 0:
        working = ImageEnhance.Color(working).enhance(max(0.0, 1.0 + (saturation / 100.0)))
    if grayscale:
        working = ImageOps.grayscale(working).convert('RGB')
    if alpha is not None:
        rgba_working = working.convert('RGBA')
        rgba_working.putalpha(alpha)
        return rgba_working
    return working


def prepare_image_for_target(img: Image.Image, target: str) -> Image.Image:
    if target in ('JPEG', 'BMP'):
        if img.mode == 'P':
            img = img.convert('RGBA')
        if img.mode in ('RGBA', 'LA'):
            background = Image.new('RGB', img.size, (255, 255, 255))
            background.paste(img.convert('RGBA'), mask=img.getchannel('A'))
            return background
        if img.mode not in ('RGB', 'L'):
            return img.convert('RGB')
        return img
    if target == 'GIF':
        if img.mode not in ('P', 'L'):
            return img.convert('P', palette=Image.ADAPTIVE)
        return img
    if target in ('PNG', 'WEBP', 'TIFF'):
        if img.mode == 'P':
            return img.convert('RGBA' if 'transparency' in img.info else 'RGB')
        if img.mode == 'CMYK':
            return img.convert('RGB')
    return img


def build_metadata_save_kwargs(original_info: Dict[str, object], target: str, strip_metadata: bool) -> Dict[str, object]:
    if strip_metadata:
        return {}

    save_kwargs: Dict[str, object] = {}
    icc_profile = original_info.get('icc_profile')
    if icc_profile and target in {'JPEG', 'PNG', 'WEBP', 'TIFF'}:
        save_kwargs['icc_profile'] = icc_profile

    exif_data = original_info.get('exif')
    if exif_data and target in {'JPEG', 'PNG', 'WEBP', 'TIFF'}:
        save_kwargs['exif'] = exif_data

    dpi = original_info.get('dpi')
    if isinstance(dpi, (tuple, list)) and len(dpi) >= 2 and target in {'JPEG', 'PNG', 'WEBP', 'TIFF'}:
        save_kwargs['dpi'] = (dpi[0], dpi[1])

    if target == 'PNG':
        pnginfo = PngImagePlugin.PngInfo()
        added_text_chunk = False
        skipped_keys = {
            'icc_profile', 'exif', 'dpi', 'transparency', 'gamma', 'aspect', 'interlace',
            'background', 'loop', 'duration', 'disposal', 'extension', 'progressive',
            'progression', 'jfif', 'jfif_version', 'jfif_unit', 'jfif_density',
        }
        for key, value in original_info.items():
            if key in skipped_keys:
                continue
            if isinstance(value, str):
                pnginfo.add_text(str(key), value)
                added_text_chunk = True
            elif isinstance(value, bytes):
                try:
                    pnginfo.add_itxt(str(key), value.decode('utf-8'))
                    added_text_chunk = True
                except Exception:
                    continue
        if added_text_chunk:
            save_kwargs['pnginfo'] = pnginfo

    return save_kwargs


def cleanup_converted_file_cache() -> None:
    """Drop expired cached converted files to keep memory bounded."""
    now = time.time()
    with CONVERTED_FILE_CACHE_LOCK:
        expired = [
            file_id for file_id, entry in CONVERTED_FILE_CACHE.items()
            if now - float(entry.get('created_at', 0)) > CONVERTED_FILE_TTL_SECONDS
        ]
        for file_id in expired:
            CONVERTED_FILE_CACHE.pop(file_id, None)

        if len(CONVERTED_FILE_CACHE) <= MAX_CONVERTED_FILE_CACHE_ITEMS:
            return

        overflow = len(CONVERTED_FILE_CACHE) - MAX_CONVERTED_FILE_CACHE_ITEMS
        oldest = sorted(
            CONVERTED_FILE_CACHE.items(),
            key=lambda item: float(item[1].get('created_at', 0))
        )[:overflow]
        for file_id, _entry in oldest:
            CONVERTED_FILE_CACHE.pop(file_id, None)


def cache_converted_file(file_bytes: bytes, filename: str, target_format: str) -> str:
    cleanup_converted_file_cache()
    file_id = uuid.uuid4().hex
    with CONVERTED_FILE_CACHE_LOCK:
        CONVERTED_FILE_CACHE[file_id] = {
            'data': file_bytes,
            'filename': filename,
            'target_format': normalize_target_format(target_format),
            'mime_type': IMAGE_OUTPUT_MIME_TYPES.get(normalize_target_format(target_format), 'application/octet-stream'),
            'created_at': time.time(),
        }
    return file_id


def cleanup_proxied_image_cache() -> None:
    """Drop expired proxied images to keep memory bounded."""
    now = time.time()
    with PROXIED_IMAGE_CACHE_LOCK:
        expired = [
            cache_key for cache_key, entry in PROXIED_IMAGE_CACHE.items()
            if now - float(entry.get('created_at', 0)) > PROXIED_IMAGE_TTL_SECONDS
        ]
        for cache_key in expired:
            PROXIED_IMAGE_CACHE.pop(cache_key, None)

        if len(PROXIED_IMAGE_CACHE) <= MAX_PROXIED_IMAGE_CACHE_ITEMS:
            return

        overflow = len(PROXIED_IMAGE_CACHE) - MAX_PROXIED_IMAGE_CACHE_ITEMS
        oldest = sorted(
            PROXIED_IMAGE_CACHE.items(),
            key=lambda item: float(item[1].get('created_at', 0))
        )[:overflow]
        for cache_key, _entry in oldest:
            PROXIED_IMAGE_CACHE.pop(cache_key, None)


def fetch_proxied_image(image_url: str) -> Dict[str, object]:
    """Fetch and cache a remote image so the browser can load it from same-origin."""
    cleanup_proxied_image_cache()
    with PROXIED_IMAGE_CACHE_LOCK:
        cached = PROXIED_IMAGE_CACHE.get(image_url)
        if cached:
            return cached

    response = requests.get(
        image_url,
        timeout=30,
        headers={
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36',
            'Accept': 'image/avif,image/webp,image/apng,image/svg+xml,image/*,*/*;q=0.8',
            'Referer': request.host_url.rstrip('/') + '/',
        },
    )
    response.raise_for_status()

    mime_type = (response.headers.get('content-type') or '').split(';', 1)[0].strip()
    if not mime_type or mime_type == 'application/octet-stream':
        mime_type = mimetypes.guess_type(urlparse(image_url).path)[0] or 'application/octet-stream'
    if mime_type.startswith('text/html'):
        raise ValueError('Remote image URL returned HTML instead of image data')

    entry = {
        'data': response.content,
        'mime_type': mime_type,
        'created_at': time.time(),
    }
    with PROXIED_IMAGE_CACHE_LOCK:
        PROXIED_IMAGE_CACHE[image_url] = entry
    return entry


def convert_image_format(
    image_data: bytes,
    source_format: str,
    target_format: str = 'JPEG',
    quality: int = 85,
    options: Dict[str, Dict[str, object]] | None = None,
) -> Tuple[bytes, Dict]:
    """Convert image bytes and return both output bytes and image metadata."""
    try:
        target = normalize_target_format(target_format)
        options = options or {}
        optimization = options.get('optimization') or {}
        strip_metadata = coerce_bool(optimization.get('strip_metadata'), default=True)

        with Image.open(io.BytesIO(image_data)) as img:
            img.load()
            original_info = dict(img.info)
            img = ImageOps.exif_transpose(img)
            img = apply_resize_options(img, options.get('resize'))
            img = apply_transform_options(img, options.get('transform'))
            img = apply_color_adjust_options(img, options.get('color_adjust'))
            img = apply_watermark_options(img, options.get('watermark'))
            img = prepare_image_for_target(img, target)

            image_info = {
                'format': target,
                'mode': img.mode,
                'size': img.size,
                'width': img.width,
                'height': img.height,
                'has_transparency': img.mode in ('RGBA', 'LA') or 'transparency' in img.info,
            }

            output_buffer = io.BytesIO()
            save_kwargs = {'format': target}
            optimize_output = bool(optimization.get('optimize', False))
            save_kwargs.update(build_metadata_save_kwargs(original_info, target, strip_metadata))

            if target == 'JPEG':
                save_kwargs['quality'] = max(1, min(100, int(quality)))
                if optimize_output:
                    save_kwargs['optimize'] = True
                if optimization.get('progressive'):
                    save_kwargs['progressive'] = True
            elif target == 'PNG':
                save_kwargs['compress_level'] = 6 if optimize_output else 1
                if optimize_output:
                    save_kwargs['optimize'] = True
            elif target == 'WEBP':
                save_kwargs['quality'] = max(1, min(100, int(quality)))
                save_kwargs['method'] = 4 if optimize_output else 0
            elif target == 'TIFF' and optimize_output:
                save_kwargs['compression'] = 'tiff_deflate'

            img.save(output_buffer, **save_kwargs)
            return output_buffer.getvalue(), image_info
    except Exception as e:
        raise ValueError(f"Failed to convert image: {str(e)}")


def download_and_convert_image(
    image_url: str,
    target_format: str = 'JPEG',
    quality: int = 85,
    options: Dict[str, Dict[str, object]] | None = None,
) -> Tuple[bytes, str, str, Dict]:
    """Download image from URL and convert it to target format"""
    try:
        # Download image
        response = requests.get(image_url, timeout=30, headers={
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36'
        })
        response.raise_for_status()
        
        # Detect source format
        source_format = 'UNKNOWN'
        content_type = response.headers.get('content-type', '').lower()
        if 'webp' in content_type:
            source_format = 'WEBP'
        elif 'jpeg' in content_type or 'jpg' in content_type:
            source_format = 'JPEG'
        elif 'png' in content_type:
            source_format = 'PNG'
        elif 'gif' in content_type:
            source_format = 'GIF'
        else:
            # Try to detect from URL extension
            url_lower = image_url.lower()
            if '.webp' in url_lower:
                source_format = 'WEBP'
            elif '.jpg' in url_lower or '.jpeg' in url_lower:
                source_format = 'JPEG'
            elif '.png' in url_lower:
                source_format = 'PNG'
            elif '.gif' in url_lower:
                source_format = 'GIF'
        
        # Convert image
        converted_data, image_info = convert_image_format(
            response.content, source_format, target_format, quality, options
        )
        
        return converted_data, source_format, normalize_target_format(target_format), image_info
        
    except Exception as e:
        raise ValueError(f"Failed to download/convert image from {image_url}: {str(e)}")


def convert_image_from_bytes(
    image_data: bytes,
    target_format: str = 'JPEG',
    quality: int = 85,
    options: Dict[str, Dict[str, object]] | None = None,
) -> Tuple[bytes, str, str, Dict]:
    """Convert image from bytes data to target format"""
    try:
        source_format = detect_source_format(image_data)
        # Convert image
        converted_data, image_info = convert_image_format(
            image_data, source_format, target_format, quality, options
        )
        
        return converted_data, source_format, normalize_target_format(target_format), image_info
        
    except Exception as e:
        raise ValueError(f"Failed to convert uploaded image: {str(e)}")

def get_image_info(image_data: bytes) -> Dict:
    """Get information about an image"""
    try:
        with Image.open(io.BytesIO(image_data)) as img:
            return {
                'format': img.format,
                'mode': img.mode,
                'size': img.size,
                'width': img.width,
                'height': img.height,
                'has_transparency': img.mode in ('RGBA', 'LA') or 'transparency' in img.info
            }
    except Exception as e:
        return {'error': str(e)}


def build_converted_image_result(
    converted_data: bytes,
    source_format: str,
    target_format_used: str,
    image_info: Dict,
    filename: str,
    index: int,
    *,
    original_url: str | None = None,
    original_filename: str | None = None,
) -> Dict:
    file_id = cache_converted_file(converted_data, filename, target_format_used)
    preview_url = f"/api/converted/{file_id}" if target_format_used in INLINE_PREVIEW_FORMATS else ''
    return {
        'success': True,
        'index': index,
        'filename': sanitize_filename_stem(filename),
        'original_filename': original_filename or filename,
        'original_url': original_url or '',
        'source_format': source_format,
        'target_format': target_format_used,
        'image_info': image_info,
        'file_size': len(converted_data),
        'download_url': f"/api/converted/{file_id}?download=1",
        'preview_url': preview_url,
    }

# -------- Flask Routes --------
@app.get('/')
def index():
    return render_template('index.html')

@app.get('/sitemap.xml')
def sitemap():
    base_url = get_public_base_url()
    urlset = ET.Element('urlset', xmlns='http://www.sitemaps.org/schemas/sitemap/0.9')

    for page in build_public_site_pages():
        url_el = ET.SubElement(urlset, 'url')
        ET.SubElement(url_el, 'loc').text = f"{base_url}{page['path']}"
        ET.SubElement(url_el, 'lastmod').text = page['lastmod']
        ET.SubElement(url_el, 'changefreq').text = page['changefreq']
        ET.SubElement(url_el, 'priority').text = page['priority']

    xml_bytes = ET.tostring(urlset, encoding='utf-8', xml_declaration=True)
    return Response(xml_bytes, mimetype='application/xml')

@app.get('/robots.txt')
def robots():
    base_url = get_public_base_url()
    content = "\n".join([
        "User-agent: *",
        "Allow: /",
        "Disallow: /api/",
        "Disallow: /test-pagination",
        f"Sitemap: {base_url}/sitemap.xml",
        "",
    ])
    return Response(content, mimetype='text/plain')

@app.get('/history')
def history():
    return render_template('history.html')

@app.get('/image-converter')
def image_converter():
    return render_template('image_converter.html')

@app.get('/test-pagination')
def test_pagination():
    """Test endpoint to verify pagination is working"""
    url = "https://www.mobilesentrix.ca/replacement-parts/samsung/galaxy-s-series/galaxy-s25-ultra"
    rules = {"percent_off": 0.0, "absolute_off": 0.0, "add_percent": 0.0}
    
    sess, _ = build_session()
    items = scrape_category_all_pages(sess, url, rules, max_pages=5, delay_ms=200)
    
    return jsonify({
        "url": url,
        "total_items": len(items),
        "items_preview": [{"title": item.title, "url": item.url} for item in items[:10]]
    })

@app.get('/api/history')
def api_history():
    """Return history data from database"""
    try:
        page = int(request.args.get('page', 1))
        limit = int(request.args.get('limit', 20))
        offset = (page - 1) * limit
        
        histories = db_manager.get_history_list(limit=limit, offset=offset)
        return jsonify({
            'histories': histories,
            'page': page,
            'limit': limit,
            'total': len(histories)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.get('/api/history/<history_id>')
def api_history_detail(history_id):
    """Return specific history entry from database"""
    try:
        history = db_manager.get_history_detail(history_id)
        if not history:
            return jsonify({'error': 'History entry not found'}), 404
        return jsonify(history)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.post('/api/history/<history_id>/export/xlsx')
def api_history_export(history_id):
    """Export a specific history entry to XLSX"""
    try:
        history = db_manager.get_history_detail(history_id)
        if not history:
            return jsonify({'error': 'History entry not found'}), 404

        wb = Workbook()
        ws = wb.active
        ws.title = "Session"

        urls = history.get('urls', [])
        rules = history.get('rules', {}) or {}
        items = history.get('items', [])
        timestamp = history.get('timestamp')

        tz_label = "Pakistan Standard Time (UTC+05:00)"
        ts_display = timestamp
        if timestamp:
            try:
                # Parse timestamp and convert to Pakistan time
                ts_str = str(timestamp)
                if '+' in ts_str or 'Z' in ts_str:
                    # Has timezone info
                    ts_obj = datetime.datetime.fromisoformat(ts_str.replace('Z', '+00:00'))
                    if ts_obj.tzinfo is None:
                        ts_obj = pytz.UTC.localize(ts_obj)
                    # Convert to Pakistan timezone
                    pakistan_tz = pytz.timezone('Asia/Karachi')
                    ts_pakistan = ts_obj.astimezone(pakistan_tz)
                else:
                    # Assume it's already in Pakistan time
                    ts_obj = datetime.datetime.fromisoformat(ts_str)
                    pakistan_tz = pytz.timezone('Asia/Karachi')
                    ts_pakistan = pakistan_tz.localize(ts_obj)
                
                ts_display = ts_pakistan.strftime('%d %b %Y %I:%M %p PKT')
            except Exception as e:
                print(f"Error formatting timestamp in export: {e}")
                ts_display = str(timestamp)

        summary_rows = [
            ["Session ID", history.get('id', '')],
            ["Timestamp", ts_display],
            ["Timezone", tz_label],
            ["URLs Crawled", len(urls)],
            ["Items Captured", history.get('items_count', 0)],
            ["Add %", rules.get('add_percent', 0)],
            ["Discount %", rules.get('percent_off', 0)],
            ["Absolute Off", rules.get('absolute_off', 0)],
        ]

        for row in summary_rows:
            ws.append(row)

        if urls:
            ws.append([])
            ws.append(["Target URLs"])
            for url in urls:
                ws.append([url])

        ws.append([])
        headers = [
            "Title","Original Price","Discounted Price","Price Text","Price Value",
            "URL","Source","Site","Currency","Image URL"
        ]
        ws.append(headers)

        for item in items:
            ws.append([
                item.get('title',''),
                item.get('original_formatted',''),
                item.get('discounted_formatted',''),
                item.get('price_text',''),
                item.get('price_value',''),
                item.get('url',''),
                item.get('source',''),
                item.get('site',''),
                item.get('price_currency',''),
                item.get('image_url','')
            ])

        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        filename = f"history_{history_id}.xlsx"
        return send_file(
            bio,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

def _delete_history_response(history_id):
    normalized_history_id = str(history_id).strip()
    if not normalized_history_id:
        return jsonify({'error': 'History ID is required'}), 400

    try:
        success = db_manager.delete_history(normalized_history_id)
        if success:
            return jsonify({'success': True, 'history_id': normalized_history_id})
        return jsonify({'error': 'History entry not found'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.delete('/api/history/<history_id>')
def api_delete_history(history_id):
    """Delete history entry from database"""
    return _delete_history_response(history_id)

@app.post('/api/history/<history_id>/delete')
def api_delete_history_fallback(history_id):
    """Delete history entry from database (POST fallback when DELETE is blocked)"""
    return _delete_history_response(history_id)

@app.get('/api/statistics')
def api_statistics():
    """Get database statistics"""
    try:
        stats = db_manager.get_statistics()
        return jsonify(stats)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.post('/api/search')
def api_search():
    """Search items in database"""
    try:
        data = request.get_json(silent=True) or {}
        query = data.get('query', '').strip()
        limit = int(data.get('limit', 100))
        
        if not query:
            return jsonify({'error': 'Search query is required'}), 400
        
        items = db_manager.search_items(query, limit)
        return jsonify({
            'query': query,
            'results': items,
            'count': len(items)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.post('/api/cleanup')
def api_cleanup():
    """Cleanup old database entries"""
    try:
        data = request.get_json(silent=True) or {}
        days = int(data.get('days', 90))
        
        deleted_count = db_manager.cleanup_old_entries(days)
        return jsonify({
            'success': True,
            'deleted_entries': deleted_count,
            'days': days
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.post('/api/convert-image')
def api_convert_image():
    """Convert a single image from URL"""
    data = request.get_json(silent=True) or {}
    image_url = data.get('url', '').strip()
    target_format = normalize_target_format(data.get('format', 'JPEG'))
    quality = int(data.get('quality', 85))
    options = parse_converter_options(data)
    
    if not image_url:
        return jsonify({'error': 'Image URL is required'}), 400
    
    if target_format not in IMAGE_OUTPUT_MIME_TYPES:
        return jsonify({'error': 'Unsupported target format'}), 400
    
    try:
        converted_data, source_format, target_format_used, image_info = download_and_convert_image(
            image_url, target_format, quality, options
        )
        name_hint = os.path.basename(urlparse(image_url).path) or 'converted_image'
        result = build_converted_image_result(
            converted_data,
            source_format,
            target_format_used,
            image_info,
            build_output_filename(name_hint, target_format_used, options.get('naming')),
            0,
            original_url=image_url,
        )
        
        return jsonify({
            **result,
            'success': True,
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.post('/api/convert-images-batch')
def api_convert_images_batch():
    """Convert multiple images from URLs"""
    data = request.get_json(silent=True) or {}
    image_urls = data.get('urls', [])
    target_format = normalize_target_format(data.get('format', 'JPEG'))
    quality = int(data.get('quality', 85))
    options = parse_converter_options(data)
    
    if not image_urls:
        return jsonify({'error': 'Image URLs are required'}), 400
    
    if target_format not in IMAGE_OUTPUT_MIME_TYPES:
        return jsonify({'error': 'Unsupported target format'}), 400
    
    urls = image_urls[:20]
    results_by_index = {}

    def convert_url_image(index: int, image_url: str) -> Tuple[int, Dict]:
        try:
            converted_data, source_format, target_format_used, image_info = download_and_convert_image(
                image_url, target_format, quality, options
            )
            name_hint = os.path.basename(urlparse(image_url).path) or f'converted_image_{index + 1}'
            result = build_converted_image_result(
                converted_data,
                source_format,
                target_format_used,
                image_info,
                build_output_filename(name_hint, target_format_used, options.get('naming')),
                index,
                original_url=image_url,
            )
            return index, result
        except Exception as e:
            return index, {
                'success': False,
                'index': index,
                'original_url': image_url,
                'error': str(e)
            }

    max_workers = min(8, max(1, len(urls)))
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {
            executor.submit(convert_url_image, index, image_url): index
            for index, image_url in enumerate(urls)
        }
        for future in as_completed(futures):
            index, result = future.result()
            results_by_index[index] = result

    results = [results_by_index[index] for index in sorted(results_by_index)]
    
    return jsonify({
        'results': results,
        'total_processed': len(results),
        'successful': len([r for r in results if r['success']]),
        'failed': len([r for r in results if not r['success']])
    })

@app.post('/api/convert-files-batch')
def api_convert_files_batch():
    """Convert multiple uploaded image files"""
    if 'files' not in request.files:
        return jsonify({'error': 'No files uploaded'}), 400
    
    files = request.files.getlist('files')
    target_format = normalize_target_format(request.form.get('format', 'JPEG'))
    quality = int(request.form.get('quality', 85))
    options = parse_converter_options(request.form)
    
    if not files:
        return jsonify({'error': 'No files uploaded'}), 400
    
    if target_format not in IMAGE_OUTPUT_MIME_TYPES:
        return jsonify({'error': 'Unsupported target format'}), 400
    
    file_payloads = []
    results_by_index = {}

    for index, file in enumerate(files[:20]):
        if file.filename == '':
            results_by_index[index] = {
                'success': False,
                'index': index,
                'filename': 'unnamed',
                'error': 'Empty filename'
            }
            continue

        file_data = file.read()
        if len(file_data) == 0:
            results_by_index[index] = {
                'success': False,
                'index': index,
                'filename': file.filename,
                'error': 'Empty file'
            }
            continue

        file_payloads.append((index, file.filename, file_data))

    def convert_uploaded_file(index: int, filename: str, file_data: bytes) -> Tuple[int, Dict]:
        try:
            converted_data, source_format, target_format_used, image_info = convert_image_from_bytes(
                file_data, target_format, quality, options
            )
            output_filename = build_output_filename(filename, target_format_used, options.get('naming'))
            result = build_converted_image_result(
                converted_data,
                source_format,
                target_format_used,
                image_info,
                output_filename,
                index,
                original_filename=filename,
            )
            return index, result
        except Exception as e:
            return index, {
                'success': False,
                'index': index,
                'filename': filename,
                'error': str(e)
            }

    max_workers = min(8, max(1, len(file_payloads)))
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {
            executor.submit(convert_uploaded_file, index, filename, file_data): index
            for index, filename, file_data in file_payloads
        }
        for future in as_completed(futures):
            index, result = future.result()
            results_by_index[index] = result

    results = [results_by_index[index] for index in sorted(results_by_index)]
    
    return jsonify({
        'results': results,
        'total_processed': len(results),
        'successful': len([r for r in results if r['success']]),
        'failed': len([r for r in results if not r['success']])
    })

@app.get('/api/converted/<file_id>')
def download_converted_image(file_id):
    """Serve a converted image from the in-memory cache."""
    cleanup_converted_file_cache()
    with CONVERTED_FILE_CACHE_LOCK:
        entry = CONVERTED_FILE_CACHE.get(file_id)

    if not entry:
        return jsonify({'error': 'Converted file not found or expired'}), 404

    as_attachment = request.args.get('download') == '1'
    return send_file(
        io.BytesIO(entry['data']),
        as_attachment=as_attachment,
        download_name=entry['filename'],
        mimetype=entry['mime_type'],
        max_age=CONVERTED_FILE_TTL_SECONDS,
    )

@app.get('/api/image-proxy')
def proxy_remote_image():
    """Proxy remote images through the app to avoid browser-side hotlink failures."""
    image_url = (request.args.get('url') or '').strip()
    if not image_url:
        return jsonify({'error': 'Image URL is required'}), 400

    parsed = urlparse(image_url)
    if parsed.scheme not in {'http', 'https'}:
        return jsonify({'error': 'Only http and https image URLs are supported'}), 400

    try:
        entry = fetch_proxied_image(image_url)
    except Exception as exc:
        app.logger.warning(f"[image-proxy] Failed to fetch {image_url}: {exc}")
        return jsonify({'error': f'Failed to load image: {exc}'}), 502

    return send_file(
        io.BytesIO(entry['data']),
        mimetype=entry['mime_type'],
        max_age=PROXIED_IMAGE_TTL_SECONDS,
    )

@app.post('/api/scrape')
def api_scrape():
    data = request.get_json(silent=True) or {}
    urls_raw = data.get('urls') or ''
    crawl_pagination = bool(data.get('crawl_pagination', True))
    max_pages = int(data.get('max_pages') or 10)  # Reduced default from 20 to 10
    max_pages = 1 if max_pages < 1 else 20 if max_pages > 20 else max_pages  # Cap at 20 instead of 100

    delay_ms = int(data.get('delay_ms') or 50)
    retries = int(data.get('retries') or 1)
    verify_ssl = bool(data.get('verify_ssl', True))
    use_curl = bool(data.get('use_curl', False))
    use_parallel = bool(data.get('use_parallel', True))
    enrich_details = bool(data.get('enrich_details', False))

    rules = {
        "add_percent": float(data.get('add_percent') or 0.0),
        "percent_off": float(data.get('percent_off') or 0.0),
        "absolute_off": float(data.get('absolute_off') or 0.0),
    }
    drop_pct = max(1.0, float(data.get('drop_pct') or 10.0))

    urls = [u.strip() for u in (urls_raw.splitlines() if isinstance(urls_raw, str) else urls_raw) if u.strip()]
    seen_u = set(); urls = [u for u in urls if not (u in seen_u or seen_u.add(u))]
    previous_history = db_manager.get_latest_history_for_urls(urls)

    items: List[Item] = []
    engine_used = {}  # Track which engine was used for each URL
    
    if use_parallel and len(urls) > 1:
        # For parallel processing, group URLs by engine type
        xcell_urls = []
        txparts_urls = []
        parts4cells_urls = []
        standard_urls = []
        
        for url in urls:
            engine_type, _ = get_scraper_for_url(url)
            if engine_type == 'xcell':
                xcell_urls.append(url)
                engine_used[url] = 'xcell_scraper_engine'
            elif engine_type == 'txparts':
                txparts_urls.append(url)
                engine_used[url] = 'txparts_scraper_engine'
            elif engine_type == 'parts4cells':
                parts4cells_urls.append(url)
                engine_used[url] = 'parts4cells_scraper_engine'
            else:
                standard_urls.append(url)
                engine_used[url] = 'scraper_engine'
        
        # Process XCellParts URLs with xcell scraper
        if xcell_urls:
            app.logger.info(f"[engine] Using XCellParts scraper for {len(xcell_urls)} URLs")
            xcell_session, _ = xcell_scraper_engine.build_session(retries=retries, verify_ssl=verify_ssl)
            for url in xcell_urls:
                items.extend(xcell_scraper_engine.scrape_url(
                    xcell_session, url, rules, crawl_pagination, max_pages, delay_ms, app.logger
                ))
        
        # Process TXParts URLs with txparts scraper
        if txparts_urls:
            app.logger.info(f"[engine] Using TXParts scraper for {len(txparts_urls)} URLs")
            txparts_session, _ = txparts_scraper_engine.build_session(retries=retries, verify_ssl=verify_ssl)
            for url in txparts_urls:
                items.extend(txparts_scraper_engine.scrape_url(
                    txparts_session, url, rules, crawl_pagination, max_pages, delay_ms, app.logger
                ))

        # Process Parts4Cells URLs with parts4cells scraper
        if parts4cells_urls:
            app.logger.info(f"[engine] Using Parts4Cells scraper for {len(parts4cells_urls)} URLs")
            p4c_session, _ = parts4cells_scraper_engine.build_session(retries=retries, verify_ssl=verify_ssl)
            for url in parts4cells_urls:
                items.extend(parts4cells_scraper_engine.scrape_url(
                    p4c_session, url, rules, crawl_pagination, max_pages, delay_ms, app.logger
                ))
        
        # Process standard URLs with main scraper (parallel)
        if standard_urls:
            app.logger.info(f"[engine] Using standard scraper for {len(standard_urls)} URLs")
            max_workers = min(3, len(standard_urls))
            items.extend(scrape_urls_parallel(
                standard_urls, rules, crawl_pagination, max_pages, delay_ms,
                retries, verify_ssl, use_curl, max_workers, app.logger
            ))
        
        using_curl = use_curl
    else:
        # Sequential processing
        for u in urls:
            engine_type, engine_module = get_scraper_for_url(u)
            
            if engine_type == 'xcell':
                # Use XCellParts scraper
                app.logger.info(f"[engine] Using XCellParts scraper for: {u}")
                engine_used[u] = 'xcell_scraper_engine'
                xcell_session, _ = xcell_scraper_engine.build_session(retries=retries, verify_ssl=verify_ssl)
                items.extend(xcell_scraper_engine.scrape_url(
                    xcell_session, u, rules, crawl_pagination, max_pages, delay_ms, app.logger
                ))
            elif engine_type == 'txparts':
                # Use TXParts scraper
                app.logger.info(f"[engine] Using TXParts scraper for: {u}")
                engine_used[u] = 'txparts_scraper_engine'
                txparts_session, _ = txparts_scraper_engine.build_session(retries=retries, verify_ssl=verify_ssl)
                items.extend(txparts_scraper_engine.scrape_url(
                    txparts_session, u, rules, crawl_pagination, max_pages, delay_ms, app.logger
                ))
            elif engine_type == 'parts4cells':
                # Use Parts4Cells (Magento 2) scraper
                app.logger.info(f"[engine] Using Parts4Cells scraper for: {u}")
                engine_used[u] = 'parts4cells_scraper_engine'
                p4c_session, _ = parts4cells_scraper_engine.build_session(retries=retries, verify_ssl=verify_ssl)
                items.extend(parts4cells_scraper_engine.scrape_url(
                    p4c_session, u, rules, crawl_pagination, max_pages, delay_ms, app.logger
                ))
            else:
                # Use standard scraper
                app.logger.info(f"[engine] Using standard scraper for: {u}")
                engine_used[u] = 'scraper_engine'
                sess, using_curl = build_session(retries=retries, verify_ssl=verify_ssl, use_curl=use_curl)
                items.extend(scrape_url(sess, u, rules, crawl_pagination, max_pages, delay_ms, app.logger))

    hydrated_from_history = hydrate_items_from_previous_history(items, previous_history)
    auto_enrich_details = (
        not enrich_details
        and len(items) <= AUTO_DETAIL_SCAN_MAX_ITEMS
        and needs_specific_stock_refresh(items)
    )
    effective_enrich_details = enrich_details or auto_enrich_details
    if auto_enrich_details:
        app.logger.info(f"[detail] Auto-enabling detail scan for {len(items)} item(s) to capture stock detail")
    items, enriched_count = enrich_scraped_items(
        items, rules, retries, verify_ssl, use_curl, enrich_details=effective_enrich_details, logger=app.logger
    )

    previous_prices = {}
    if previous_history:
        for previous_item in previous_history.get('items', []):
            snapshot = normalize_item_snapshot(previous_item)
            url = snapshot.get('url')
            price = snapshot.get('effective_price')
            if not url or price is None:
                continue
            previous_prices[url] = {
                'price': price,
                'title': snapshot.get('title', ''),
                'site': snapshot.get('site', ''),
                'history_id': previous_history.get('id'),
                'timestamp': previous_history.get('timestamp'),
            }
    price_drops = build_price_drop_alerts(items, previous_prices, drop_pct)
    comparison = build_session_comparison(previous_history, items)

    # Store in database instead of memory
    history_id = str(int(time.time() * 1000))  # timestamp-based ID
    history_saved = False
    try:
        history_saved = db_manager.save_fetch_history(history_id, urls, items, rules)
        if not history_saved:
            app.logger.error("Failed to save fetch history to database")
    except Exception as e:
        app.logger.error(f"Database error: {e}")

    return jsonify({
        "rules": rules,
        "count": len(items),
        "drop_pct": drop_pct,
        "price_drops": price_drops,
        "comparison": comparison,
        "using_curl": using_curl if 'using_curl' in locals() else False,
        "using_parallel": use_parallel and len(urls) > 1,
        "engines_used": engine_used,  # Show which engine was used for each URL
        "enrich_details": effective_enrich_details,
        "enrich_details_requested": enrich_details,
        "auto_enrich_details": auto_enrich_details,
        "details_hydrated_from_history": hydrated_from_history,
        "details_enriched": enriched_count,
        "items": [asdict(i) for i in items],
        "history_id": history_id,
        "history_saved": history_saved
    })

@app.post('/api/export/xlsx')
def export_xlsx():
    data = request.get_json(silent=True) or {}
    rows = data.get('rows') or []
    wb = Workbook()
    ws = wb.active
    ws.title = "Extract"
    headers = []
    # dynamic headers from keys (preserve a friendly order if present)
    preferred = ["image_url","title","price","adjusted_price","original","percent_off","absolute_off","url","source","model"]
    if rows:
        keys = list(rows[0].keys())
        for k in preferred:
            if k in keys and k not in headers: headers.append(k)
        for k in keys:
            if k not in headers: headers.append(k)
    else:
        headers = preferred
    ws.append([k for k in headers])
    for r in rows:
        ws.append([r.get(k, "") for k in headers])
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(bio, as_attachment=True, download_name="export.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.post('/api/comparison/upload')
def upload_comparison_file():
    """Accept a CSV/XLSX file and return normalized comparison rows."""
    uploaded = request.files.get('file')
    if not uploaded or not uploaded.filename:
        return jsonify({'status': 'error', 'error': 'No file uploaded.'}), 400

    filename = uploaded.filename
    ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
    allowed = {'csv', 'txt', 'xlsx', 'xlsm', 'xltx', 'xltm'}
    if ext not in allowed:
        return jsonify({'status': 'error', 'error': 'Unsupported file type. Please upload a CSV or XLSX file.'}), 400

    raw = uploaded.read()
    if not raw:
        return jsonify({'status': 'error', 'error': 'Uploaded file is empty.'}), 400

    title_fields = ('title', 'name', 'product', 'product_name', 'clean_title', 'model')
    price_fields = ('final', 'price', 'compare_price', 'original', 'cost', 'amount', 'my_price', 'list_price', 'sale_price')
    site_fields = ('site', 'source', 'store', 'market', 'domain')
    url_fields = ('url', 'product_url', 'link')

    def extract_row(row: Dict[str, object]) -> Dict[str, object] | None:
        title_value = ''
        for field in title_fields:
            value = row.get(field)
            if value is None:
                continue
            text = str(value).strip()
            if text:
                title_value = text
                break
        if not title_value:
            return None

        price_value = None
        for field in price_fields:
            value = row.get(field)
            if value in (None, ''):
                continue
            if isinstance(value, (int, float)):
                price_value = float(value)
            else:
                price_value = parse_price_number(str(value))
            if price_value is not None:
                break
        if price_value is None:
            return None

        site_value = ''
        for field in site_fields:
            value = row.get(field)
            if value in (None, ''):
                continue
            text = str(value).strip()
            if text:
                site_value = text
                break

        url_value = ''
        for field in url_fields:
            value = row.get(field)
            if value in (None, ''):
                continue
            text = str(value).strip()
            if text:
                url_value = text
                break

        return {
            'title': title_value,
            'price': float(round(price_value, 4)),
            'site': site_value,
            'url': url_value
        }

    extracted: List[Dict[str, object]] = []
    skipped = 0

    try:
        if ext in {'csv', 'txt'}:
            text = raw.decode('utf-8-sig', errors='ignore')
            reader = csv.DictReader(io.StringIO(text))
            if not reader.fieldnames:
                raise ValueError('No headers found in CSV file.')
            for row in reader:
                result = extract_row(row)
                if result is None:
                    skipped += 1
                    continue
                extracted.append(result)
        else:
            workbook = load_workbook(io.BytesIO(raw), data_only=True)
            sheet = workbook.active
            rows_iter = list(sheet.iter_rows(values_only=True))
            if not rows_iter:
                raise ValueError('Spreadsheet is empty.')

            headers_raw = rows_iter[0]
            headers = [str(h or '').strip().lower() for h in headers_raw]
            if not any(headers):
                raise ValueError('Header row is missing in the spreadsheet.')

            for row_values in rows_iter[1:]:
                row_dict = {}
                for idx, header in enumerate(headers):
                    if not header:
                        continue
                    value = row_values[idx] if idx < len(row_values) else None
                    row_dict[header] = value
                result = extract_row(row_dict)
                if result is None:
                    skipped += 1
                    continue
                extracted.append(result)
    except ValueError as ve:
        return jsonify({'status': 'error', 'error': str(ve)}), 400
    except Exception as exc:
        return jsonify({'status': 'error', 'error': f'Failed to process file: {exc}'}), 400

    if not extracted:
        return jsonify({'status': 'error', 'error': 'No valid rows found. Ensure the file includes both title and price columns.'}), 400

    message = f"Loaded {len(extracted)} comparison rows"
    if skipped:
        message += f" (skipped {skipped} rows without title or price)"

    return jsonify({'status': 'success', 'message': message, 'rows': extracted})

# -------- Main --------
def find_free_port(start=5000, end=5050):
    for p in range(start, end + 1):
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            s.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
            try:
                s.bind(('0.0.0.0', p))
                return p
            except OSError:
                continue
    return 0

if __name__ == '__main__':
    # Get port from environment (for Fly.io) or find free port locally
    port = int(os.getenv("PORT", "0")) or find_free_port()
    if not port:
        raise SystemExit("No free port in 5000–5050. Set PORT env var to a free port.")
    
    # Check if running in production (Fly.io sets FLY_APP_NAME)
    is_production = os.getenv("FLY_APP_NAME") is not None
    debug_mode = not is_production
    
    app.run(host='0.0.0.0', port=port, debug=debug_mode)
