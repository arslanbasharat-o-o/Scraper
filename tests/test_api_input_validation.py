import importlib
import sys
from io import BytesIO
from types import SimpleNamespace

from openpyxl import load_workbook


def _fresh_app(tmp_path, monkeypatch):
    site_dbs = tmp_path / "site_dbs"
    monkeypatch.setenv("DATABASES_DIR", str(site_dbs))
    monkeypatch.delenv("DATABASE_PATH", raising=False)

    for module_name in ("app", "database"):
        sys.modules.pop(module_name, None)

    app_module = importlib.import_module("app")
    app_module.app.config["TESTING"] = True
    app_module.ensure_automation_scheduler_started = lambda: None
    return app_module


def test_scrape_bad_numeric_options_return_validation_error(tmp_path, monkeypatch):
    app_module = _fresh_app(tmp_path, monkeypatch)

    with app_module.app.test_client() as client:
        response = client.post("/api/scrape", json={
            "urls": "",
            "max_pages": "abc",
            "delay_ms": "abc",
            "retries": "abc",
            "drop_pct": "abc",
        })

    payload = response.get_json()

    assert response.status_code == 400
    assert payload["error"] == "At least one URL is required."
    assert payload["history_saved"] is False


def test_scrape_rejects_non_supplier_url_before_browser_launch(tmp_path, monkeypatch):
    app_module = _fresh_app(tmp_path, monkeypatch)

    with app_module.app.test_client() as client:
        response = client.post("/api/scrape", json={
            "urls": "http://127.0.0.1:5000/api/statistics",
        })

    payload = response.get_json()

    assert response.status_code == 400
    assert "approved supplier domain" in payload["error"]
    assert payload["history_saved"] is False


def test_supplier_url_validation_rejects_private_dns_result(tmp_path, monkeypatch):
    app_module = _fresh_app(tmp_path, monkeypatch)
    monkeypatch.setattr(
        app_module.socket,
        "getaddrinfo",
        lambda *_args, **_kwargs: [
            (app_module.socket.AF_INET, app_module.socket.SOCK_STREAM, 6, "", ("127.0.0.1", 443))
        ],
    )

    try:
        app_module.validate_supplier_remote_url("https://xcellparts.com/product")
    except ValueError as exc:
        assert "Private" in str(exc)
    else:
        raise AssertionError("Private supplier DNS result was accepted")


def test_image_proxy_rejects_unapproved_host(tmp_path, monkeypatch):
    app_module = _fresh_app(tmp_path, monkeypatch)

    with app_module.app.test_client() as client:
        response = client.get("/api/image-proxy?url=https://example.com/image.jpg")

    assert response.status_code == 400
    assert "approved supplier domain" in response.get_json()["error"]


def test_health_endpoint_and_security_headers(tmp_path, monkeypatch):
    app_module = _fresh_app(tmp_path, monkeypatch)

    with app_module.app.test_client() as client:
        response = client.get("/api/health")

    assert response.status_code == 200
    assert response.get_json()["browser_engine"] == "botasaurus"
    assert response.headers["X-Content-Type-Options"] == "nosniff"
    assert response.headers["X-Frame-Options"] == "SAMEORIGIN"
    assert response.headers["Cache-Control"] == "no-store"


def test_automation_runs_can_opt_into_compact_model_summary(tmp_path, monkeypatch):
    app_module = _fresh_app(tmp_path, monkeypatch)
    manager = app_module.db_manager
    job = manager.save_automation_job({
        "name": "Compact Payload",
        "scraper_key": "xcell",
        "category_query": "iphone",
        "root_url": "https://xcellparts.com/",
        "enabled": False,
    }, targets=[])
    run = manager.create_automation_run(job["id"], trigger_type="manual", target_urls=[])
    manager.update_automation_run_progress(run["id"], summary={
        "current_items": 2,
        "models": [
            {"model": "iPhone 15", "items": 1},
            {"model": "iPhone 16", "items": 1},
        ],
    })

    with app_module.app.test_client() as client:
        compact_response = client.get("/api/automation/runs?limit=1&include_models=0")
        compatible_response = client.get("/api/automation/runs?limit=1")

    compact_summary = compact_response.get_json()["runs"][0]["summary"]
    compatible_summary = compatible_response.get_json()["runs"][0]["summary"]
    assert compact_response.status_code == 200
    assert "models" not in compact_summary
    assert compact_summary["model_count"] == 2
    assert len(compatible_summary["models"]) == 2


def test_cross_origin_state_change_is_rejected(tmp_path, monkeypatch):
    app_module = _fresh_app(tmp_path, monkeypatch)

    with app_module.app.test_client() as client:
        response = client.post(
            "/api/search",
            json={"query": "screen"},
            headers={"Origin": "https://attacker.example"},
        )

    assert response.status_code == 403
    assert "Cross-origin" in response.get_json()["error"]


def test_cleanup_bad_days_uses_safe_default(tmp_path, monkeypatch):
    app_module = _fresh_app(tmp_path, monkeypatch)

    with app_module.app.test_client() as client:
        response = client.post(
            "/api/cleanup",
            json={"days": "abc"},
            headers={"X-Confirm-Destructive": "permanently-delete"},
        )

    payload = response.get_json()

    assert response.status_code == 200
    assert payload["success"] is True
    assert payload["days"] == 90


def test_cleanup_delete_all_explicitly_removes_recent_history(tmp_path, monkeypatch):
    app_module = _fresh_app(tmp_path, monkeypatch)
    manager = app_module.db_manager.managers["standard"]
    assert manager.save_fetch_history(
        "1784319000001",
        ["https://www.mobilesentrix.com/tablets"],
        [{"title": "Tablet Screen", "url": "https://www.mobilesentrix.com/tablets/item"}],
        {},
    )

    with app_module.app.test_client() as client:
        response = client.post(
            "/api/cleanup",
            json={"delete_all": True},
            headers={"X-Confirm-Destructive": "permanently-delete"},
        )

    payload = response.get_json()

    assert response.status_code == 200
    assert payload["success"] is True
    assert payload["delete_all"] is True
    assert payload["days"] == 99999
    assert payload["deleted_entries"] >= 1
    assert manager.get_history_detail("1784319000001") is None


def test_destructive_endpoint_requires_explicit_confirmation(tmp_path, monkeypatch):
    app_module = _fresh_app(tmp_path, monkeypatch)

    with app_module.app.test_client() as client:
        response = client.post("/api/cleanup", json={"days": 90})

    assert response.status_code == 428
    assert "confirmation" in response.get_json()["error"]


def test_xlsx_export_sanitizes_illegal_cell_characters(tmp_path, monkeypatch):
    app_module = _fresh_app(tmp_path, monkeypatch)

    with app_module.app.test_client() as client:
        response = client.post("/api/export/xlsx", json={
            "rows": [{
                "title": "Bad\x01Title",
                "description": "Description\x02with control",
                "url": "https://example.com/product",
            }]
        })

    assert response.status_code == 200
    assert response.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    wb = load_workbook(BytesIO(response.data))
    ws = wb.active
    assert ws["A2"].value == "BadTitle"
    assert ws["C2"].value == "Descriptionwith control"


def test_history_xlsx_export_sanitizes_illegal_cell_characters(tmp_path, monkeypatch):
    app_module = _fresh_app(tmp_path, monkeypatch)
    history_id = "1784319000000"
    item = app_module.Item(
        title="Bad\x01History",
        url="https://example.com/product",
        site="example.com",
        price_value=9.99,
        price_currency="USD",
        price_text="$9.99",
        discounted_value=9.99,
        discounted_formatted="$9.99",
        original_formatted="$9.99",
        source="test",
        image_url="",
    )
    item.description = "Description\x02with control"
    app_module.db_manager.save_fetch_history(history_id, ["https://example.com/product"], [item], {})

    with app_module.app.test_client() as client:
        response = client.post(f"/api/history/{history_id}/export/xlsx")

    assert response.status_code == 200
    assert response.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    wb = load_workbook(BytesIO(response.data))
    ws = wb.active
    assert ws["A2"].value == "BadHistory"
    assert ws["E2"].value == "Descriptionwith control"


def test_export_xlsx_standardized_headers_and_format(tmp_path, monkeypatch):
    app_module = _fresh_app(tmp_path, monkeypatch)

    headers = ["Title", "Price", "SKU", "Description", "Image", "URL", "Source Name", "Change Details"]
    row_data = {
        "Title": "LCD Assembly For iPhone 17e",
        "Price": "CA$23.91",
        "SKU": "107082127854",
        "Description": "LCD Assembly replacement part",
        "Image": "https://static.mobilesentrix.com/img.jpg",
        "URL": "https://www.mobilesentrix.ca/item",
        "Source Name": "www.mobilesentrix.ca",
        "Change Details": "Price: CA$20.00 -> CA$23.91",
    }

    with app_module.app.test_client() as client:
        response = client.post("/api/export/xlsx", json={
            "rows": [row_data],
            "headers": headers,
        })

    assert response.status_code == 200
    assert response.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    wb = load_workbook(BytesIO(response.data))
    ws = wb.active
    assert ws.title == "Extract"
    assert ws.freeze_panes == "A2"

    exported_headers = [cell.value for cell in ws[1]]
    assert exported_headers == headers

    exported_row = [cell.value for cell in ws[2]]
    assert exported_row == [row_data[h] for h in headers]

