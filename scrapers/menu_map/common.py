from __future__ import annotations

import argparse
import asyncio
import csv
import html
import json
import logging
import re
import time
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Awaitable, Callable, Iterable
from urllib.parse import parse_qsl, urlencode, urljoin, urlparse, urlunparse

import aiohttp
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from scrapers.botasaurus_wrapper import (
    Driver,
    browser,
    close_botasaurus_driver,
    resolve_chrome_executable,
    resolve_chrome_profile_root,
)
from scrapers.browser_fetcher import MOBILESENTRIX_CANADA_POPUP_DISMISS_JS


TRACKING_PARAMS = {
    "utm_source",
    "utm_medium",
    "utm_campaign",
    "utm_term",
    "utm_content",
    "gclid",
    "fbclid",
    "srsltid",
}

BAD_TEXT_RE = re.compile(
    r"^(search|cart|checkout|login|sign in|my account|account|logo|home|shipping|"
    r"privacy statement|skip to main content|learn more|shop now|view all models)$",
    re.I,
)
BAD_HREF_RE = re.compile(r"^(?:javascript:|mailto:|tel:|#?$)", re.I)


@dataclass(slots=True)
class SiteConfig:
    website: str
    website_url: str
    output_slug: str
    base_url: str
    parent_nav_selector: str
    parent_item_selector: str
    mega_menu_selector: str
    sub_child_panel_selector: str
    sub_child_item_selector: str
    active_sub_child_selector: str
    child_panel_selector: str
    child_link_selector: str
    scroll_container_selector: str
    menu_close_selector: str
    search_selector: str
    mobile_menu_selector: str
    parent_open_method: str = "click"
    sub_child_activation_method: str = "dom-inspection"


@dataclass(slots=True)
class CategoryRecord:
    website: str
    website_url: str
    parent_name: str
    parent_url: str
    parent_display_order: int
    parent_open_method: str
    sub_child_name: str = ""
    sub_child_url: str = ""
    sub_child_display_order: int = 0
    sub_child_activation_method: str = ""
    child_name: str = ""
    child_url: str = ""
    child_display_order: int = 0
    column_number: int = 0
    row_number: int = 0
    hierarchy_level: int = 1
    discovery_method: str = "dom-inspection"
    source_selector: str = ""
    normalized_url: str = ""
    url_missing: bool = False
    is_duplicate: bool = False
    scraped_at: str = ""
    http_status: int | None = None
    final_url: str = ""
    redirect_url: str = ""
    response_error: str = ""
    validation_timestamp: str = ""


@dataclass(slots=True)
class ScrapeError:
    website: str
    parent_name: str = ""
    sub_child_name: str = ""
    failed_action: str = ""
    error_type: str = ""
    error_message: str = ""
    retry_count: int = 0
    screenshot_path: str = ""
    html_snapshot_path: str = ""
    timestamp: str = ""


@dataclass(slots=True)
class ScrapeResult:
    records: list[CategoryRecord] = field(default_factory=list)
    errors: list[ScrapeError] = field(default_factory=list)
    inspection: list[dict[str, Any]] = field(default_factory=list)
    start_time: str = ""
    completion_time: str = ""
    runtime_seconds: float = 0.0
    browser_version: str = ""
    preserve_previous_output: bool = False


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def normalize_text(value: Any) -> str:
    text = html.unescape(str(value or ""))
    text = re.sub(r"[\u200b\u200c\u200d\ufeff]", "", text)
    text = re.sub(r"[\u2190-\u21ff\u25b2-\u25c4]+", " ", text)
    text = re.sub(r"\b(New|Sale)\b\s*$", "", text, flags=re.I)
    return re.sub(r"\s+", " ", text).strip()


def is_meaningful_name(value: str) -> bool:
    text = normalize_text(value)
    return bool(text and len(text) > 1 and not BAD_TEXT_RE.match(text))


def normalize_url(raw_url: str, base_url: str) -> str:
    raw_url = html.unescape((raw_url or "").strip())
    if not raw_url or BAD_HREF_RE.match(raw_url):
        return ""
    absolute = urljoin(base_url, raw_url)
    parsed = urlparse(absolute)
    if not parsed.scheme or not parsed.netloc:
        return ""
    netloc = parsed.netloc.lower()
    path = parsed.path
    if path != "/":
        path = path.rstrip("/")
    query_pairs = [
        (k, v)
        for k, v in parse_qsl(parsed.query, keep_blank_values=True)
        if k.lower() not in TRACKING_PARAMS
    ]
    query = urlencode(sorted(query_pairs), doseq=True)
    return urlunparse((parsed.scheme, netloc, path, "", query, ""))


def normalized_key(*parts: str) -> str:
    return "|".join(re.sub(r"\s+", " ", (p or "").strip().lower()) for p in parts)


def safe_filename(value: str, fallback: str = "artifact") -> str:
    text = normalize_text(value).lower()
    text = re.sub(r"[^a-z0-9._-]+", "-", text).strip("-")
    return text[:100] or fallback


def setup_output_dir(base: str | Path, slug: str) -> Path:
    path = Path(base) / slug
    (path / "screenshots").mkdir(parents=True, exist_ok=True)
    (path / "html_snapshots").mkdir(parents=True, exist_ok=True)
    return path


def setup_logger(output_dir: Path, level: str) -> logging.Logger:
    logger = logging.getLogger(f"category_scraper.{output_dir.name}")
    logger.handlers.clear()
    logger.setLevel(getattr(logging, level.upper(), logging.INFO))
    fmt = logging.Formatter("%(asctime)s %(levelname)s %(message)s")
    file_handler = logging.FileHandler(output_dir / "scraper.log", encoding="utf-8")
    file_handler.setFormatter(fmt)
    console_handler = logging.StreamHandler()
    console_handler.setFormatter(fmt)
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)
    return logger


class BotasaurusLocator:
    """Small async locator adapter used by the menu-map extractors."""

    def __init__(self, driver: Driver, selector: str, index: int | None = None):
        self.driver = driver
        self.selector = selector
        self.index = index

    def nth(self, index: int) -> "BotasaurusLocator":
        return BotasaurusLocator(self.driver, self.selector, index)

    def _elements(self):
        return self.driver.select_all(self.selector, wait=1) or []

    def _element(self):
        elements = self._elements()
        index = self.index if self.index is not None else 0
        if index < 0 or index >= len(elements):
            raise LookupError(f"No element {index} for selector {self.selector!r}")
        return elements[index]

    async def count(self) -> int:
        return len(self._elements())

    async def hover(self) -> None:
        self._element().move_mouse_here(is_jump=True)

    async def click(self, force: bool = False) -> None:
        del force
        self._element().click()

    async def focus(self) -> None:
        self._element().run_js("(element) => { element.focus(); return true; }")


class BotasaurusPage:
    """Async page adapter backed exclusively by Botasaurus."""

    def __init__(self, driver: Driver):
        self.driver = driver
        self.default_timeout_ms = 60_000

    def set_default_timeout(self, timeout: int) -> None:
        self.default_timeout_ms = max(1, int(timeout))

    def locator(self, selector: str) -> BotasaurusLocator:
        return BotasaurusLocator(self.driver, selector)

    async def goto(self, url: str, *, wait_until: str = "domcontentloaded", timeout: int = 60_000) -> None:
        del wait_until
        self.driver.get(url, timeout=max(1, int(timeout / 1000)))

    async def wait_for_load_state(self, state: str, *, timeout: int = 15_000) -> None:
        del state, timeout
        self.driver.run_js("return document.readyState")

    async def wait_for_timeout(self, delay_ms: int) -> None:
        self.driver.sleep(max(0, int(delay_ms)) / 1000)

    async def evaluate(self, script: str, arg: Any = None) -> Any:
        source = str(script or "").strip()
        if "=>" in source and not source.rstrip(";").endswith(")()"):
            if arg is None:
                source = f"return ({source})()"
            else:
                arg_literal = json.dumps(arg, ensure_ascii=False)
                source = f"const __scraper_arg = {arg_literal}; return ({source})(__scraper_arg)"
                arg = None
        return self.driver.run_js(source, args=arg)

    async def screenshot(self, *, path: str, full_page: bool = True) -> None:
        del full_page
        self.driver.save_screenshot(filename=path)

    async def content(self) -> str:
        return self.driver.page_html or ""

    async def title(self) -> str:
        return self.driver.title or ""


Page = BotasaurusPage


async def save_screenshot(page: Page, output_dir: Path, label: str) -> str:
    path = output_dir / "screenshots" / f"{safe_filename(label)}-{int(time.time())}.png"
    await page.screenshot(path=str(path), full_page=True)
    return str(path)


async def save_html_snapshot(page: Page, output_dir: Path, label: str) -> str:
    path = output_dir / "html_snapshots" / f"{safe_filename(label)}-{int(time.time())}.html"
    path.write_text(await page.content(), encoding="utf-8")
    return str(path)


async def record_error(
    errors: list[ScrapeError],
    page: Page,
    output_dir: Path,
    website: str,
    failed_action: str,
    exc: BaseException,
    parent_name: str = "",
    sub_child_name: str = "",
    retry_count: int = 0,
) -> None:
    screenshot = ""
    snapshot = ""
    try:
        screenshot = await save_screenshot(page, output_dir, f"{failed_action}-{parent_name}-{sub_child_name}")
        snapshot = await save_html_snapshot(page, output_dir, f"{failed_action}-{parent_name}-{sub_child_name}")
    except Exception:
        pass
    errors.append(
        ScrapeError(
            website=website,
            parent_name=parent_name,
            sub_child_name=sub_child_name,
            failed_action=failed_action,
            error_type=type(exc).__name__,
            error_message=str(exc),
            retry_count=retry_count,
            screenshot_path=screenshot,
            html_snapshot_path=snapshot,
            timestamp=utc_now(),
        )
    )


async def click_or_hover(page: Page, selector: str, index: int, method: str, delay_ms: int) -> None:
    locator = page.locator(selector).nth(index)
    if method == "hover":
        await locator.hover()
    elif method == "hover-and-click":
        await locator.hover()
        await page.wait_for_timeout(delay_ms)
        await locator.click(force=True)
    elif method == "focus":
        await locator.focus()
    else:
        await locator.click(force=True)
    await page.wait_for_timeout(delay_ms)


async def collect_dom_inspection(page: Page, selectors: Iterable[str]) -> list[dict[str, Any]]:
    selector_list = list(dict.fromkeys(s for s in selectors if s))
    return await page.evaluate(
        """
        (selectors) => {
          const textOf = el => (el.innerText || el.textContent || el.getAttribute('aria-label') || '').replace(/\\s+/g, ' ').trim();
          const cssPath = el => {
            if (!el || !el.nodeType) return '';
            if (el.id) return '#' + CSS.escape(el.id);
            const parts = [];
            let n = el;
            for (let i = 0; n && n.nodeType === 1 && i < 7; i++, n = n.parentElement) {
              let p = n.tagName.toLowerCase();
              if (n.id) { p += '#' + CSS.escape(n.id); parts.unshift(p); break; }
              const cls = [...n.classList].slice(0, 4);
              if (cls.length) p += '.' + cls.map(c => CSS.escape(c)).join('.');
              parts.unshift(p);
            }
            return parts.join(' > ');
          };
          const xpath = el => {
            if (el.id) return `//*[@id="${el.id}"]`;
            const parts = [];
            for (; el && el.nodeType === 1; el = el.parentNode) {
              let ix = 1, sib = el.previousSibling;
              while (sib) { if (sib.nodeType === 1 && sib.nodeName === el.nodeName) ix++; sib = sib.previousSibling; }
              parts.unshift(`${el.nodeName.toLowerCase()}[${ix}]`);
            }
            return '/' + parts.join('/');
          };
          const visible = el => {
            const r = el.getBoundingClientRect(), s = getComputedStyle(el);
            return r.width > 1 && r.height > 1 && s.display !== 'none' && s.visibility !== 'hidden' && +s.opacity !== 0;
          };
          const picked = new Set();
          const elements = [];
          for (const selector of selectors) {
            for (const el of document.querySelectorAll(selector)) {
              if (!picked.has(el)) { picked.add(el); elements.push(el); }
            }
          }
          return elements.map((el, order) => {
            const r = el.getBoundingClientRect(), s = getComputedStyle(el);
            const p = el.parentElement;
            const nearestNav = el.closest('nav, header, [role="navigation"], #nav, #mainMenu, #ninjamenus4, #ast-hf-menu-1');
            return {
              order,
              tag_name: el.tagName.toLowerCase(),
              text: textOf(el),
              href: el.getAttribute('href') || '',
              id: el.id || '',
              class: el.className ? String(el.className) : '',
              role: el.getAttribute('role') || '',
              aria_label: el.getAttribute('aria-label') || '',
              aria_expanded: el.getAttribute('aria-expanded') || '',
              aria_controls: el.getAttribute('aria-controls') || '',
              aria_selected: el.getAttribute('aria-selected') || '',
              tabindex: el.getAttribute('tabindex') || '',
              data_attributes: Object.fromEntries([...el.attributes].filter(a => a.name.startsWith('data-')).map(a => [a.name, a.value])),
              bounding_box: {x:r.x, y:r.y, width:r.width, height:r.height},
              visibility: visible(el),
              computed_display: s.display,
              computed_position: s.position,
              computed_overflow_y: s.overflowY,
              scrollHeight: el.scrollHeight,
              clientHeight: el.clientHeight,
              scrollTop: el.scrollTop,
              parent_element_summary: p ? `${p.tagName.toLowerCase()}#${p.id || ''}.${p.className || ''}` : '',
              child_element_summary: [...el.children].slice(0, 5).map(c => `${c.tagName.toLowerCase()}#${c.id || ''}.${c.className || ''}`),
              nearest_navigation_container: nearestNav ? cssPath(nearestNav) : '',
              css_selector_candidate: cssPath(el),
              xpath_candidate: xpath(el),
            };
          });
        }
        """,
        selector_list,
    )


def write_inspection(output_dir: Path, inspection: list[dict[str, Any]]) -> None:
    (output_dir / "dom_inspection.json").write_text(
        json.dumps(inspection, indent=2, ensure_ascii=False), encoding="utf-8"
    )
    rows = []
    for item in inspection:
        rows.append(
            "<tr>"
            + "".join(
                f"<td>{html.escape(str(item.get(k, ''))[:1500])}</td>"
                for k in (
                    "order",
                    "tag_name",
                    "text",
                    "href",
                    "id",
                    "class",
                    "role",
                    "aria_expanded",
                    "computed_display",
                    "computed_overflow_y",
                    "scrollHeight",
                    "clientHeight",
                    "visibility",
                    "css_selector_candidate",
                    "xpath_candidate",
                )
            )
            + "</tr>"
        )
    output = """
    <html><head><meta charset="utf-8"><style>
    body{font-family:Arial,sans-serif} table{border-collapse:collapse;width:100%}
    th,td{border:1px solid #ddd;padding:6px;vertical-align:top;font-size:12px}
    th{position:sticky;top:0;background:#111;color:#fff}.true{background:#e8ffe8}
    </style></head><body><h1>Menu DOM Inspection</h1>
    <p>Use this report to identify parent nav, sub-child groups, child links, active states, and scrollable containers.</p>
    <table><thead><tr>
    """
    headers = [
        "order",
        "tag",
        "text",
        "href",
        "id",
        "class",
        "role",
        "aria-expanded",
        "display",
        "overflow-y",
        "scrollHeight",
        "clientHeight",
        "visible",
        "CSS selector",
        "XPath",
    ]
    output += "".join(f"<th>{h}</th>" for h in headers) + "</tr></thead><tbody>"
    output += "\n".join(rows) + "</tbody></table></body></html>"
    (output_dir / "dom_inspection.html").write_text(output, encoding="utf-8")


def mark_duplicates(records: list[CategoryRecord]) -> list[dict[str, Any]]:
    exact_counts = Counter(
        normalized_key(r.parent_name, r.sub_child_name, r.child_name, r.normalized_url)
        for r in records
    )
    url_paths: dict[str, set[str]] = defaultdict(set)
    name_urls: dict[str, set[str]] = defaultdict(set)
    for r in records:
        if r.normalized_url:
            path_key = normalized_key(r.parent_name, r.sub_child_name, r.child_name)
            url_paths[r.normalized_url].add(path_key)
            if r.child_name or r.sub_child_name or r.parent_name:
                name_urls[normalized_key(r.child_name or r.sub_child_name or r.parent_name)].add(r.normalized_url)
    duplicates = []
    for r in records:
        key = normalized_key(r.parent_name, r.sub_child_name, r.child_name, r.normalized_url)
        if exact_counts[key] > 1:
            r.is_duplicate = True
            duplicates.append({**asdict(r), "duplicate_type": "exact"})
        elif r.normalized_url and len(url_paths[r.normalized_url]) > 1:
            duplicates.append({**asdict(r), "duplicate_type": "cross_hierarchy_url"})
        elif r.normalized_url and len(name_urls[normalized_key(r.child_name or r.sub_child_name or r.parent_name)]) > 1:
            duplicates.append({**asdict(r), "duplicate_type": "same_name_different_url"})
    return duplicates


def parent_category_rows(records: list[CategoryRecord]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    seen: set[tuple[str, int]] = set()
    for record in records:
        key = (record.parent_name, record.parent_display_order)
        if key in seen:
            continue
        seen.add(key)
        row = asdict(record)
        row.update(
            {
                "sub_child_name": "",
                "sub_child_url": "",
                "sub_child_display_order": 0,
                "sub_child_activation_method": "",
                "child_name": "",
                "child_url": "",
                "child_display_order": 0,
                "column_number": 0,
                "row_number": 0,
                "hierarchy_level": 1,
                "normalized_url": record.parent_url,
                "url_missing": not bool(record.parent_url),
            }
        )
        rows.append(row)
    return rows


def sub_child_category_rows(records: list[CategoryRecord]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    seen: set[tuple[str, int, str, int]] = set()
    for record in records:
        if not record.sub_child_name:
            continue
        key = (
            record.parent_name,
            record.parent_display_order,
            record.sub_child_name,
            record.sub_child_display_order,
        )
        if key in seen:
            continue
        seen.add(key)
        row = asdict(record)
        row.update(
            {
                "child_name": "",
                "child_url": "",
                "child_display_order": 0,
                "column_number": 0,
                "row_number": 0,
                "hierarchy_level": 2,
                "normalized_url": record.sub_child_url,
                "url_missing": not bool(record.sub_child_url),
            }
        )
        rows.append(row)
    return rows


def nested_json(records: list[CategoryRecord]) -> list[dict[str, Any]]:
    parents: dict[tuple[str, int], dict[str, Any]] = {}
    sub_map: dict[tuple[str, int, str, int], dict[str, Any]] = {}
    for r in records:
        pkey = (r.parent_name, r.parent_display_order)
        if pkey not in parents:
            parents[pkey] = {
                "parent_name": r.parent_name,
                "parent_url": r.parent_url,
                "display_order": r.parent_display_order,
                "open_method": r.parent_open_method,
                "sub_children": [],
            }
        if not r.sub_child_name:
            continue
        skey = (r.parent_name, r.parent_display_order, r.sub_child_name, r.sub_child_display_order)
        if skey not in sub_map:
            sub = {
                "sub_child_name": r.sub_child_name,
                "sub_child_url": r.sub_child_url,
                "display_order": r.sub_child_display_order,
                "activation_method": r.sub_child_activation_method,
                "children": [],
            }
            sub_map[skey] = sub
            parents[pkey]["sub_children"].append(sub)
        if r.child_name:
            sub_map[skey]["children"].append(
                {
                    "child_name": r.child_name,
                    "child_url": r.child_url,
                    "display_order": r.child_display_order,
                    "column_number": r.column_number,
                    "row_number": r.row_number,
                }
            )
    return list(parents.values())


async def validate_urls(records: list[CategoryRecord], concurrency: int = 8, timeout: int = 12) -> None:
    urls = sorted({r.normalized_url for r in records if r.normalized_url})
    semaphore = asyncio.Semaphore(concurrency)

    async def fetch(session: aiohttp.ClientSession, url: str) -> dict[str, Any]:
        async with semaphore:
            for attempt in range(3):
                try:
                    async with session.head(url, allow_redirects=True, timeout=timeout) as resp:
                        return {"status": resp.status, "final_url": str(resp.url), "error": ""}
                except Exception as head_exc:
                    try:
                        async with session.get(url, allow_redirects=True, timeout=timeout) as resp:
                            return {"status": resp.status, "final_url": str(resp.url), "error": ""}
                    except Exception as get_exc:
                        if attempt == 2:
                            return {"status": None, "final_url": "", "error": f"{type(get_exc).__name__}: {get_exc or head_exc}"}
                        await asyncio.sleep(0.5 * (attempt + 1))
            return {"status": None, "final_url": "", "error": "validation failed"}

    async with aiohttp.ClientSession(headers={"User-Agent": "Mozilla/5.0 category-menu-scraper"}) as session:
        results = dict(zip(urls, await asyncio.gather(*(fetch(session, url) for url in urls))))
    stamp = utc_now()
    for r in records:
        result = results.get(r.normalized_url)
        if not result:
            continue
        r.http_status = result["status"]
        r.final_url = result["final_url"]
        r.redirect_url = result["final_url"] if result["final_url"] and result["final_url"] != r.normalized_url else ""
        r.response_error = result["error"]
        r.validation_timestamp = stamp


def export_outputs(
    config: SiteConfig,
    output_dir: Path,
    result: ScrapeResult,
    headless: bool,
    duplicates: list[dict[str, Any]],
) -> None:
    records = result.records
    all_rows = [asdict(r) for r in records]
    parent_rows = parent_category_rows(records)
    sub_child_rows = sub_child_category_rows(records)
    child_rows = [asdict(r) for r in records if r.hierarchy_level == 3]
    fieldnames = list(CategoryRecord.__dataclass_fields__.keys())
    with (output_dir / "categories.csv").open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(all_rows)
    (output_dir / "categories.json").write_text(
        json.dumps(nested_json(records), indent=2, ensure_ascii=False), encoding="utf-8"
    )
    (output_dir / "scraping_errors.json").write_text(
        json.dumps([asdict(e) for e in result.errors], indent=2, ensure_ascii=False), encoding="utf-8"
    )
    pd.DataFrame(parent_rows).to_csv(output_dir / "parent_categories.csv", index=False)
    pd.DataFrame(sub_child_rows).to_csv(output_dir / "sub_child_categories.csv", index=False)
    pd.DataFrame(child_rows).to_csv(output_dir / "child_categories.csv", index=False)
    pd.DataFrame(duplicates).to_csv(output_dir / "duplicate_urls.csv", index=False)

    missing_rows = [row for row in [*parent_rows, *sub_child_rows, *all_rows] if row.get("url_missing")]
    summary = {
        "Website name": config.website,
        "Website URL": config.website_url,
        "Total Parent Categories": len({(r.parent_name, r.parent_display_order) for r in records}),
        "Total Sub-Child Categories": len({(r.parent_name, r.sub_child_name, r.sub_child_display_order) for r in records if r.sub_child_name}),
        "Total Child Categories": len([r for r in records if r.child_name]),
        "Total hierarchy records": len(records),
        "Unique URLs": len({r.normalized_url for r in records if r.normalized_url}),
        "Missing Parent URLs": len([row for row in parent_rows if row.get("url_missing")]),
        "Missing Sub-Child URLs": len([row for row in sub_child_rows if row.get("url_missing")]),
        "Missing Child URLs": len([r for r in records if r.hierarchy_level == 3 and r.url_missing]),
        "Exact duplicates": len([d for d in duplicates if d.get("duplicate_type") == "exact"]),
        "Cross-hierarchy duplicate URLs": len([d for d in duplicates if d.get("duplicate_type") == "cross_hierarchy_url"]),
        "Failed Parent Categories": len({e.parent_name for e in result.errors if e.parent_name and not e.sub_child_name}),
        "Failed Sub-Child Categories": len({(e.parent_name, e.sub_child_name) for e in result.errors if e.sub_child_name}),
        "Empty Child panels": len([e for e in result.errors if e.failed_action == "empty_child_panel"]),
        "Start time": result.start_time,
        "Completion time": result.completion_time,
        "Total runtime": f"{result.runtime_seconds:.2f}s",
        "Headless mode": headless,
        "Browser version": result.browser_version,
    }
    sheets = {
        "All Categories": pd.DataFrame(all_rows),
        "Parent Categories": pd.DataFrame(parent_rows),
        "Sub-Child Categories": pd.DataFrame(sub_child_rows),
        "Child Categories": pd.DataFrame(child_rows),
        "Duplicate URLs": pd.DataFrame(duplicates),
        "Missing URLs": pd.DataFrame(missing_rows),
        "Errors": pd.DataFrame([asdict(e) for e in result.errors]),
        "Summary": pd.DataFrame([{"metric": k, "value": v} for k, v in summary.items()]),
    }
    xlsx = output_dir / "categories.xlsx"
    with pd.ExcelWriter(xlsx, engine="openpyxl") as writer:
        for name, df in sheets.items():
            df.to_excel(writer, sheet_name=name[:31], index=False)
    format_workbook(xlsx)


def format_workbook(path: Path) -> None:
    wb = load_workbook(path)
    fill_missing = PatternFill("solid", fgColor="FFF2CC")
    fill_bad = PatternFill("solid", fgColor="F4CCCC")
    fill_dup = PatternFill("solid", fgColor="D9EAD3")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        if ws.max_row and ws.max_column:
            ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True)
        headers = {cell.value: i + 1 for i, cell in enumerate(ws[1])}
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = cell.alignment.copy(wrap_text=True)
                if isinstance(cell.value, str) and cell.value.startswith(("http://", "https://")):
                    cell.hyperlink = cell.value
                    cell.style = "Hyperlink"
            if "url_missing" in headers and row[headers["url_missing"] - 1].value is True:
                for cell in row:
                    cell.fill = fill_missing
            if "response_error" in headers and row[headers["response_error"] - 1].value:
                for cell in row:
                    cell.fill = fill_bad
            if "is_duplicate" in headers and row[headers["is_duplicate"] - 1].value is True:
                for cell in row:
                    cell.fill = fill_dup
        for col_idx in range(1, ws.max_column + 1):
            values = [str(ws.cell(row=i, column=col_idx).value or "") for i in range(1, min(ws.max_row, 200) + 1)]
            width = min(max([len(v) for v in values] + [10]) + 2, 60)
            ws.column_dimensions[get_column_letter(col_idx)].width = width
    wb.save(path)


def build_arg_parser(default_output_slug: str) -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser()
    mode = parser.add_mutually_exclusive_group()
    mode.add_argument("--headless", action="store_true", help="Run Chromium headless.")
    mode.add_argument("--visible", action="store_true", help="Run Chromium with a visible window.")
    parser.add_argument("--inspect-only", action="store_true")
    parser.add_argument("--output-dir", default="output")
    parser.add_argument("--timeout", type=int, default=60000)
    parser.add_argument("--interaction-delay", type=int, default=600)
    parser.add_argument("--scroll-delay", type=int, default=250)
    parser.add_argument("--max-retries", type=int, default=2)
    parser.add_argument("--validate-urls", action="store_true")
    parser.add_argument("--skip-url-validation", action="store_true")
    parser.add_argument("--save-parent-screenshots", action="store_true")
    parser.add_argument("--log-level", default="INFO")
    parser.add_argument("--site-output-slug", default=default_output_slug, help=argparse.SUPPRESS)
    return parser


SEMANTIC_MENU_JS = r"""
(preferredSelector) => {
  const clean = value => (value || '').replace(/\s+/g, ' ').trim();
  const badName = /^(search|cart|checkout|login|sign in|my account|account|home|shop now|view all models)$/i;
  const usableName = value => clean(value).length > 1 && !badName.test(clean(value));
  const usableUrl = anchor => {
    const raw = anchor?.getAttribute('href') || '';
    if (!raw || /^(javascript:|mailto:|tel:|#?$)/i.test(raw)) return '';
    try {
      const base = new URL(document.baseURI || location.href);
      const url = new URL(raw, base);
      return url.origin === base.origin ? url.href : '';
    } catch (_error) {
      return '';
    }
  };
  const directAnchor = element => {
    if (!element) return null;
    for (const child of element.children) {
      if (child.matches?.('a, button, [role="menuitem"]')) return child;
    }
    return element.querySelector(':scope > div > a, :scope > div > button, :scope > span > a, :scope > span > button');
  };
  const css = element => {
    if (!element) return '';
    if (element.id) return '#' + CSS.escape(element.id);
    const parts = [];
    for (let node = element, depth = 0; node?.nodeType === 1 && depth < 7; node = node.parentElement, depth += 1) {
      let part = node.tagName.toLowerCase();
      if (node.id) {
        part += '#' + CSS.escape(node.id);
        parts.unshift(part);
        break;
      }
      const classes = [...node.classList].filter(Boolean).slice(0, 3);
      if (classes.length) part += '.' + classes.map(name => CSS.escape(name)).join('.');
      parts.unshift(part);
    }
    return parts.join(' > ');
  };
  const immediateItems = list => [...(list?.children || [])].filter(child => child.matches?.('li, [role="menuitem"]'));
  const candidateRoots = [...new Set([
    ...document.querySelectorAll(preferredSelector || '#nav'),
    ...document.querySelectorAll('#nav, nav[aria-label], nav, header ul')
  ])];
  const rootScore = root => {
    const direct = immediateItems(root).length || immediateItems(root.querySelector(':scope > ul')).length;
    const links = [...root.querySelectorAll('a[href]')].filter(anchor => usableUrl(anchor)).length;
    return direct * 20 + Math.min(links, 2000);
  };
  const root = candidateRoots.sort((a, b) => rootScore(b) - rootScore(a))[0];
  if (!root) return [];
  const rootList = immediateItems(root).length ? root : root.querySelector(':scope > ul');
  const hierarchy = [];
  immediateItems(rootList).forEach((parentItem, parentIndex) => {
    const parentAnchor = directAnchor(parentItem);
    const parentName = clean(parentAnchor?.innerText || parentAnchor?.textContent);
    if (!usableName(parentName)) return;
    const parent = {
      order: parentIndex + 1,
      name: parentName,
      url: usableUrl(parentAnchor),
      selector: css(parentAnchor),
      method: 'adaptive-dom',
      sub_children: []
    };

    const groupCandidates = [...parentItem.querySelectorAll('ul')].filter(list => {
      const items = immediateItems(list);
      if (items.length < 2) return false;
      const first = directAnchor(items[0]);
      const title = clean(first?.innerText || first?.textContent);
      const validLinks = items.slice(1).map(directAnchor).filter(anchor => usableUrl(anchor)).length;
      return usableName(title) && validLinks > 0;
    });
    const deepestGroups = groupCandidates.filter(candidate =>
      !groupCandidates.some(other => other !== candidate && candidate.contains(other))
    );

    deepestGroups.forEach((group, groupIndex) => {
      const items = immediateItems(group);
      const titleAnchor = directAnchor(items[0]);
      const subName = clean(titleAnchor?.innerText || titleAnchor?.textContent);
      if (!usableName(subName)) return;
      const sub = {
        order: groupIndex + 1,
        name: subName,
        url: usableUrl(titleAnchor),
        selector: css(titleAnchor),
        method: 'adaptive-dom',
        children: []
      };
      items.slice(1).forEach(item => {
        const anchor = directAnchor(item);
        const name = clean(anchor?.innerText || anchor?.textContent);
        const url = usableUrl(anchor);
        if (!usableName(name) || !url) return;
        sub.children.push({
          order: sub.children.length + 1,
          name,
          url,
          column: 0,
          row: sub.children.length + 1,
          selector: css(anchor),
          method: 'adaptive-dom'
        });
      });
      if (sub.children.length) parent.sub_children.push(sub);
    });

    if (!parent.sub_children.length) {
      const links = [...parentItem.querySelectorAll('a[href]')];
      const children = [];
      const seen = new Set();
      links.forEach(anchor => {
        const name = clean(anchor.innerText || anchor.textContent);
        const url = usableUrl(anchor);
        if (!usableName(name) || !url || seen.has(url)) return;
        seen.add(url);
        children.push({order: children.length + 1, name, url, column: 0, row: children.length + 1, selector: css(anchor), method: 'adaptive-dom'});
      });
      if (children.length) {
        parent.sub_children.push({order: 1, name: parentName, url: parent.url, selector: css(parentAnchor), method: 'adaptive-dom', children});
      }
    }
    hierarchy.push(parent);
  });
  return hierarchy;
}
"""


def hierarchy_health(records: list[CategoryRecord]) -> dict[str, int]:
    return {
        "parents": len({(record.parent_name, record.parent_display_order) for record in records}),
        "sub_children": len({
            (record.parent_name, record.sub_child_name, record.sub_child_display_order)
            for record in records if record.sub_child_name
        }),
        "children": len([record for record in records if record.child_name]),
        "unique_urls": len({record.normalized_url or record.child_url or record.sub_child_url or record.parent_url for record in records if record.normalized_url or record.child_url or record.sub_child_url or record.parent_url}),
    }


def hierarchy_score(records: list[CategoryRecord]) -> int:
    health = hierarchy_health(records)
    return health["unique_urls"] * 5 + health["children"] * 3 + health["sub_children"] * 2 + health["parents"]


def hierarchy_needs_healing(records: list[CategoryRecord], previous_health: dict[str, Any] | None = None) -> bool:
    health = hierarchy_health(records)
    if not records or (health["parents"] >= 3 and health["children"] == 0):
        return True
    if health["parents"] and health["sub_children"] >= 3 and health["children"] == 0:
        return True
    previous_urls = int((previous_health or {}).get("unique_urls") or 0)
    return previous_urls >= 10 and health["unique_urls"] < max(3, int(previous_urls * 0.45))


async def adaptively_extract_menu(
    page: Page,
    config: SiteConfig,
    args: argparse.Namespace,
    output_dir: Path,
    logger: logging.Logger,
    extractor: Callable[[Page, SiteConfig, argparse.Namespace, Path, logging.Logger], Awaitable[list[CategoryRecord]]],
) -> list[CategoryRecord]:
    profile_path = output_dir / "healing_profile.json"
    previous_health: dict[str, Any] = {}
    try:
        previous_health = json.loads(profile_path.read_text(encoding="utf-8")).get("health", {})
    except (OSError, ValueError, TypeError, AttributeError):
        pass

    primary_records = await extractor(page, config, args, output_dir, logger)
    chosen_records = primary_records
    healing_used = False
    if hierarchy_needs_healing(primary_records, previous_health):
        logger.warning("Menu hierarchy looks incomplete; rebuilding it from the live navigation DOM")
        hierarchy = await page.evaluate(SEMANTIC_MENU_JS, config.parent_nav_selector)
        adaptive_records = records_from_hierarchy(config, hierarchy)
        if hierarchy_score(adaptive_records) > hierarchy_score(primary_records):
            chosen_records = adaptive_records
            healing_used = True
            logger.info("Adaptive menu healing recovered %s hierarchy records", len(adaptive_records))
        else:
            logger.warning("Adaptive menu healing did not improve the result; retaining site-specific output")

    health = hierarchy_health(chosen_records)
    still_unhealthy = hierarchy_needs_healing(chosen_records, previous_health)
    profile = {
        "website": config.website,
        "updated_at": utc_now(),
        "healing_used": healing_used,
        "healthy": not still_unhealthy,
        "health": health,
        "source_selectors": dict(Counter(record.source_selector for record in chosen_records if record.source_selector).most_common(20)),
    }
    target_profile = output_dir / ("healing_attempt.json" if still_unhealthy and previous_health else "healing_profile.json")
    target_profile.write_text(json.dumps(profile, indent=2, ensure_ascii=False), encoding="utf-8")
    if still_unhealthy and int(previous_health.get("unique_urls") or 0) >= 10:
        logger.error("Healing could not match the last healthy menu; keeping the previous output files")
        return []
    return chosen_records


async def run_site(
    config: SiteConfig,
    args: argparse.Namespace,
    extractor: Callable[[Page, SiteConfig, argparse.Namespace, Path, logging.Logger], Awaitable[list[CategoryRecord]]],
) -> ScrapeResult:
    output_dir = setup_output_dir(args.output_dir, config.output_slug)
    logger = setup_logger(output_dir, args.log_level)
    result = ScrapeResult(start_time=utc_now())
    start = time.monotonic()
    headless = bool(args.headless or not args.visible)

    profile_root = resolve_chrome_profile_root(Path("data") / "browser_profiles")
    profile_dir = profile_root / f"menu-map-{config.output_slug}"
    profile_dir.mkdir(parents=True, exist_ok=True)
    chrome_executable = resolve_chrome_executable()
    logger.info(
        "Starting Botasaurus with Chrome executable %s and profile %s",
        chrome_executable or "Botasaurus default discovery",
        profile_dir,
    )

    @browser(
        headless=headless,
        profile=str(profile_dir),
        window_size=(1536, 900),
        lang="en-US",
        user_agent=(
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36"
        ),
        output=None,
        raise_exception=True,
        create_error_logs=False,
        close_on_crash=True,
    )
    def _run_with_botasaurus(driver: Driver, _data):
        page = BotasaurusPage(driver)
        page.set_default_timeout(args.timeout)
        result.browser_version = driver.user_agent or "Botasaurus Chromium"

        async def _extract() -> None:
            try:
                logger.info("Opening %s with Botasaurus", config.website_url)
                await page.goto(config.website_url, wait_until="domcontentloaded", timeout=args.timeout)
                await page.wait_for_load_state("networkidle", timeout=min(args.timeout, 15000))
                await page.wait_for_timeout(args.interaction_delay)
                if "mobilesentrix.ca" in config.website_url.lower():
                    for _attempt in range(8):
                        try:
                            if await page.evaluate(MOBILESENTRIX_CANADA_POPUP_DISMISS_JS):
                                logger.info("Dismissed MobileSentrix Canada location prompt")
                                await page.wait_for_timeout(750)
                                break
                        except Exception as exc:
                            logger.warning("Could not dismiss Canada location prompt: %s", exc)
                            break
                        await page.wait_for_timeout(500)
                inspect_selectors = [
                    config.parent_nav_selector,
                    config.parent_item_selector,
                    config.mega_menu_selector,
                    config.sub_child_panel_selector,
                    config.sub_child_item_selector,
                    config.child_panel_selector,
                    config.child_link_selector,
                    config.scroll_container_selector,
                    "header",
                    "nav",
                    "[role='navigation']",
                ]
                result.inspection = await collect_dom_inspection(page, inspect_selectors)
                write_inspection(output_dir, result.inspection)
                if not result.inspection or "Just a moment" in (await page.title()):
                    raise RuntimeError("Site returned an access verification page or no menu DOM was available.")
                if not args.inspect_only:
                    result.records = await adaptively_extract_menu(page, config, args, output_dir, logger, extractor)
            except Exception as exc:
                await record_error(result.errors, page, output_dir, config.website, "site_scrape", exc)
                logger.exception("Site scrape failed")
            finally:
                if not args.inspect_only:
                    try:
                        await save_screenshot(page, output_dir, "final-menu-state")
                    except Exception:
                        pass
                close_botasaurus_driver(driver, logger)

        asyncio.run(_extract())

    try:
        await asyncio.to_thread(_run_with_botasaurus, None)
    except Exception as exc:
        result.errors.append(ScrapeError(
            website=config.website,
            failed_action="browser_start",
            error_type=type(exc).__name__,
            error_message=str(exc),
            timestamp=utc_now(),
        ))
        logger.exception(
            "Botasaurus menu-map run failed while starting or connecting to Chrome executable %s",
            chrome_executable or "Botasaurus default discovery",
        )
    result.completion_time = utc_now()
    result.runtime_seconds = time.monotonic() - start
    for r in result.records:
        r.normalized_url = normalize_url(r.child_url or r.sub_child_url or r.parent_url, config.base_url)
        r.url_missing = not bool(r.normalized_url)
        if not r.scraped_at:
            r.scraped_at = result.completion_time
    duplicates = mark_duplicates(result.records)
    if args.validate_urls and not args.skip_url_validation and result.records:
        await validate_urls(result.records)
    if not args.inspect_only:
        has_previous_output = (output_dir / "categories.json").exists()
        if result.records or not has_previous_output:
            export_outputs(config, output_dir, result, headless=headless, duplicates=duplicates)
        else:
            result.preserve_previous_output = True
            logger.error("No healthy replacement hierarchy was produced; previous Menu Map output was preserved")
    print_validation_summary(config, output_dir, result, duplicates)
    return result


def print_validation_summary(config: SiteConfig, output_dir: Path, result: ScrapeResult, duplicates: list[dict[str, Any]]) -> None:
    parents = len({(r.parent_name, r.parent_display_order) for r in result.records})
    subs = len({(r.parent_name, r.sub_child_name, r.sub_child_display_order) for r in result.records if r.sub_child_name})
    children = len([r for r in result.records if r.child_name])
    missing = len([r for r in result.records if r.url_missing])
    print(f"Website: {config.website}")
    print(f"Parents found: {parents}")
    print(f"Sub-children found: {subs}")
    print(f"Children found: {children}")
    print(f"Missing URLs: {missing}")
    print(f"Duplicate hierarchy rows: {len([d for d in duplicates if d.get('duplicate_type') == 'exact'])}")
    print(f"Errors: {len(result.errors)}")
    print(f"Output directory: {output_dir}")


def make_record(
    config: SiteConfig,
    parent: dict[str, Any],
    sub: dict[str, Any] | None = None,
    child: dict[str, Any] | None = None,
    hierarchy_level: int = 1,
    scraped_at: str | None = None,
) -> CategoryRecord:
    sub = sub or {}
    child = child or {}
    url = child.get("url") or sub.get("url") or parent.get("url") or ""
    return CategoryRecord(
        website=config.website,
        website_url=config.website_url,
        parent_name=normalize_text(parent.get("name")),
        parent_url=normalize_url(parent.get("url", ""), config.base_url),
        parent_display_order=int(parent.get("order") or 0),
        parent_open_method=parent.get("method") or config.parent_open_method,
        sub_child_name=normalize_text(sub.get("name")),
        sub_child_url=normalize_url(sub.get("url", ""), config.base_url),
        sub_child_display_order=int(sub.get("order") or 0),
        sub_child_activation_method=sub.get("method") or (config.sub_child_activation_method if sub else ""),
        child_name=normalize_text(child.get("name")),
        child_url=normalize_url(child.get("url", ""), config.base_url),
        child_display_order=int(child.get("order") or 0),
        column_number=int(child.get("column") or 0),
        row_number=int(child.get("row") or 0),
        hierarchy_level=hierarchy_level,
        discovery_method=child.get("method") or sub.get("method") or parent.get("method") or "dom-inspection",
        source_selector=child.get("selector") or sub.get("selector") or parent.get("selector") or "",
        normalized_url=normalize_url(url, config.base_url),
        url_missing=not bool(normalize_url(url, config.base_url)),
        scraped_at=scraped_at or utc_now(),
    )


def records_from_hierarchy(config: SiteConfig, hierarchy: list[dict[str, Any]]) -> list[CategoryRecord]:
    records: list[CategoryRecord] = []
    stamp = utc_now()
    for parent in hierarchy:
        subs = parent.get("sub_children") or []
        if not subs:
            records.append(make_record(config, parent, hierarchy_level=1, scraped_at=stamp))
            continue
        for sub in subs:
            children = sub.get("children") or []
            if not children:
                records.append(make_record(config, parent, sub, hierarchy_level=2, scraped_at=stamp))
                continue
            for child in children:
                records.append(make_record(config, parent, sub, child, hierarchy_level=3, scraped_at=stamp))
    return records
